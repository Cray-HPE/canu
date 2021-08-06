"""CANU switch config commands."""
from collections import defaultdict
import json
import os
from pathlib import Path
import re
import sys

import click
from click_help_colors import HelpColorsCommand
from jinja2 import Environment, FileSystemLoader, StrictUndefined
import natsort
from network_modeling.NetworkNodeFactory import NetworkNodeFactory
from openpyxl import load_workbook
import requests
import ruamel.yaml
import urllib3

from canu.validate.shcd.shcd import node_model_from_shcd

yaml = ruamel.yaml.YAML()

# To disable warnings about unsecured HTTPS requests
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


# Get project root directory
if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):  # pragma: no cover
    project_root = sys._MEIPASS
else:
    prog = __file__
    project_root = Path(__file__).resolve().parent.parent.parent.parent

# Schema and Data files
hardware_schema_file = os.path.join(
    project_root, "network_modeling", "schema", "cray-network-hardware-schema.yaml"
)
hardware_spec_file = os.path.join(
    project_root, "network_modeling", "models", "cray-network-hardware.yaml"
)
architecture_schema_file = os.path.join(
    project_root, "network_modeling", "schema", "cray-network-architecture-schema.yaml"
)
architecture_spec_file = os.path.join(
    project_root, "network_modeling", "models", "cray-network-architecture.yaml"
)

canu_cache_file = os.path.join(project_root, "canu", "canu_cache.yaml")
canu_config_file = os.path.join(project_root, "canu", "canu.yaml")

# Import templates
network_templates_folder = os.path.join(
    project_root, "network_modeling", "configs", "templates"
)
env = Environment(
    loader=FileSystemLoader(network_templates_folder),
    # trim_blocks=True,
    undefined=StrictUndefined,
)

# Get Shasta versions from canu.yaml
with open(canu_config_file, "r") as file:
    canu_config = yaml.load(file)

shasta_options = canu_config["shasta_versions"]


@click.command(
    cls=HelpColorsCommand,
    help_headers_color="yellow",
    help_options_color="blue",
)
@click.option(
    "--shasta",
    "-s",
    type=click.Choice(shasta_options),
    help="Shasta network version",
    prompt="Shasta network version",
    required=True,
    show_choices=True,
)
@click.option(
    "--architecture",
    "-a",
    type=click.Choice(["Full", "TDS", "V1"], case_sensitive=False),
    help="Shasta architecture",
    required=True,
    prompt="Architecture type",
)
@click.option(
    "--shcd",
    help="SHCD file",
    type=click.File("rb"),
    required=True,
)
@click.option(
    "--tabs",
    help="The tabs on the SHCD file to check, e.g. 10G_25G_40G_100G,NMN,HMN.",
)
@click.option(
    "--corners",
    help="The corners on each tab, comma separated e.g. 'J37,U227,J15,T47,J20,U167'.",
)
@click.option(
    "--name",
    "switch_name",
    required=True,
    help="The name of the switch to generate config e.g. 'sw-spine-001'",
    prompt="Switch Name",
)
@click.option("--csi-folder", help="Directory containing the CSI json file")
@click.option(
    "--auth-token",
    envvar="SLS_TOKEN",
    help="Token for SLS authentication",
)
@click.option("--sls-address", default="api-gw-service-nmn.local", show_default=True)
@click.option(
    "--out", help="Output results to a file", type=click.File("w"), default="-"
)
@click.pass_context
def config(
    ctx,
    shasta,
    architecture,
    shcd,
    tabs,
    corners,
    switch_name,
    csi_folder,
    auth_token,
    sls_address,
    out,
):
    """Switch config commands.

    In order to generate switch config, a valid SHCD must be passed in and system variables must be read in from either
    CSI output or the SLS API.

    To generate config for a specific switch, a hostname must also be passed in using the `--name HOSTNAME` flag.
    To output the config to a file, append the `--out FILENAME` flag.

    \f
    # noqa: D301

    Args:
        ctx: CANU context settings
        shasta: Shasta version
        architecture: Shasta architecture
        shcd: SHCD file
        tabs: The tabs on the SHCD file to check, e.g. 10G_25G_40G_100G,NMN,HMN.
        corners: The corners on each tab, comma separated e.g. 'J37,U227,J15,T47,J20,U167'.
        switch_name: Switch name
        csi_folder: Directory containing the CSI json file
        auth_token: Token for SLS authentication
        sls_address: The address of SLS
        out: Name of the output file
    """
    if architecture.lower() == "full":
        architecture = "network_v2"
        template_folder = "full"
    elif architecture.lower() == "tds":
        architecture = "network_v2_tds"
    elif architecture.lower() == "v1":
        architecture = "network_v1"

    # SHCD Parsing
    sheets = []

    if not tabs:
        wb = load_workbook(shcd, read_only=True)
        click.secho("What tabs would you like to check in the SHCD?")
        tab_options = wb.sheetnames
        for x in tab_options:
            click.secho(f"{x}", fg="green")

        tabs = click.prompt(
            "Please enter the tabs to check separated by a comma, e.g. 10G_25G_40G_100G,NMN,HMN.",
            type=str,
        )

    if corners:
        if len(tabs.split(",")) * 2 != len(corners.split(",")):
            click.secho("Not enough corners.\n", fg="red")
            click.secho(
                f"Make sure each tab: {tabs.split(',')} has 2 corners.\n", fg="red"
            )
            click.secho(
                f"There were {len(corners.split(','))} corners entered, but there should be {len(tabs.split(',')) * 2}.",
                fg="red",
            )
            click.secho(
                f"{corners}\n",
                fg="red",
            )
            return

        # Each tab should have 2 corners entered in comma separated
        for i in range(len(tabs.split(","))):
            # 0 -> 0,1
            # 1 -> 2,3
            # 2 -> 4,5

            sheets.append(
                (
                    tabs.split(",")[i],
                    corners.split(",")[i * 2].strip(),
                    corners.split(",")[i * 2 + 1].strip(),
                )
            )
    else:
        for tab in tabs.split(","):
            click.secho(f"\nFor the Sheet {tab}", fg="green")
            range_start = click.prompt(
                "Enter the cell of the upper left corner (Labeled 'Source')",
                type=str,
            )
            range_end = click.prompt(
                "Enter the cell of the lower right corner", type=str
            )
            sheets.append((tab, range_start, range_end))

    # Create Node factory
    factory = NetworkNodeFactory(
        hardware_schema=hardware_schema_file,
        hardware_data=hardware_spec_file,
        architecture_schema=architecture_schema_file,
        architecture_data=architecture_spec_file,
        architecture_version=architecture,
    )

    # Get nodes from SHCD
    shcd_node_list, shcd_warnings = node_model_from_shcd(
        factory=factory, spreadsheet=shcd, sheets=sheets
    )

    # Parse sls_input_file.json file from CSI
    if csi_folder:
        try:
            with open(os.path.join(csi_folder, "sls_input_file.json"), "r") as f:
                input_json = json.load(f)

                # Format the input to be like the SLS JSON
                sls_json = [
                    network[x]
                    for network in [input_json.get("Networks", {})]
                    for x in network
                ]

        except FileNotFoundError:
            click.secho(
                "The file sls_input_file.json was not found, check that this is the correct CSI directory.",
                fg="red",
            )
            return
    else:
        # Get SLS config
        token = os.environ.get("SLS_TOKEN")

        # Token file takes precedence over the environmental variable
        if auth_token != token:
            try:
                with open(auth_token) as f:
                    data = json.load(f)
                    token = data["access_token"]

            except Exception:
                click.secho(
                    "Invalid token file, generate another token or try again.",
                    fg="white",
                    bg="red",
                )
                return

        # SLS
        url = "https://" + sls_address + "/apis/sls/v1/networks"
        try:
            response = requests.get(
                url,
                headers={
                    "Content-Type": "application/json",
                    "Authorization": f"Bearer {token}",
                },
                verify=False,
            )
            response.raise_for_status()

            sls_json = response.json()

        except requests.exceptions.ConnectionError:
            return click.secho(
                f"Error connecting to SLS {sls_address}, check the address or pass in a new address using --sls-address.",
                fg="white",
                bg="red",
            )
        except requests.exceptions.HTTPError:
            bad_token_reason = (
                "environmental variable 'SLS_TOKEN' is correct."
                if auth_token == token
                else "token is valid, or generate a new one."
            )
            return click.secho(
                f"Error connecting SLS {sls_address}, check that the {bad_token_reason}",
                fg="white",
                bg="red",
            )
    sls_variables = parse_sls_for_config(sls_json)

    # For versions of Shasta < 1.6, the SLS Hostnames need to be renamed
    if shasta:
        if float(shasta) < 1.6:
            sls_variables = rename_sls_hostnames(sls_variables)

    switch_config, devices = generate_switch_config(
        shcd_node_list, factory, switch_name, sls_variables, template_folder
    )

    dash = "-" * 60
    click.echo("\n")
    click.secho(f"{switch_name} Config", fg="bright_white")
    click.echo(dash)

    click.echo(switch_config, file=out)
    return


def get_shasta_name(name, mapper):
    """Parse mapper to get Shasta name."""
    for node in mapper:
        shasta_name = node[1]
        if shasta_name in name:
            return shasta_name


def generate_switch_config(
    shcd_node_list, factory, switch_name, sls_variables, template_folder
):
    """Generate switch config.

    Args:
        shcd_node_list: List of nodes from the SHCD
        factory: Node factory object
        switch_name: Switch hostname
        sls_variables: Dictionary containing SLS variables
        template_folder: Architecture folder contaning the switch templates

    Returns:
        switch_config: The generated switch configuration
    """
    node_shasta_name = get_shasta_name(switch_name, factory.lookup_mapper())

    if node_shasta_name is None:
        return Exception(
            click.secho(
                f"For switch {switch_name}, the type cannot be determined. Please check the switch name and try again.",
                fg="red",
            )
        )
    elif node_shasta_name not in ["sw-cdu", "sw-leaf-bmc", "sw-leaf", "sw-spine"]:
        return Exception(
            click.secho(
                f"{switch_name} is not a switch. Only switch config can be generated.",
                fg="red",
            )
        )

    is_primary, primary, secondary = switch_is_primary(switch_name)

    templates = {
        "sw-spine": {
            "primary": f"{template_folder}/sw-spine.primary.j2",
            "secondary": f"{template_folder}/sw-spine.secondary.j2",
        },
        "sw-cdu": {
            "primary": "common/sw-cdu.primary.j2",
            "secondary": "common/sw-cdu.secondary.j2",
        },
        "sw-leaf": {
            "primary": f"{template_folder}/sw-leaf.primary.j2",
            "secondary": f"{template_folder}/sw-leaf.secondary.j2",
        },
        "sw-leaf-bmc": {
            "primary": f"{template_folder}/sw-leaf-bmc.j2",
            "secondary": f"{template_folder}/sw-leaf-bmc.j2",
        },
    }
    template_name = templates[node_shasta_name][
        "primary" if is_primary else "secondary"
    ]
    template = env.get_template(template_name)
    variables = {
        "HOSTNAME": switch_name,
        "NCN_W001": sls_variables["ncn_w001"],
        "NCN_W002": sls_variables["ncn_w002"],
        "NCN_W003": sls_variables["ncn_w003"],
        "NMN": sls_variables["NMN"],
        "HMN": sls_variables["HMN"],
        "HMN_MTN": sls_variables["HMN_MTN"],
        "NMN_MTN": sls_variables["NMN_MTN"],
        "HMN_IP_GATEWAY": sls_variables["HMN_IP_GATEWAY"],
        "MTL_IP_GATEWAY": sls_variables["MTL_IP_GATEWAY"],
        "NMN_IP_GATEWAY": sls_variables["NMN_IP_GATEWAY"],
        "CAN_IP_GATEWAY": sls_variables["CAN_IP_GATEWAY"],
        "CAN_IP_PRIMARY": sls_variables["CAN_IP_PRIMARY"],
        "CAN_IP_SECONDARY": sls_variables["CAN_IP_SECONDARY"],
        "NMN_MTN_CABINETS": sls_variables["NMN_MTN_CABINETS"],
        "HMN_MTN_CABINETS": sls_variables["HMN_MTN_CABINETS"],
    }

    cabling = {}
    cabling["nodes"] = get_switch_nodes(switch_name, shcd_node_list, factory)

    if switch_name not in sls_variables["HMN_IPs"].keys():
        click.secho(f"Cannot find {switch_name} in CSI / SLS nodes.", fg="red")
        exit(1)

    variables["HMN_IP"] = sls_variables["HMN_IPs"][switch_name]
    variables["MTL_IP"] = sls_variables["MTL_IPs"][switch_name]
    variables["NMN_IP"] = sls_variables["NMN_IPs"][switch_name]

    last_octet = variables["HMN_IP"].split(".")[3]
    variables["LOOPBACK_IP"] = "10.2.0." + last_octet + "/32"
    variables["IPV6_IP"] = "2001:db8:beef:99::" + last_octet + "/128"

    if node_shasta_name in ["sw-spine", "sw-leaf", "sw-cdu"]:
        # Get connections to switch pair
        pair_connections = get_pair_connections(cabling["nodes"], switch_name)
        variables["VSX_KEEPALIVE"] = pair_connections[0]
        variables["VSX_ISL_PORT1"] = pair_connections[1]
        variables["VSX_ISL_PORT2"] = pair_connections[2]

    switch_config = template.render(
        variables=variables,
        cabling=cabling,
    )
    devices = set()
    for node in cabling["nodes"]:
        devices.add(node["subtype"])

    return switch_config, devices


def get_pair_connections(nodes, switch_name):
    """Given a hostname and nodes, return connections to the primary or secondary switch.

    Args:
        nodes: List of nodes connected to the switch
        switch_name: Switch hostname

    Returns:
        List of connections to the paired switch
    """
    is_primary, primary, secondary = switch_is_primary(switch_name)

    if is_primary:
        pair_hostname = secondary
    else:
        pair_hostname = primary

    connections = []
    for x in nodes:
        if pair_hostname in x["config"]["DESCRIPTION"]:
            connections.append(x["config"]["PORT"])

    connections = natsort.natsorted(connections)
    return connections


def get_switch_nodes(switch_name, shcd_node_list, factory):
    """Get the nodes connected to the switch ports.

    Args:
        switch_name: Switch hostname
        shcd_node_list: List of nodes from the SHCD
        factory: Node factory object

    Returns:
        List of nodes connected to the switch
    """
    nodes = []
    nodes_by_name = {}
    nodes_by_id = {}

    # Make 2 dictionaries for easy node lookup
    for node in shcd_node_list:
        node_tmp = node.serialize()
        name = node_tmp["common_name"]

        nodes_by_name[name] = node_tmp
        nodes_by_id[node_tmp["id"]] = node_tmp

    if switch_name not in nodes_by_name.keys():
        click.secho(
            f"For switch {switch_name}, the type cannot be determined. Please check the switch name and try again.",
            fg="red",
        )
        exit(1)

    for port in nodes_by_name[switch_name]["ports"]:
        destination_node_id = port["destination_node_id"]
        destination_node_name = nodes_by_id[port["destination_node_id"]]["common_name"]
        source_port = port["port"]
        destination_port = port["destination_port"]
        destination_slot = port["destination_slot"]

        shasta_name = get_shasta_name(destination_node_name, factory.lookup_mapper())

        primary_port = get_primary_port(nodes_by_name, switch_name, destination_node_id)
        if shasta_name == "ncn-m":
            new_node = {
                "subtype": "master",
                "slot": destination_slot,
                "destination_port": destination_port,
                "config": {
                    "DESCRIPTION": f"{switch_name}:{source_port}==>{destination_node_name}:{destination_slot}:{destination_port}",
                    "PORT": f"1/1/{source_port}",
                    "LAG_NUMBER": primary_port,
                },
            }
            nodes.append(new_node)
        elif shasta_name == "ncn-s":
            # ncn-s also needs destination_port to find the match
            primary_port_ncn_s = get_primary_port(
                nodes_by_name, switch_name, destination_node_id, destination_port
            )
            new_node = {
                "subtype": "storage",
                "slot": destination_slot,
                "destination_port": destination_port,
                "config": {
                    "DESCRIPTION": f"{switch_name}:{source_port}==>{destination_node_name}:{destination_slot}:{destination_port}",
                    "PORT": f"1/1/{source_port}",
                    "LAG_NUMBER": primary_port_ncn_s,
                },
            }
            nodes.append(new_node)
        elif shasta_name == "ncn-w":
            new_node = {
                "subtype": "worker",
                "slot": destination_slot,
                "destination_port": destination_port,
                "config": {
                    "DESCRIPTION": f"{switch_name}:{source_port}==>{destination_node_name}:{destination_slot}:{destination_port}",
                    "PORT": f"1/1/{source_port}",
                    "LAG_NUMBER": primary_port,
                },
            }
            nodes.append(new_node)
        elif shasta_name == "cec":
            new_node = {
                "subtype": "cec",
                "slot": None,
                "config": {
                    "DESCRIPTION": f"{switch_name}:{source_port}==>{destination_node_name}:{destination_port}",
                    "INTERFACE_NUMBER": f"1/1/{source_port}",
                },
            }
            nodes.append(new_node)
        elif shasta_name == "cmm":
            new_node = {
                "subtype": "cmm",
                "slot": None,
                "config": {
                    "DESCRIPTION": f"{switch_name}:{source_port}==>{destination_node_name}:{destination_port}",
                    "PORT": f"1/1/{source_port}",
                    "LAG_NUMBER": primary_port,
                },
            }
            nodes.append(new_node)
        elif shasta_name == "uan":
            new_node = {
                "subtype": "uan",
                "slot": destination_slot,
                "destination_port": destination_port,
                "config": {
                    "DESCRIPTION": f"{switch_name}:{source_port}==>{destination_node_name}:{destination_slot}:{destination_port}",
                    "PORT": f"1/1/{source_port}",
                    "LAG_NUMBER": primary_port,
                },
            }
            nodes.append(new_node)
        elif shasta_name == "sw-spine":
            # sw-leaf ==> sw-spine
            if switch_name.startswith("sw-leaf"):
                is_primary, primary, secondary = switch_is_primary(switch_name)
                digits = re.findall(r"(\d+)", primary)[0]
                lag_number = 100 + int(digits)

            # sw-cdu ==> sw-spine
            elif switch_name.startswith("sw-cdu"):
                lag_number = 255

            # sw-leaf-bmc ==> sw-spine
            elif switch_name.startswith("sw-leaf-bmc"):
                lag_number = 255

            # sw-spine ==> sw-spine
            elif switch_name.startswith("sw-spine"):
                lag_number = 256
            new_node = {
                "subtype": "spine",
                "slot": None,
                "config": {
                    "DESCRIPTION": f"{switch_name}:{source_port}==>{destination_node_name}:{destination_port}",
                    "LAG_NUMBER": lag_number,
                    "PORT": f"1/1/{source_port}",
                },
            }
            nodes.append(new_node)
        elif shasta_name == "sw-cdu":
            is_primary, primary, secondary = switch_is_primary(destination_node_name)
            # sw-spine ==> sw-cdu
            if switch_name.startswith("sw-spine"):
                digits = re.findall(r"(\d+)", primary)[0]
                lag_number = 200 + int(digits)

            # sw-cdu ==> sw-cdu
            elif switch_name.startswith("sw-cdu"):
                lag_number = 256
            new_node = {
                "subtype": "cdu",
                "slot": None,
                "primary": is_primary,
                "config": {
                    "DESCRIPTION": f"{switch_name}:{source_port}==>{destination_node_name}:{destination_port}",
                    "LAG_NUMBER": lag_number,
                    "PORT": f"1/1/{source_port}",
                },
            }
            nodes.append(new_node)
        elif shasta_name == "sw-leaf":
            # sw-spine ==> sw-leaf
            is_primary, primary, secondary = switch_is_primary(destination_node_name)
            if switch_name.startswith("sw-spine"):
                digits = re.findall(r"(\d+)", primary)[0]
                lag_number = 100 + int(digits)

            # sw-leaf-bmc ==> sw-leaf
            elif switch_name.startswith("sw-leaf-bmc"):
                lag_number = 255

            # sw-leaf ==> sw-leaf
            elif switch_name.startswith("sw-leaf"):
                lag_number = 256
            new_node = {
                "subtype": "leaf",
                "slot": None,
                "primary": is_primary,
                "config": {
                    "DESCRIPTION": f"{switch_name}:{source_port}==>{destination_node_name}:{destination_port}",
                    "LAG_NUMBER": lag_number,
                    "PORT": f"1/1/{source_port}",
                },
            }
            nodes.append(new_node)
        elif shasta_name == "sw-leaf-bmc":
            # sw-leaf ==> sw-leaf-bmc
            if switch_name.startswith("sw-leaf"):
                digits = re.findall(r"(\d+)", destination_node_name)[0]
                lag_number = 150 + int(digits)

            # sw-spine ==> sw-leaf-bmc
            elif switch_name.startswith("sw-spine"):
                digits = re.findall(r"(\d+)", destination_node_name)[0]
                lag_number = 150 + int(digits)
            new_node = {
                "subtype": "leaf-bmc",
                "slot": None,
                "config": {
                    "DESCRIPTION": f"{switch_name}:{source_port}==>{destination_node_name}:{destination_port}",
                    "LAG_NUMBER": lag_number,
                    "PORT": f"1/1/{source_port}",
                },
            }
            nodes.append(new_node)
        elif shasta_name == "sw-edge":
            pass
        else:  # pragma: no cover
            print("*********************************")
            print("Cannot determine destination connection")
            print("Source: ", switch_name)
            print("Port: ", port)
            print("Destination: ", destination_node_name)
            print("shasta_name", shasta_name)
            print("*********************************")
            new_node = {
                "subtype": "unknown",
                "slot": None,
                "config": {
                    "DESCRIPTION": f"{switch_name}:{source_port}==>{destination_node_name}:{destination_port}",
                },
            }
            nodes.append(new_node)
    return nodes


def switch_is_primary(switch):
    """Determine if the switch is primary or secondary.

    Args:
        switch: Switch hostname

    Returns:
        is_primary: Bool if the switch is primary
        primary: Primary switch hostname
        secondary: Secondary switch hostname
    """
    # Determine if PRIMARY or SECONDARY
    digits = re.findall(r"(\d+)", switch)[0]
    middle = re.findall(r"(?:sw-)([a-z-]+)", switch)[0]

    if int(digits) % 2 == 0:  # Switch is Secondary
        is_primary = False
        primary = f"sw-{middle.rstrip('-')}-{int(digits)-1 :03d}"
        secondary = switch
    else:  # Switch is Primary
        is_primary = True
        secondary = f"sw-{middle.rstrip('-')}-{int(digits)+1 :03d}"
        primary = switch

    return is_primary, primary, secondary


def parse_sls_for_config(input_json):
    """Parse the `sls_input_file.json` file or the JSON from SLS `/networks` API for config variables.

    Args:
        input_json: JSON from the SLS `/networks` API

    Returns:
        sls_variables: Dictionary containing SLS variables.
    """
    networks_list = []

    sls_variables = {
        "CAN": None,
        "HMN": None,
        "MTL": None,
        "NMN": None,
        "HMN_MTN": None,
        "NMN_MTN": None,
        "CAN_IP_GATEWAY": None,
        "HMN_IP_GATEWAY": None,
        "MTL_IP_GATEWAY": None,
        "NMN_IP_GATEWAY": None,
        "ncn_w001": None,
        "ncn_w002": None,
        "ncn_w003": None,
        "CAN_IP_PRIMARY": None,
        "CAN_IP_SECONDARY": None,
        "HMN_IPs": defaultdict(),
        "MTL_IPs": defaultdict(),
        "NMN_IPs": defaultdict(),
        "NMN_MTN_CABINETS": [],
        "HMN_MTN_CABINETS": [],
    }

    for sls_network in input_json:
        name = sls_network.get("Name", "")

        if name == "CAN":
            sls_variables["CAN"] = sls_network.get("ExtraProperties", {}).get(
                "CIDR", ""
            )
            for subnets in sls_network.get("ExtraProperties", {}).get("Subnets", {}):
                if subnets["Name"] == "bootstrap_dhcp":
                    sls_variables["CAN_IP_GATEWAY"] = subnets["Gateway"]
                    for ip in subnets["IPReservations"]:
                        if ip["Name"] == "can-switch-1":
                            sls_variables["CAN_IP_PRIMARY"] = ip["IPAddress"]
                        elif ip["Name"] == "can-switch-2":
                            sls_variables["CAN_IP_SECONDARY"] = ip["IPAddress"]

        elif name == "HMN":
            sls_variables["HMN"] = sls_network.get("ExtraProperties", {}).get(
                "CIDR", ""
            )
            for subnets in sls_network.get("ExtraProperties", {}).get("Subnets", {}):
                if subnets["Name"] == "network_hardware":
                    sls_variables["HMN_IP_GATEWAY"] = subnets["Gateway"]
                    for ip in subnets["IPReservations"]:
                        sls_variables["HMN_IPs"][ip["Name"]] = ip["IPAddress"]

        elif name == "MTL":
            sls_variables["MTL"] = sls_network.get("ExtraProperties", {}).get(
                "CIDR", ""
            )
            for subnets in sls_network.get("ExtraProperties", {}).get("Subnets", {}):
                if subnets["Name"] == "network_hardware":
                    sls_variables["MTL_IP_GATEWAY"] = subnets["Gateway"]
                    for ip in subnets["IPReservations"]:
                        sls_variables["MTL_IPs"][ip["Name"]] = ip["IPAddress"]

        elif name == "NMN":
            sls_variables["NMN"] = sls_network.get("ExtraProperties", {}).get(
                "CIDR", ""
            )
            for subnets in sls_network.get("ExtraProperties", {}).get("Subnets", {}):
                if subnets["Name"] == "bootstrap_dhcp":
                    for ip in subnets["IPReservations"]:
                        if ip["Name"] == "ncn-w001":
                            sls_variables["ncn_w001"] = ip["IPAddress"]
                        elif ip["Name"] == "ncn-w002":
                            sls_variables["ncn_w002"] = ip["IPAddress"]
                        elif ip["Name"] == "ncn-w003":
                            sls_variables["ncn_w003"] = ip["IPAddress"]
                elif subnets["Name"] == "network_hardware":
                    sls_variables["NMN_IP_GATEWAY"] = subnets["Gateway"]
                    for ip in subnets["IPReservations"]:
                        sls_variables["NMN_IPs"][ip["Name"]] = ip["IPAddress"]

        elif name == "NMN_MTN":
            sls_variables["NMN_MTN"] = sls_network.get("ExtraProperties", {}).get(
                "CIDR", ""
            )
            sls_variables["NMN_MTN_CABINETS"] = [
                subnet
                for subnet in sls_network.get("ExtraProperties", {}).get("Subnets", {})
            ]
        elif name == "HMN_MTN":
            sls_variables["HMN_MTN"] = sls_network.get("ExtraProperties", {}).get(
                "CIDR", ""
            )
            sls_variables["HMN_MTN_CABINETS"] = [
                subnet
                for subnet in sls_network.get("ExtraProperties", {}).get("Subnets", {})
            ]

        for subnets in sls_network.get("ExtraProperties", {}).get("Subnets", {}):

            vlan = subnets.get("VlanID", "")
            networks_list.append([name, vlan])

    networks_list = set(tuple(x) for x in networks_list)

    return sls_variables


def rename_sls_hostnames(sls_variables):
    """Parse and rename SLS switch names.

    The operation needs to be done in two passes to prevent naming conflicts.

    Args:
        sls_variables: Dictionary containing SLS variables.

    Returns:
        sls_variables: Dictionary containing renamed SLS variables.
    """
    # First pass rename leaf ==> leaf-bmc
    for key, value in sls_variables["HMN_IPs"].copy().items():
        new_name = key.replace("-leaf-", "-leaf-bmc-")
        sls_variables["HMN_IPs"].pop(key)
        sls_variables["HMN_IPs"][new_name] = value

    for key, value in sls_variables["MTL_IPs"].copy().items():
        new_name = key.replace("-leaf-", "-leaf-bmc-")
        sls_variables["MTL_IPs"].pop(key)
        sls_variables["MTL_IPs"][new_name] = value

    for key, value in sls_variables["NMN_IPs"].copy().items():
        new_name = key.replace("-leaf-", "-leaf-bmc-")
        sls_variables["NMN_IPs"].pop(key)
        sls_variables["NMN_IPs"][new_name] = value

    # Second pass rename agg ==> leaf
    for key, value in sls_variables["HMN_IPs"].copy().items():
        new_name = key.replace("-agg-", "-leaf-")
        sls_variables["HMN_IPs"].pop(key)
        sls_variables["HMN_IPs"][new_name] = value

    for key, value in sls_variables["MTL_IPs"].copy().items():
        new_name = key.replace("-agg-", "-leaf-")
        sls_variables["MTL_IPs"].pop(key)
        sls_variables["MTL_IPs"][new_name] = value

    for key, value in sls_variables["NMN_IPs"].copy().items():
        new_name = key.replace("-agg-", "-leaf-")
        sls_variables["NMN_IPs"].pop(key)
        sls_variables["NMN_IPs"][new_name] = value

    return sls_variables


def get_primary_port(
    nodes_by_name, switch_name, destination_node_id, destination_port=None
):
    """Return the primary switch port number for a connection to a node.

    Args:
        nodes_by_name: Dictionary containing the node list where hostname is the key
        switch_name: Switch hostname
        destination_node_id: The node id of the destination device.
        destination_port: (Optional, only used with ncn-s) The destination port

    Returns:
        port: Port number of the primary connection to a device.
    """
    is_primary, primary, secondary = switch_is_primary(switch_name)

    for y in nodes_by_name[primary]["ports"]:
        if y["destination_node_id"] == destination_node_id:
            if not destination_port:
                return y["port"]
            # Since ncn-s can have multiple connections to a device, returns the correct one
            elif destination_port and y["destination_port"] == destination_port:
                return y["port"]
