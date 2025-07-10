#
# MIT License
#
# (C) Copyright 2022-2025 Hewlett Packard Enterprise Development LP
#
# Permission is hereby granted, free of charge, to any person obtaining a
# copy of this software and associated documentation files (the "Software"),
# to deal in the Software without restriction, including without limitation
# the rights to use, copy, modify, merge, publish, distribute, sublicense,
# and/or sell copies of the Software, and to permit persons to whom the
# Software is furnished to do so, subject to the following conditions:
#
# The above copyright notice and this permission notice shall be included
# in all copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL
# THE AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR
# OTHER LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE,
# ARISING FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR
# OTHER DEALINGS IN THE SOFTWARE.
#
"""CANU commands that validate the shcd."""
import datetime
import json
import logging
import re
import sys
from collections import defaultdict
from importlib import metadata
from os import path
from pathlib import Path

import click
import natsort
from openpyxl import load_workbook

from canu.style import Style
from network_modeling.NetworkNodeFactory import NetworkNodeFactory
from network_modeling.NetworkPort import NetworkPort
from network_modeling.NodeLocation import NodeLocation

# Get project root directory
if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):  # pragma: no cover
    project_root = sys._MEIPASS
else:
    prog = __file__
    project_root = Path(__file__).resolve().parent.parent.parent.parent

version = metadata.version("canu")

log = logging.getLogger("validate_shcd")


@click.command(
    cls=Style.CanuHelpColorsCommand,
)
@click.option(
    "--architecture",
    "-a",
    type=click.Choice(["Full", "TDS", "V1"], case_sensitive=False),
    help="CSM architecture",
    required=True,
    prompt="Architecture type",
)
@click.option(
    "--shcd",
    help="SHCD file",
    type=click.File("rb"),
    required=True,
)
@click.option(
    "--tabs",
    help="The tabs on the SHCD file to check, e.g. 10G_25G_40G_100G,NMN,HMN.",
)
@click.option(
    "--corners",
    help="The corners on each tab, comma separated e.g. 'J37,U227,J15,T47,J20,U167'.",
)
@click.option(
    "--edge",
    type=click.Choice(["Aruba", "Arista"], case_sensitive=False),
    help="Vendor of Edge router",
    required=True,
    default="Arista",
)
@click.option(
    "--out",
    help="Output results to a file",
    type=click.File("w"),
    default="-",
)
@click.option("--json", "json_", is_flag=True, help="Output JSON model to a file")
@click.option(
    "--log",
    "log_",
    help="Level of logging.",
    type=click.Choice(["DEBUG", "INFO", "WARNING", "ERROR"]),
    default="ERROR",
)
@click.pass_context
def shcd(ctx, architecture, shcd, tabs, corners, edge, out, json_, log_):
    """Validate a SHCD file.

    CANU can be used to validate that an SHCD (SHasta Cabling Diagram) passes basic validation checks.

    - Use the '--tabs' flag to select which tabs on the spreadsheet will be included.

    - The '--corners' flag is used to input the upper left and lower right corners of the table on each tab of the worksheet. If the corners are not specified, you will be prompted to enter them for each tab.

    - The table should contain the 11 headers: Source, Rack, Location, Slot, (Blank), Port, Destination, Rack, Location, (Blank), Port.

    --------
    \f
    # noqa: D301, B950

    Args:
        ctx: CANU context settings
        architecture: CSM architecture
        shcd: SHCD file
        tabs: The tabs on the SHCD file to check, e.g. 10G_25G_40G_100G,NMN,HMN.
        corners: The corners on each tab, comma separated e.g. 'J37,U227,J15,T47,J20,U167'.
        edge: Vendor of the edge router
        out: Filename for the JSON Topology if requested.
        json_: Bool indicating json output
        log_: Level of logging.
    """
    logging.basicConfig(format="%(name)s - %(levelname)s: %(message)s", level=log_)

    # This should really be a lookup table in cray-network-architecture.yaml
    if architecture.lower() == "full":
        architecture = "network_v2"
    elif architecture.lower() == "tds":
        architecture = "network_v2_tds"
    elif architecture.lower() == "v1":
        architecture = "network_v1"

    # SHCD Parsing
    try:
        sheets = shcd_to_sheets(shcd, tabs, corners)
    except Exception as err:
        click.secho(err, fg="red")
        return

    # Create Node factory
    factory = NetworkNodeFactory(architecture_version=architecture)

    node_list, warnings = node_model_from_shcd(
        factory=factory,
        spreadsheet=shcd,
        sheets=sheets,
        edge=edge.lower(),
    )

    if json_:
        json_output(node_list, factory, architecture, ctx, out)
    else:
        print_node_list(node_list, "SHCD", out)

        node_list_warnings(node_list, warnings, out)

        switch_unused_ports(node_list)


def shcd_to_sheets(shcd, tabs, corners):
    """Parse SHCD tabs and corners into sheets.

    Args:
        shcd: SHCD file
        tabs: The tabs on the SHCD file to check, e.g. 10G_25G_40G_100G,NMN,HMN.
        corners: The corners on each tab, comma separated e.g. 'J37,U227,J15,T47,J20,U167'.

    Returns:
        sheets

    Raises:
        Exception: If there are not pairs of corners
    """
    sheets = []

    if not tabs:
        wb = load_workbook(shcd, read_only=True, data_only=True)
        click.secho("What tabs would you like to check in the SHCD?")
        tab_options = wb.sheetnames
        for x in tab_options:
            click.secho(f"{x}", fg="green")

        tabs = click.prompt(
            "Please enter the tabs to check separated by a comma, e.g. 10G_25G_40G_100G,NMN,HMN.",
            type=str,
        )

    if corners:
        if len(tabs.split(",")) * 2 != len(corners.split(",")):
            click.secho("Not enough corners.\n", fg="red")
            click.secho(
                f"Make sure each tab: {tabs.split(',')} has 2 corners.\n",
                fg="red",
            )
            click.secho(
                f"There were {len(corners.split(','))} corners entered, but there should be {len(tabs.split(',')) * 2}.",
                fg="red",
            )
            click.secho(
                f"{corners}\n",
                fg="red",
            )
            raise Exception

        # Each tab should have 2 corners entered in comma separated
        for i in range(len(tabs.split(","))):
            # 0 -> 0,1
            # 1 -> 2,3
            # 2 -> 4,5

            sheets.append(
                (
                    tabs.split(",")[i],
                    corners.split(",")[i * 2].strip(),
                    corners.split(",")[i * 2 + 1].strip(),
                ),
            )
    else:
        for tab in tabs.split(","):
            click.secho(f"\nFor the Sheet {tab}", fg="green")
            range_start = click.prompt(
                "Enter the cell of the upper left corner (Labeled 'Source')",
                type=str,
            )
            range_end = click.prompt(
                "Enter the cell of the lower right corner",
                type=str,
            )
            sheets.append((tab, range_start, range_end))

    return sheets


def get_node_common_name(name, rack_number, rack_elevation, mapper):
    """Map SHCD device names to hostname.

    Args:
        name: A string from SHCD representing the device name
        rack_number: A string for rack the device is in (e.g. x1001)
        rack_elevation: A string for the position of the device in the rack (e.g. u19)
        mapper: An array of tuples (SHCD name, hostname, device type)

    Returns:
        common_name: A string of the hostname

    Raises:
        ValueError: When name cannot be appropriately found
    """
    common_name = None
    if name is not None:
        name = name.lower()
    for node in mapper:
        for lookup_name in node[0]:
            if re.match("^{}".format(lookup_name.strip()), name):
                # One naming convention for switches, another for else.
                tmp_name = None
                if node[1].find("sw-hsn") != -1:
                    tmp_name = node[1] + "-" + rack_number + "-"
                elif node[1].find("sw-") != -1:
                    tmp_name = node[1] + "-"
                elif node[1].find("cmm") != -1:
                    tmp_name = node[1] + "-" + rack_number + "-"
                elif node[1].find("cec") != -1:
                    tmp_name = node[1] + "-" + rack_number + "-"
                elif node[1].find("pdu") != -1:
                    tmp_name = node[1] + "-" + rack_number + "-"
                elif node[1].find("kvm") != -1:
                    tmp_name = node[1] + "-"
                elif node[1].find("subrack") != -1:
                    tmp_name = node[1] + "-"
                else:
                    tmp_name = node[1]

                if tmp_name == "sw-cdu-" and not name.startswith("sw-cdu"):
                    # cdu0sw1 --> sw-cdu-001
                    # cdu0sw2 --> sw-cdu-002
                    # cdu1sw1 --> sw-cdu-003
                    # cdu1sw2 --> sw-cdu-004
                    # cdu2sw1 --> sw-cdu-005
                    # cdu2sw2 --> sw-cdu-006
                    digits = re.findall(r"\d+", name)
                    try:
                        tmp_id = int(digits[0]) * 2 + int(digits[1])
                    except IndexError as err:
                        click.secho(
                            f"Naming of CDU switch {name} is non-standard and cannot be parsed properly.",
                            fg="red",
                        )
                        raise ValueError(
                            f"Valid name standards for CDU switches are {node[0]} or {node[1]}.",
                        ) from err
                    common_name = f"{tmp_name}{tmp_id:0>3}"
                elif tmp_name.startswith("pdu"):
                    digits = re.findall(r"\d+", name)
                    try:
                        digit = digits[-1]
                    except IndexError as err:
                        click.secho(
                            f"Naming of PDU {name} is non-standard and cannot be parsed properly.",
                            fg="red",
                        )
                        raise ValueError(
                            f"Valid name standards for PDUs are {node[0]} or {node[1]}.",
                        ) from err
                    # Original names of:
                    #    pdu1 in x3113, or
                    #    x3113pdu1 in x3113, or
                    #    x3113p1 in x3113
                    # Becomes pdu-xXXXX-NNN, or pdu-x3113-001
                    common_name = f"{tmp_name}{digit:0>3}"
                else:
                    tmp_id = re.sub(
                        "^({})0*([1-9]*)".format(lookup_name),
                        r"\2",
                        name,
                    ).strip("-")
                    common_name = f"{tmp_name}{tmp_id:0>3}"
                return common_name
    return common_name


def get_node_type(name, mapper):
    """Map SHCD device name to device type.

    Args:
        name: A string from SHCD representing the device name
        mapper: An array of tuples (SHCD name, hostname, device type)

    Returns:
        node_type: A string with the device type
    """
    if name is not None:
        name = name.lower()
    node_type = None
    for node in mapper:
        for lookup_name in node[0]:
            if re.match("^{}".format(lookup_name.strip()), name):
                node_type = node[2]
                return node_type
    return node_type


def validate_shcd_slot_data(cell, sheet, warnings, is_src_slot=False):
    """Ensure that a slot from the SHCD is a proper string.

    Args:
        cell: Cell object of a port from the SHCD spreadsheet
        sheet: SHCD spreadsheet sheet/tab name
        warnings: Existing list of warnings to post to
        is_src_slot: Boolean hack to work around SHCD inconsistencies.

    Returns:
        slot: A cleaned up string value from the cell
    """
    valid_slot_names = ["ocp", "pcie-slot1", "bmc", "mgmt", "onboard", "s", "pci", None]
    location = None
    if cell.value is not None:
        location = cell.coordinate
    slot = cell.value
    if isinstance(slot, str):
        slot = slot.strip()
        if slot and slot[0] == "s":
            warnings["shcd_slot_data"].append(sheet + ":" + location)
            log.warning(
                'Prepending the character "s" to a slot will not be allowed in the future. '
                + f"Please correct cell {sheet}:{location} in the SHCD with value {slot} and prefer pcie-slot1.",
            )
            slot = "pcie-slot" + slot[1:]
        if slot and slot == "pci":
            warnings["shcd_slot_data"].append(sheet + ":" + location)
            log.warning(
                "The name pcie alone as a slot will not be allowed in the future"
                + f"Please correct cell {sheet}:{location} in the SHCD with value {slot} and prefer pcie-slot1.",
            )
            slot = "pcie-slot1"
        if slot not in valid_slot_names:
            warnings["shcd_slot_data"].append(sheet + ":" + location)
            log.warning(
                f"Slots must be named from the following list {valid_slot_names}."
                + f"Please correct cell {sheet}:{location} in the SHCD with value {slot}.",
            )
        if not slot:
            slot = None
    # Awful hack around the convention that src slot can be blank and a bmc
    # is noted by port 3 when there is physically one port.
    # NOTE: This is required for the port to get fixed.
    if is_src_slot:
        if sheet == "HMN" and slot is None:
            warnings["shcd_slot_data"].append(f"{sheet}:{cell.coordinate}")
            log.warning(
                'A source slot of type "bmc" for servers or "mgmt" for switches must be specified in the HMN tab. '
                + f"Please correct the SHCD for {sheet}:{cell.coordinate} with an empty value.",
            )
            slot = "bmc"
        elif sheet == "NMN" and slot is None:
            warnings["shcd_slot_data"].append(f"{sheet}:{cell.coordinate}")
            log.warning(
                'A source slot of type "onboard" for servers must be specified in the NMN tab. '
                + f"Please correct the SHCD for {sheet}:{cell.coordinate} with an empty value.",
            )
            slot = "onboard"

    return slot


def validate_shcd_port_data(cell, sheet, warnings, is_src_port=False, node_type=None):
    """Ensure that a port from the SHCD is a proper integer.

    Args:
        cell: Cell object of a port from the SHCD spreadsheet
        sheet: SHCD spreadsheet sheet/tab name
        warnings: Existing list of warnings to post to
        is_src_port: (optional) Boolean triggers a hack to work around SHCD inconsistencies
        node_type: (optional) if node_type is 'subrack' returns a None instead of an Exception

    Returns:
        port: A cleaned up integer value from the cell
    """
    location = None
    if cell.value is not None:
        location = cell.coordinate
    port = cell.value
    # Clean up the port if the cell is messy enough to be interpreted as a string
    if isinstance(port, str):
        port = port.strip()
        if not port:
            if node_type == "subrack":
                return None
            log.fatal(
                "A port number must be specified. "
                + f"Please correct the SHCD for {sheet}:{location} with an empty value",
            )
            sys.exit(1)
        if port[0].lower() == "j":
            warnings["shcd_port_data"].append(f"{sheet}:{location}")
            log.warning(
                'Prepending the character "j" to a port will not be allowed in the future. '
                + f'Please correct cell {sheet}:{location} in the SHCD with value "{port}"',
            )
            port = port[1:]
        # For SubRacks
        if port.upper() == "CMC" or port.upper() == "RCM":
            port = "1"
        if re.search(r"\D", port) is not None:
            click.secho(
                "Port numbers must be integers. "
                + f'Please correct in the SHCD for cell {sheet}:{location} with value "{port!r}"',
                fg="red",
            )
            sys.exit(1)
        if int(port) < 1:
            click.secho(
                "Ports numbers must be greater than 1. Port numbering must begin at 1. "
                + f'Please correct in the SHCD for cell {sheet}:{location} with value "{port!r}"',
                fg="red",
            )
            sys.exit(1)

        # String tests and corrections completed, convert the port to an int.
        port = int(port)

    if is_src_port:
        # Awful hack around the convention that src slot can be blank and a bmc
        # is noted by port 3 when there is physically one port.
        # NOTE: This assumes that the slot has already been corrected to "bmc"
        if sheet == "HMN" and port == 3:
            warnings["shcd_port_conventions"].append(f"{sheet}:{location}")
            log.warning(
                f'Bad slot/port convention for port "j{port}" in location {sheet}:{location}.'
                + 'This should be slot "bmc" for servers and "mgmt" for switches, and port "1".',
            )
            port = 1

    if port is None:
        if node_type == "subrack":
            return None
        else:
            click.secho(
                "A port number must be specified. "
                + f"Please correct the SHCD for {sheet}:{cell.coordinate} with an empty value",
                fg="red",
            )
            sys.exit(1)

    return int(port)


def node_model_from_shcd(factory, spreadsheet, sheets, edge=None):
    """Create a list of nodes from SHCD.

    Args:
        factory: Node factory object
        spreadsheet: The SHCD spreadsheet
        sheets: An array of tabs and their corners on the spreadsheet
        edge: Allow override of edge router

    Returns:
        node_list: A list of created nodes
        warnings: A list of warnings
    """
    # Generated nodes
    node_list = []
    node_name_list = []
    warnings = defaultdict(list)

    wb = load_workbook(spreadsheet, read_only=True, data_only=True)

    for tab in sheets:

        sheet = tab[0]
        range_start = tab[1]
        range_end = tab[2]

        log.info("---------------------------------------------")
        log.info(f"Working on tab/worksheet {sheet}")
        log.info("---------------------------------------------")
        log.info("")

        if sheet not in wb.sheetnames:
            click.secho(f"Tab {sheet} not found in {spreadsheet.name}\n", fg="red")
            click.secho(f"Available tabs: {wb.sheetnames}", fg="red")
            sys.exit(1)

        ws = wb[sheet]
        try:
            block = ws[range_start:range_end]
        except ValueError as err:
            log.fatal(err)
            click.secho(f"Bad range of cells entered for tab {sheet}.", fg="red")
            click.secho(f"{range_start}:{range_end}\n", fg="red")
            click.secho(
                "Ensure that the upper left corner (Labeled 'Source'), and the lower right corner of the table is entered.",
                fg="red",
            )
            sys.exit(1)

        # Process Headers
        required_header = [
            "Source",
            "Rack",
            "Location",
            "Slot",
            "Port",
            "Destination",
            "Rack",
            "Location",
            "Port",
        ]
        original_header = required_header.copy()

        header = block[0]
        if len(header) == 0 or len(header) < len(required_header):
            click.secho(
                f"Bad range of cells entered for tab {sheet}:{range_start}:{range_end}.",
                fg="red",
            )
            click.secho(
                "Not enough columns exist.\n"
                "Columns must exist in the following order, but may have other columns in between:\n"
                f"{original_header}\n"
                "Ensure that the upper left corner (Labeled 'Source'), and the lower right corner of the table is entered.",
                fg="red",
            )
            sys.exit(1)

        log.info(
            f"Expecting header with {len(required_header):>2} columns: {required_header}",
        )
        log.info(
            f"Found header with     {len(header):>2} columns: {[x.value for x in header]}",
        )
        log.info("Mapping required columns to actual.")

        subrack = None
        start_index = 0
        for required_index in range(len(required_header)):
            found = None
            for current_index in range(start_index, len(header)):
                if header[current_index].value == "Parent":
                    subrack = current_index
                elif header[current_index].value == required_header[required_index]:
                    found = current_index
                    break
                else:
                    found = None
            if found is not None:
                log.info(
                    f"Required header column {required_header[required_index]} "
                    f"found in spreadsheet cell {header[current_index].coordinate}",
                )
                required_header[required_index] = found
                start_index = current_index + 1
                found = None
            else:
                click.secho(
                    f"On tab {sheet}, header column {required_header[required_index]} not found.",
                    fg="red",
                )
                click.secho(
                    f"On tab {sheet}, the header is formatted incorrectly.\n"
                    "Columns must exist in the following order, but may have other columns in between:\n"
                    f"{original_header}",
                    fg="red",
                )
                sys.exit(1)

        # Process Data
        block = block[1:]
        for row in block:
            # Cable source
            try:
                current_row = row[required_header[0]].row
                log.debug(f"---- Working in sheet {sheet} on row {current_row} ----")
                src_name = row[required_header[0]].value.strip()
                tmp_slot = row[required_header[3]]
                tmp_port = row[required_header[4]]
                src_rack = None
                if row[required_header[1]].value:
                    src_rack = row[required_header[1]].value.strip().lower()
                src_elevation = None
                if row[required_header[2]].value:
                    src_elevation = row[required_header[2]].value.strip().lower()
                src_location = NodeLocation(src_rack, src_elevation)
            except AttributeError as err:
                click.secho(err, fg="red")
                click.secho(
                    f"Bad cell data or range of cells entered for sheet {sheet} in row {current_row} for source data.",
                    fg="red",
                )
                click.secho(
                    "Ensure the range entered does not contain empty cells.",
                    fg="red",
                )
                sys.exit(1)

            src_slot = validate_shcd_slot_data(
                tmp_slot,
                sheet,
                warnings,
                is_src_slot=True,
            )

            try:
                node_name = get_node_common_name(
                    src_name,
                    src_rack,
                    src_elevation,
                    factory.lookup_mapper(),
                )
            except ValueError as err:
                click.secho(err, fg="red")
                click.secho(
                    f"A parsing error occured in sheet {sheet} in row {current_row} for Source name.",
                    fg="red",
                )
                sys.exit(1)

            log.debug(f"Source Name Lookup:  {node_name}")
            log.debug(f"Source rack {src_rack} in location {src_elevation}")
            node_type = get_node_type(src_name, factory.lookup_mapper())
            log.debug(f"Source Node Type Lookup:  {node_type}")

            # Create src_node if it does not exist
            src_node = None
            src_index = None
            parent = None
            if node_type is not None and node_name is not None:
                if node_name not in node_name_list:
                    log.info(f"Creating new node {node_name} of type {node_type}")
                    try:
                        # hack for edge routers
                        if node_type == "customer_edge_router" and edge == "aruba":
                            node_type = "aruba_JL636A_edge_router"
                            print(node_type)

                        src_node = factory.generate_node(node_type)
                    except Exception as err:
                        log.fatal(err)
                        sys.exit(1)

                    src_node.common_name(node_name)
                    src_node.location(src_location)

                    node_list.append(src_node)
                    node_name_list.append(node_name)
                else:
                    # If the src is a subrack, add location info
                    if node_type == "subrack":
                        node_list[node_name_list.index(node_name)].location(
                            src_location,
                        )
                    log.debug(
                        f"Node {node_name} already exists, skipping node creation.",
                    )

                src_index = node_name_list.index(node_name)

                # update the node with parent location if it exists
                if subrack is not None:
                    parent = row[subrack].value
                    if parent is not None:
                        parent = parent.strip()
                        src_node = node_list[src_index]
                        src_location.parent(parent)
                        src_node.location(src_location)
            else:
                warnings["node_type"].append(
                    f"{src_name}@@{sheet}@@{row[required_header[0]].coordinate}",
                )
                log.warning(
                    f"Node type for {src_name} cannot be determined by node type ({node_type}) or node name ({node_name})",
                )

            # If a parent is specified, need to make a connection between the src and the parent
            if parent:
                src_parent_port = 1
                src_parent_slot = "cmc"
                src_node_port = NetworkPort(
                    number=src_parent_port,
                    slot=src_parent_slot,
                )

                parent_slot = "cmc"
                parent_rack = None
                parent_elevation = None
                try:
                    node_name_parent = get_node_common_name(
                        parent,
                        parent_rack,
                        parent_elevation,
                        factory.lookup_mapper(),
                    )
                except ValueError as err:
                    click.secho(err, fg="red")
                    click.secho(
                        f"A parsing error occured in sheet {sheet} in row {current_row} for Parent field.",
                        fg="red",
                    )
                    sys.exit(1)

                node_type_parent = get_node_type(parent, factory.lookup_mapper())

                parent_location = None
                parent_index = create_dst_node(
                    node_type_parent,
                    node_name_parent,
                    node_name_list,
                    parent_location,
                    node_list,
                    factory,
                )

                # We do not know what port numbers to connect to on the SubRack,
                # Get the next_free_port, and use that as the port number
                parent_port = node_list[parent_index].available_ports(
                    speed=1,
                    slot="cmc",
                    next_free_port=True,
                )

                parent_node_port = NetworkPort(number=parent_port, slot=parent_slot)
                src_node = node_list[src_index]
                parent_node = node_list[parent_index]
                try:
                    connect_src_dst(
                        src_node,
                        parent_node,
                        src_node_port,
                        parent_node_port,
                    )
                except Exception as err:
                    log.fatal(err)
                    click.secho(
                        f"Failed to connect {src_node.common_name()} "
                        + f"to parent {parent_node.common_name()} "
                        + f"while working on sheet {sheet}, row {current_row}.",
                        fg="red",
                    )
                    sys.exit(1)
                # If the tmp_port is None, make one connection:
                #   src ==> parent
                # Continue to connect the rest of the nodes
                if tmp_port.value is None:
                    continue

            src_port = validate_shcd_port_data(
                tmp_port,
                sheet,
                warnings,
                is_src_port=True,
                node_type=node_type,
            )

            # Sometimes a CMC is in the SHCD with no destination
            # If this is the case, continue to connect the other nodes
            if node_type == "subrack" and src_port is None:
                continue

            log.debug(f"Corrected Source Data:  {src_name} {src_slot} {src_port}")
            # Create the source port for the node
            src_node_port = NetworkPort(number=src_port, slot=src_slot)

            # Cable destination
            try:
                dst_name = row[required_header[5]].value.strip()
                dst_slot = None  # dst is always a switch
                dst_port = validate_shcd_port_data(
                    row[required_header[8]],
                    sheet,
                    warnings,
                )
                dst_rack = None
                if row[required_header[6]].value:
                    dst_rack = row[required_header[6]].value.strip()
                dst_elevation = None
                if row[required_header[7]].value:
                    dst_elevation = row[required_header[7]].value.strip()
                dst_location = NodeLocation(dst_rack, dst_elevation)
            except AttributeError as err:
                log.fatal(err)
                click.secho(
                    f"Bad cell data or range of cells entered for sheet {sheet} in row {current_row} for destination data.",
                    fg="red",
                )
                click.secho(
                    "Ensure the range entered does not contain empty cells.",
                    fg="red",
                )
                sys.exit(1)

            log.debug(f"Destination Data:  {dst_name} {dst_slot} {dst_port}")
            try:
                node_name = get_node_common_name(
                    dst_name,
                    dst_rack,
                    dst_elevation,
                    factory.lookup_mapper(),
                )
            except ValueError as err:
                click.secho(err, fg="red")
                click.secho(
                    f"A parsing error occured in sheet {sheet} in row {current_row} for Destination name.",
                    fg="red",
                )
                sys.exit(1)
            log.debug(f"Destination Name Lookup:  {node_name}")
            node_type = get_node_type(dst_name, factory.lookup_mapper())
            log.debug(f"Destination Node Type Lookup:  {node_type}")

            # Create dst_node if it does not exist
            try:
                dst_index = create_dst_node(
                    node_type,
                    node_name,
                    node_name_list,
                    dst_location,
                    node_list,
                    factory,
                )
            except Exception:
                warnings["node_type"].append(
                    f"{dst_name}@@{sheet}@@{row[required_header[5]].coordinate}",
                )

                log.warning(
                    f"Node type for {dst_name} cannot be determined by node type ({node_type}) or node name ({node_name})",
                )
                dst_index = None

            # Create the destination port
            dst_node_port = NetworkPort(number=dst_port, slot=dst_slot)

            if src_index is not None and dst_index is not None:
                src_node = node_list[src_index]
                dst_node = node_list[dst_index]
                try:
                    connect_src_dst(
                        src_node,
                        dst_node,
                        src_node_port,
                        dst_node_port,
                    )
                except Exception as err:
                    click.secho(err, fg="red")
                    click.secho(
                        f"Failed to connect {src_node.common_name()} "
                        + f"to {dst_node.common_name()} "
                        + f"while working on sheet {sheet}, row {current_row}.",
                        fg="red",
                    )
                    sys.exit(1)

    wb.close()
    return node_list, warnings


def create_dst_node(
    node_type,
    node_name,
    node_name_list,
    dst_location,
    node_list,
    factory,
):
    """Create a destination node.

    Args:
        node_type: Type of node
        node_name: Node name
        node_name_list: List of node names
        dst_location: Location object for the destination node
        node_list: A list of nodes
        factory: Node factory object

    Returns:
        dst_index: Index in the node_name_list of the destination

    Raises:
        Exception: If a node cannot be determined
    """
    dst_node = None
    dst_index = None
    if node_type is not None and node_name is not None:
        if node_name not in node_name_list:
            log.info(f"Creating new node {node_name} of type {node_type}")
            try:
                dst_node = factory.generate_node(node_type)
            except Exception as err:
                click.secho(err, fg="red")
                sys.exit(1)

            dst_node.common_name(node_name)
            dst_node.location(dst_location)

            node_list.append(dst_node)
            node_name_list.append(node_name)
        else:
            log.debug(
                f"Node {node_name} already exists, skipping node creation",
            )

        dst_index = node_name_list.index(node_name)

    else:
        raise Exception

    return dst_index


def connect_src_dst(
    src_node,
    dst_node,
    src_node_port,
    dst_node_port,
):
    """Connect a source node to a destination node.

    Args:
        src_node: Source node
        dst_node: Destination node
        src_node_port: Source node port
        dst_node_port: Destination node port

    Raises:
        Exception: If nodes cannot be connected
    """
    try:
        connected = src_node.connect(
            dst_node,
            src_port=src_node_port,
            dst_port=dst_node_port,
            strict=True,
        )
    except Exception as err:
        raise err

    if connected:
        log.info(
            f"Connected {src_node.common_name()} to {dst_node.common_name()} ",
        )
    else:
        raise Exception(
            f"Failed to connect {src_node.common_name()} to {dst_node.common_name()}. "
            "More information is often found by re-running with --log DEBUG enabled. "
            "Often the source of the error is that ports are being used twice on the device, "
            "that the correct network architecture model was specified, or plan-of-record hardware "
            "is not being used.",
        )


def node_list_warnings(node_list, warnings, out="-"):
    """Print the warnings found while validating the SHCD.

    Args:
        node_list: A list of nodes
        warnings: A dictionary of warnings
        out: Defaults to stdout, but will print to the file name passed in
    """
    dash = "-" * 80
    # Generate warnings
    # Additional warnings about the data will be entered here
    for node in node_list:
        if len(node.edges()) == 0:
            warnings["zero_connections"].append(node.common_name())

    # Print Warnings
    if warnings:
        click.secho("\nWarnings", fg="red", file=out)
        if warnings["node_type"]:
            click.secho(
                "\nNode type or port number could not be determined for the following."
                + "\nThese nodes are not currently included in the model."
                + "\n(This may be a missing architectural definition/lookup or a spelling error)",
                fg="red",
                file=out,
            )
            click.secho(dash, file=out)
            nodes = set(warnings["node_type"])
            nodes = natsort.natsorted(nodes)
            has_mac = False
            node_type_warnings = defaultdict(list)
            for node in nodes:
                # Missing nodes with '@@' are from the SHCD
                if "@@" in node:
                    name = node.split("@@")[0]
                    sheet = node.split("@@")[1]
                    cell = node.split("@@")[2]
                    node_type_warnings[sheet].append(f"Cell: {cell:<8s} Name: {name}")

                else:
                    # If the string has a mac address in it, set to True
                    if bool(re.search(r"(?:[0-9a-fA-F]:?){12}", str(node))):
                        has_mac = True
                    click.secho(node, fg="bright_white", file=out)
            if len(node_type_warnings) > 0:
                for sheet, cell_list in node_type_warnings.items():
                    click.secho(f"Sheet: {sheet}", fg="bright_white", file=out)
                    for cell in cell_list:
                        click.secho(cell, file=out)
                    click.secho("", file=out)
            if has_mac is True:
                click.secho(
                    "Nodes that show up as MAC addresses might need to have LLDP enabled.",
                    file=out,
                )
        if warnings["zero_connections"]:
            click.secho(
                "\nThe following nodes have zero connections"
                + "\n(The node type may not have been found or no connections are present)",
                fg="red",
                file=out,
            )
            click.secho(dash, file=out)
            nodes = set(warnings["zero_connections"])
            nodes = natsort.natsorted(nodes)
            for node in nodes:
                click.secho(node, fg="bright_white", file=out)
        if warnings["rename"]:
            click.secho(
                "\nThe following nodes should be renamed",
                fg="red",
                file=out,
            )
            click.secho(dash, file=out)
            nodes = set()
            for x in warnings["rename"]:
                new_name = x[1]
                if new_name == "":  # pragma: no cover
                    new_name = "(could not identify node)"
                nodes.add((x[0], new_name))
            nodes = natsort.natsorted(nodes)
            for node in nodes:
                click.secho(
                    f"{node[0]} should be renamed {node[1]}",
                    fg="bright_white",
                    file=out,
                )
        if warnings["shcd_port_data"]:
            click.secho(
                '\nSHCD port definitions are using a deprecated "j" prefix'
                + '\n(Ports should be an integer, remove the "j" in each cell)',
                fg="red",
                file=out,
            )
            click.secho(dash, file=out)
            port_warnings = defaultdict(list)
            for x in warnings["shcd_port_data"]:
                sheet = x.split(":")[0]
                cell = x.split(":")[1]
                port_warnings[sheet].append(cell)

            for sheet, cell_list in port_warnings.items():
                click.secho(f"Sheet: {sheet}", fg="bright_white", file=out)
                click.secho(f"{', '.join(cell_list)}\n", file=out)

        if warnings["shcd_port_conventions"]:
            click.secho(
                "\nSHCD port convention in the HMN tab is to use port 3 to represent BMCs."
                + '\n(For servers, correct the following cells to use a Slot of "bmc" and a port of "1")'
                + '\n(For Switches, correct the following cells to use a Slot of "mgmt" and a port of "1")',
                fg="red",
                file=out,
            )
            click.secho(dash, file=out)
            slot_warnings = defaultdict(list)
            for x in warnings["shcd_port_conventions"]:
                sheet = x.split(":")[0]
                cell = x.split(":")[1]
                slot_warnings[sheet].append(cell)

            for sheet, cell_list in slot_warnings.items():
                click.secho(f"Sheet: {sheet}", fg="bright_white", file=out)
                click.secho(f"{', '.join(cell_list)}\n", file=out)

        if warnings["shcd_slot_data"]:
            click.secho(
                "\nSHCD slot definitions used are either deprecated, missing or incorrect."
                + '\n(The cells below can be blank but should one of the following ["bmc", "ocp", "pcie-slot1, "mgmt", "onboard"])',
                fg="red",
                file=out,
            )
            click.secho(dash, file=out)
            def_warnings = defaultdict(list)
            for x in warnings["shcd_slot_data"]:
                sheet = x.split(":")[0]
                cell = x.split(":")[1]
                def_warnings[sheet].append(cell)

            for sheet, cell_list in def_warnings.items():
                click.secho(f"Sheet: {sheet}", fg="bright_white", file=out)
                click.secho(f"{', '.join(cell_list)}\n", file=out)


def switch_unused_ports(node_list):
    """Create a dictionary of unused ports.

    Args:
        node_list: A list of nodes

    Returns:
        unused_ports: Dictionary of switches and their unused ports
    """
    unused_ports = {}
    for node in node_list:
        if "sw" in node.common_name() and "sw-hsn" not in node.common_name():

            unused_ports[node.common_name()] = []
            unused_block = []
            logical_index = 1
            for port in node.ports():
                if port is None:
                    unused_ports[node.common_name()].append(logical_index)
                    unused_block.append(logical_index)
                    logical_index += 1
                    continue
                if unused_block:
                    unused_block = []  # reset
                logical_index += 1
            unused_ports[node.common_name()].pop()
    return unused_ports


def print_node_list(node_list, title, out="-"):
    """Print the nodes found in the SHCD.

    Args:
        node_list: A list of nodes
        title: Title to be printed
        out: Defaults to stdout, but will print to the file name passed in
    """
    dash = "-" * 60
    click.echo("\n", file=out)
    click.secho(f"{title} Node Connections", fg="bright_white", file=out)
    click.echo(dash, file=out)

    for node in node_list:
        click.echo(
            f"{node.id()}: {node.common_name()} connects to {len(node.edges())} nodes: {node.edges()}",
            file=out,
        )

    dash = "-" * 60
    click.echo("\n", file=out)
    click.secho(f"{title} Port Usage", fg="bright_white", file=out)
    click.echo(dash, file=out)

    for node in node_list:
        click.echo(
            f"{node.id()}: {node.common_name()} has the following port usage:",
            file=out,
        )

        unused_block = []
        logical_index = 1
        for port in node.ports():
            if port is None:
                unused_block.append(logical_index)
                logical_index += 1
                continue
            if unused_block:
                if len(unused_block) == 1:
                    port_string = f"{unused_block[0]:02}==>UNUSED"
                else:
                    port_string = f"{unused_block[0]:02}-{unused_block[len(unused_block) - 1]:02}==>UNUSED"
                unused_block = []  # reset
                click.secho(f"        {port_string}", fg="green", file=out)

            destination_node_name = [x.common_name() for x in node_list if x.id() == port["destination_node_id"]]
            destination_node_name = destination_node_name[0]
            destination_port_slot = None
            if port["destination_slot"] is None:
                destination_port_slot = f'{port["destination_port"]}'
            else:
                destination_port_slot = f'{port["destination_slot"]}:{port["destination_port"]}'
            if port["slot"] is None:
                port_string = f'{port["port"]:>02}==>{destination_node_name}:{destination_port_slot}'
            else:
                port_string = f'{port["slot"]}:{port["port"]}==>{destination_node_name}:{destination_port_slot}'
            click.echo(f"        {port_string}", file=out)
            logical_index += 1


def json_output(node_list, factory, architecture, ctx, out):
    """Create a schema-validated JSON Topology file from the model."""
    topology = []
    for node in node_list:
        topology.append(node.serialize())

    paddle = {
        "canu_version": version,
        "architecture": architecture,
        "shcd_file": path.basename(ctx.params["shcd"].name),
        "tabs": ctx.params["tabs"],
        "corners": ctx.params["corners"],
        "edge": ctx.params["edge"],
        "updated_at": (
            datetime.datetime.now().strftime(
                "%Y-%m-%d %H:%M:%S",
            )
        ),
        "topology": topology,
    }

    factory.validate_paddle(paddle)

    json_model = json.dumps(paddle, indent=2)
    click.echo(json_model, file=out)
