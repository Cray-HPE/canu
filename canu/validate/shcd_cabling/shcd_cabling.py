# MIT License
#
# (C) Copyright [2020] Hewlett Packard Enterprise Development LP
#
# Permission is hereby granted, free of charge, to any person obtaining a
# copy of this software and associated documentation files (the "Software"),
# to deal in the Software without restriction, including without limitation
# the rights to use, copy, modify, merge, publish, distribute, sublicense,
# and/or sell copies of the Software, and to permit persons to whom the
# Software is furnished to do so, subject to the following conditions:
#
# The above copyright notice and this permission notice shall be included
# in all copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL
# THE AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR
# OTHER LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE,
# ARISING FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR
# OTHER DEALINGS IN THE SOFTWARE.
"""CANU commands that validate the shcd against the current network cabling."""
from collections import defaultdict
import ipaddress
import logging
import os
from pathlib import Path
import sys

import click
from click_help_colors import HelpColorsCommand
from click_option_group import optgroup, RequiredMutuallyExclusiveOptionGroup
from click_params import IPV4_ADDRESS, Ipv4AddressListParamType
import click_spinner
from network_modeling.NetworkNodeFactory import NetworkNodeFactory
from openpyxl import load_workbook
import requests
import ruamel.yaml

from canu.report.switch.cabling.cabling import get_lldp
from canu.validate.network.cabling.cabling import node_model_from_canu
from canu.validate.shcd.shcd import (
    node_list_warnings,
    node_model_from_shcd,
    print_node_list,
)

yaml = ruamel.yaml.YAML()

# Get project root directory
if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):  # pragma: no cover
    project_root = sys._MEIPASS
else:
    prog = __file__
    project_root = Path(__file__).resolve().parent.parent.parent.parent

# Schema and Data files
hardware_schema_file = os.path.join(
    project_root, "network_modeling", "schema", "cray-network-hardware-schema.yaml"
)
hardware_spec_file = os.path.join(
    project_root, "network_modeling", "models", "cray-network-hardware.yaml"
)
architecture_schema_file = os.path.join(
    project_root, "network_modeling", "schema", "cray-network-architecture-schema.yaml"
)
architecture_spec_file = os.path.join(
    project_root, "network_modeling", "models", "cray-network-architecture.yaml"
)

canu_cache_file = os.path.join(project_root, "canu", "canu_cache.yaml")

log = logging.getLogger("validate_shcd")


@click.command(
    cls=HelpColorsCommand,
    help_headers_color="yellow",
    help_options_color="blue",
)
@click.option(
    "--architecture",
    "-a",
    type=click.Choice(["Full", "TDS"], case_sensitive=False),
    help="Shasta architecture",
    required=True,
    prompt="Architecture type",
)
@click.option(
    "--shcd",
    help="SHCD file",
    type=click.File("rb"),
    required=True,
)
@click.option(
    "--tabs",
    help="The tabs on the SHCD file to check, e.g. 10G_25G_40G_100G,NMN,HMN.",
)
@click.option(
    "--corners",
    help="The corners on each tab, comma separated e.g. 'J37,U227,J15,T47,J20,U167'.",
    # required=True,
)
@optgroup.group(
    "Network cabling IPv4 input sources",
    cls=RequiredMutuallyExclusiveOptionGroup,
)
@optgroup.option(
    "--ips",
    help="Comma separated list of IPv4 addresses of switches",
    type=Ipv4AddressListParamType(),
)
@optgroup.option(
    "--ips-file",
    help="File with one IPv4 address per line",
    type=click.File("r"),
)
@click.option("--username", default="admin", show_default=True, help="Switch username")
@click.option(
    "--password",
    prompt=True,
    hide_input=True,
    confirmation_prompt=False,
    help="Switch password",
)
@click.option(
    "--log",
    "log_",
    help="Level of logging.",
    type=click.Choice(["DEBUG", "INFO", "WARNING", "ERROR"]),
    # required=True,
    default="ERROR",
)
@click.pass_context
def shcd_cabling(
    ctx, architecture, shcd, tabs, corners, ips, ips_file, username, password, log_
):
    """Validate a SHCD file against the current network cabling .

    Pass in a SHCD file to validate that it works architecturally.

    This command will also use LLDP to determine the neighbors of the IP addresses passed in to validate that the network
    is properly connected architecturally.

    The validation will ensure that spine switches, leaf switches,
    edge switches, and nodes all are connected properly.

    \f
    # noqa: D301

    Args:
        ctx: CANU context settings
        architecture: Shasta architecture
        shcd: SHCD file
        tabs: The tabs on the SHCD file to check, e.g. 10G_25G_40G_100G,NMN,HMN.
        corners: The corners on each tab, comma separated e.g. 'J37,U227,J15,T47,J20,U167'.
        ips: Comma separated list of IPv4 addresses of switches
        ips_file: File with one IPv4 address per line
        username: Switch username
        password: Switch password
        log_: Level of logging.
    """
    logging.basicConfig(format="%(name)s - %(levelname)s: %(message)s", level=log_)

    if architecture.lower() == "full":
        architecture = "network_v2"
    elif architecture.lower() == "tds":
        architecture = "network_v2_tds"

    # SHCD Parsing
    sheets = []

    if not tabs:
        wb = load_workbook(shcd, read_only=True)
        click.secho("What tabs would you like to check in the SHCD?")
        tab_options = wb.sheetnames
        for x in tab_options:
            click.secho(f"{x}", fg="green")

        tabs = click.prompt(
            "Please enter the tabs to check separated by a comma, e.g. 10G_25G_40G_100G,NMN,HMN.",
            type=str,
        )

    if corners:
        if len(tabs.split(",")) * 2 != len(corners.split(",")):
            log.error("")
            click.secho("Not enough corners.\n", fg="red")
            click.secho(
                f"Make sure each tab: {tabs.split(',')} has 2 corners.\n", fg="red"
            )
            click.secho(
                f"There were {len(corners.split(','))} corners entered, but there should be {len(tabs.split(',')) * 2}.",
                fg="red",
            )
            click.secho(
                f"{corners}\n",
                fg="red",
            )
            return

        # Each tab should have 2 corners entered in comma separated
        for i in range(len(tabs.split(","))):
            # 0 -> 0,1
            # 1 -> 2,3
            # 2 -> 4,5

            sheets.append(
                (
                    tabs.split(",")[i],
                    corners.split(",")[i * 2].strip(),
                    corners.split(",")[i * 2 + 1].strip(),
                )
            )
    else:
        for tab in tabs.split(","):
            click.secho(f"\nFor the Sheet {tab}", fg="green")
            range_start = click.prompt(
                "Enter the cell of the upper left corner (Labeled 'Source')",
                type=str,
            )
            range_end = click.prompt(
                "Enter the cell of the lower right corner", type=str
            )
            sheets.append((tab, range_start, range_end))

    # IP Address / LLDP Parsing
    if ips_file:
        ips = []
        lines = [line.strip().replace(",", "") for line in ips_file]
        ips.extend([ipaddress.ip_address(line) for line in lines if IPV4_ADDRESS(line)])

    credentials = {"username": username, "password": password}

    errors = []
    ips_length = len(ips)

    if ips:
        with click_spinner.spinner():
            for i, ip in enumerate(ips, start=1):
                print(
                    f"  Connecting to {ip} - Switch {i} of {ips_length}        ",
                    end="\r",
                )
                try:
                    # Get LLDP info (stored in cache)
                    get_lldp(str(ip), credentials, True)

                except requests.exceptions.HTTPError:
                    errors.append(
                        [
                            str(ip),
                            f"Error connecting to switch {ip}, "
                            + "check that this IP is an Aruba switch, or check the username or password.",
                        ]
                    )
                except requests.exceptions.ConnectionError:
                    errors.append(
                        [
                            str(ip),
                            f"Error connecting to switch {ip},"
                            + " check the IP address and try again.",
                        ]
                    )
                except requests.exceptions.RequestException:  # pragma: no cover
                    errors.append(
                        [
                            str(ip),
                            f"Error connecting to switch {ip}.",
                        ]
                    )

    # Create SHCD Node factory
    shcd_factory = NetworkNodeFactory(
        hardware_schema=hardware_schema_file,
        hardware_data=hardware_spec_file,
        architecture_schema=architecture_schema_file,
        architecture_data=architecture_spec_file,
        architecture_version=architecture,
    )

    shcd_node_list, shcd_warnings = node_model_from_shcd(
        factory=shcd_factory, spreadsheet=shcd, sheets=sheets
    )
    dash = "-" * 100
    double_dash = "=" * 100
    # SHCD
    click.echo(double_dash)
    click.secho(
        "SHCD                                                              ",
        fg="bright_white",
    )
    click.echo(double_dash)
    print_node_list(shcd_node_list, "SHCD")

    node_list_warnings(shcd_node_list, shcd_warnings)

    # Create Cabling Node factory
    cabling_factory = NetworkNodeFactory(
        hardware_schema=hardware_schema_file,
        hardware_data=hardware_spec_file,
        architecture_schema=architecture_schema_file,
        architecture_data=architecture_spec_file,
        architecture_version=architecture,
    )

    # Open the updated cache to model nodes
    with open(canu_cache_file, "r+") as file:
        canu_cache = yaml.load(file)

    cabling_node_list, cabling_warnings = node_model_from_canu(
        cabling_factory, canu_cache, ips
    )

    click.echo("\n")
    click.echo(double_dash)
    click.secho("Cabling", fg="bright_white")
    click.echo(double_dash)

    print_node_list(cabling_node_list, "Cabling")

    node_list_warnings(cabling_node_list, cabling_warnings)

    click.echo("\n")
    click.echo(double_dash)
    click.secho("SHCD vs Cabling", fg="bright_white")
    click.echo(double_dash)
    compare_shcd_cabling(shcd_node_list, cabling_node_list)

    if len(errors) > 0:
        click.echo("\n")
        click.secho("Errors", fg="red")
        click.echo(dash)
        for error in errors:
            click.echo("{:<15s} - {}".format(error[0], error[1]))


def compare_shcd_cabling(shcd_node_list, cabling_node_list):
    """Print comparison of the SHCD and network.

    Args:
        shcd_node_list: A list of shcd nodes
        cabling_node_list: A list of nodes found on the network
    """
    dash = "-" * 60

    shcd_dict = node_list_to_dict(shcd_node_list)
    for x in shcd_dict:
        for y in shcd_dict[x]["ports"]:
            del y["destination_node_id"]
    cabling_dict = node_list_to_dict(cabling_node_list)
    for x in cabling_dict:
        for y in cabling_dict[x]["ports"]:
            del y["destination_node_id"]

    click.secho(
        "\nSHCD / Cabling Comparison",
        fg="bright_white",
    )
    click.secho(dash)

    for node in shcd_dict:

        if node in cabling_dict.keys():

            shcd_missing_connections = []
            for port in shcd_dict[node]["ports"]:
                if port not in cabling_dict[node]["ports"]:
                    shcd_missing_connections.append(
                        f'port {port["port"]} ==> {port["destination_node_name"]}'
                    )

            if len(shcd_missing_connections) > 0:
                click.secho(
                    f"{node : <16}: Found in SHCD and on the network, but missing the following connections on the "
                    + "network that were found in the SHCD:",
                    fg="green",
                )

                click.secho(f"                {str(shcd_missing_connections)}")

        else:
            click.secho(
                f"{node : <16}: Found in SHCD but not found on the network.",
                fg="blue",
            )
    print("--------")
    for node in cabling_dict:

        if node in shcd_dict.keys():

            cabling_missing_connections = []

            for port in cabling_dict[node]["ports"]:
                if port not in shcd_dict[node]["ports"]:
                    cabling_missing_connections.append(
                        f'port {port["port"]} ==> {port["destination_node_name"]}'
                    )

            if len(cabling_missing_connections) > 0:
                click.secho(
                    f"{node : <16}: Found on the network and in the SHCD, but missing the following connections in the"
                    + " SHCD that were found on the network:",
                    fg="green",
                )

                click.secho(f"                {str(cabling_missing_connections)}")

        if node not in shcd_dict.keys():
            click.secho(
                f"{node : <16}: Found on the network but not found in the SHCD.",
                fg="blue",
            )


def node_list_to_dict(node_list):
    """Convert node list to a dict.

    TODO:  Move this to model serialization

    Args:
        node_list: A list of nodes

    Returns:
        node_dict: The nodes converted to a dict
    """
    node_dict = defaultdict()
    node_id_dict = defaultdict()

    for node in node_list:
        node_id_dict[node.id()] = node.common_name()

    node_tmp = [node.serialize() for node in node_list]

    for node in node_tmp:
        for port in node["ports"]:
            port["destination_node_name"] = node_id_dict[port["destination_node_id"]]
        node_dict[node["common_name"]] = node

    return node_dict
