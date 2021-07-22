"""CANU commands that generate the config of the entire Shasta network."""
import json
import os
from pathlib import Path
import sys

import click
from click_help_colors import HelpColorsCommand
from jinja2 import Environment, FileSystemLoader, StrictUndefined
from network_modeling.NetworkNodeFactory import NetworkNodeFactory
from openpyxl import load_workbook
import requests
import ruamel.yaml
import urllib3

from canu.switch.config.config import (
    generate_switch_config,
    get_shasta_name,
    parse_sls_for_config,
    rename_sls_hostnames,
)
from canu.validate.shcd.shcd import node_model_from_shcd

yaml = ruamel.yaml.YAML()

# To disable warnings about unsecured HTTPS requests
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


# Get project root directory
if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):  # pragma: no cover
    project_root = sys._MEIPASS
else:
    prog = __file__
    project_root = Path(__file__).resolve().parent.parent.parent.parent

# Schema and Data files
hardware_schema_file = os.path.join(
    project_root, "network_modeling", "schema", "cray-network-hardware-schema.yaml"
)
hardware_spec_file = os.path.join(
    project_root, "network_modeling", "models", "cray-network-hardware.yaml"
)
architecture_schema_file = os.path.join(
    project_root, "network_modeling", "schema", "cray-network-architecture-schema.yaml"
)
architecture_spec_file = os.path.join(
    project_root, "network_modeling", "models", "cray-network-architecture.yaml"
)

canu_cache_file = os.path.join(project_root, "canu", "canu_cache.yaml")

# Import templates
network_templates_folder = os.path.join(
    project_root, "network_modeling", "configs", "templates"
)
env = Environment(
    loader=FileSystemLoader(network_templates_folder),
    # trim_blocks=True,
    undefined=StrictUndefined,
)


@click.command(
    cls=HelpColorsCommand,
    help_headers_color="yellow",
    help_options_color="blue",
)
@click.option(
    "--architecture",
    "-a",
    type=click.Choice(["Full", "TDS"], case_sensitive=False),
    help="Shasta architecture",
    required=True,
    prompt="Architecture type",
)
@click.option(
    "--shcd",
    help="SHCD file",
    type=click.File("rb"),
    required=True,
)
@click.option(
    "--tabs",
    help="The tabs on the SHCD file to check, e.g. 10G_25G_40G_100G,NMN,HMN.",
)
@click.option(
    "--corners",
    help="The corners on each tab, comma separated e.g. 'J37,U227,J15,T47,J20,U167'.",
)
@click.option("--csi-folder", help="Directory containing the CSI json file")
@click.option(
    "--auth-token",
    envvar="SLS_TOKEN",
    help="Token for SLS authentication",
)
@click.option("--sls-address", default="api-gw-service-nmn.local", show_default=True)
@click.option(
    "--folder",
    help="Folder to store config files",
    required=True,
    prompt="Folder for configs",
)
@click.option(
    "--password",
    prompt=True,
    hide_input=True,
    confirmation_prompt=False,
    help="Switch password",
)
@click.pass_context
def config(
    ctx,
    architecture,
    shcd,
    tabs,
    corners,
    csi_folder,
    auth_token,
    sls_address,
    folder,
    password,
):
    """Generate the config of all Aruba switches (API v10.04) on the network using the SHCD.

    In order to generate network switch config, a valid SHCD must be passed in and system variables must be read in from either
    CSI output or the SLS API.

    Use the `--folder FOLDERNAME` flag to output all the switch configs to a folder.
    \f
    # noqa: D301

    Args:
        ctx: CANU context settings
        architecture: Shasta architecture
        shcd: SHCD file
        tabs: The tabs on the SHCD file to check, e.g. 10G_25G_40G_100G,NMN,HMN.
        corners: The corners on each tab, comma separated e.g. 'J37,U227,J15,T47,J20,U167'.
        csi_folder: Directory containing the CSI json file
        auth_token: Token for SLS authentication
        sls_address: The address of SLS
        folder: Folder to store config files
        password: Switch password
    """
    if architecture.lower() == "full":
        architecture = "network_v2"
        template_folder = "full"
    elif architecture.lower() == "tds":
        architecture = "network_v2_tds"
        template_folder = "tds"

    # SHCD Parsing
    sheets = []

    if not tabs:
        wb = load_workbook(shcd, read_only=True)
        click.secho("What tabs would you like to check in the SHCD?")
        tab_options = wb.sheetnames
        for x in tab_options:
            click.secho(f"{x}", fg="green")

        tabs = click.prompt(
            "Please enter the tabs to check separated by a comma, e.g. 10G_25G_40G_100G,NMN,HMN.",
            type=str,
        )

    if corners:
        if len(tabs.split(",")) * 2 != len(corners.split(",")):
            click.secho("Not enough corners.\n", fg="red")
            click.secho(
                f"Make sure each tab: {tabs.split(',')} has 2 corners.\n", fg="red"
            )
            click.secho(
                f"There were {len(corners.split(','))} corners entered, but there should be {len(tabs.split(',')) * 2}.",
                fg="red",
            )
            click.secho(
                f"{corners}\n",
                fg="red",
            )
            return

        # Each tab should have 2 corners entered in comma separated
        for i in range(len(tabs.split(","))):
            # 0 -> 0,1
            # 1 -> 2,3
            # 2 -> 4,5

            sheets.append(
                (
                    tabs.split(",")[i],
                    corners.split(",")[i * 2].strip(),
                    corners.split(",")[i * 2 + 1].strip(),
                )
            )
    else:
        for tab in tabs.split(","):
            click.secho(f"\nFor the Sheet {tab}", fg="green")
            range_start = click.prompt(
                "Enter the cell of the upper left corner (Labeled 'Source')",
                type=str,
            )
            range_end = click.prompt(
                "Enter the cell of the lower right corner", type=str
            )
            sheets.append((tab, range_start, range_end))

    # Create Node factory
    factory = NetworkNodeFactory(
        hardware_schema=hardware_schema_file,
        hardware_data=hardware_spec_file,
        architecture_schema=architecture_schema_file,
        architecture_data=architecture_spec_file,
        architecture_version=architecture,
    )

    # Get nodes from SHCD
    shcd_node_list, shcd_warnings = node_model_from_shcd(
        factory=factory, spreadsheet=shcd, sheets=sheets
    )

    # Parse sls_input_file.json file from CSI
    if csi_folder:
        try:
            with open(os.path.join(csi_folder, "sls_input_file.json"), "r") as f:
                input_json = json.load(f)

                # Format the input to be like the SLS JSON
                sls_json = [
                    network[x]
                    for network in [input_json.get("Networks", {})]
                    for x in network
                ]

        except FileNotFoundError:
            click.secho(
                "The file sls_input_file.json was not found, check that this is the correct CSI directory.",
                fg="red",
            )
            return
    else:
        # Get SLS config
        token = os.environ.get("SLS_TOKEN")

        # Token file takes precedence over the environmental variable
        if auth_token != token:
            try:
                with open(auth_token) as f:
                    data = json.load(f)
                    token = data["access_token"]

            except Exception:
                click.secho(
                    "Invalid token file, generate another token or try again.",
                    fg="white",
                    bg="red",
                )
                return

        # SLS
        url = "https://" + sls_address + "/apis/sls/v1/networks"
        try:
            response = requests.get(
                url,
                headers={
                    "Content-Type": "application/json",
                    "Authorization": f"Bearer {token}",
                },
                verify=False,
            )
            response.raise_for_status()

            sls_json = response.json()

        except requests.exceptions.ConnectionError:
            return click.secho(
                f"Error connecting to SLS {sls_address}, check the address or pass in a new address using --sls-address.",
                fg="white",
                bg="red",
            )
        except requests.exceptions.HTTPError:
            bad_token_reason = (
                "environmental variable 'SLS_TOKEN' is correct."
                if auth_token == token
                else "token is valid, or generate a new one."
            )
            return click.secho(
                f"Error connecting SLS {sls_address}, check that the {bad_token_reason}",
                fg="white",
                bg="red",
            )
    sls_variables = parse_sls_for_config(sls_json)

    # For versions of Shasta < 1.6, the SLS Hostnames need to be renamed
    if ctx.obj["shasta"]:
        shasta = ctx.obj["shasta"]
        if float(shasta) < 1.6:
            sls_variables = rename_sls_hostnames(sls_variables)

    # make folder
    if not os.path.exists(folder):
        os.makedirs(folder)

    for node in shcd_node_list:
        switch_name = node.common_name()
        node_shasta_name = get_shasta_name(switch_name, factory.lookup_mapper())

        if node_shasta_name in ["sw-cdu", "sw-leaf-bmc", "sw-leaf", "sw-spine"]:

            switch_config = generate_switch_config(
                shcd_node_list,
                factory,
                switch_name,
                sls_variables,
                template_folder,
                password,
            )

            with open(f"{folder}/{switch_name}.aos", "w+") as f:
                f.write(switch_config)

            click.secho(f"{switch_name} Config Generated", fg="bright_white")

    return
