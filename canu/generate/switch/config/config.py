# MIT License
#
# (C) Copyright [2021] Hewlett Packard Enterprise Development LP
#
# Permission is hereby granted, free of charge, to any person obtaining a
# copy of this software and associated documentation files (the "Software"),
# to deal in the Software without restriction, including without limitation
# the rights to use, copy, modify, merge, publish, distribute, sublicense,
# and/or sell copies of the Software, and to permit persons to whom the
# Software is furnished to do so, subject to the following conditions:
#
# The above copyright notice and this permission notice shall be included
# in all copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL
# THE AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR
# OTHER LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE,
# ARISING FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR
# OTHER DEALINGS IN THE SOFTWARE.
"""CANU generate switch config commands."""
from collections import defaultdict
import json
import os
from os import environ, path
from pathlib import Path
import re
import sys

import click
from click_help_colors import HelpColorsCommand
from hier_config import HConfig, Host
from jinja2 import Environment, FileSystemLoader, StrictUndefined
import natsort
import netaddr
from network_modeling.NetworkNodeFactory import NetworkNodeFactory
from openpyxl import load_workbook
import requests
from ruamel.yaml import YAML
import urllib3

from canu.utils.cache import cache_directory
from canu.validate.shcd.shcd import node_model_from_shcd

yaml = YAML()

# To disable warnings about unsecured HTTPS requests
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


# Get project root directory
if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):  # pragma: no cover
    project_root = sys._MEIPASS
else:
    prog = __file__
    project_root = Path(__file__).resolve().parent.parent.parent.parent.parent

# Schema and Data files
hardware_schema_file = path.join(
    project_root,
    "network_modeling",
    "schema",
    "cray-network-hardware-schema.yaml",
)
hardware_spec_file = path.join(
    project_root,
    "network_modeling",
    "models",
    "cray-network-hardware.yaml",
)
architecture_schema_file = path.join(
    project_root,
    "network_modeling",
    "schema",
    "cray-network-architecture-schema.yaml",
)
architecture_spec_file = path.join(
    project_root,
    "network_modeling",
    "models",
    "cray-network-architecture.yaml",
)

canu_cache_file = path.join(cache_directory(), "canu_cache.yaml")
canu_config_file = path.join(project_root, "canu", "canu.yaml")

# Import templates
network_templates_folder = path.join(
    project_root,
    "network_modeling",
    "configs",
    "templates",
)
env = Environment(
    loader=FileSystemLoader(network_templates_folder),
    undefined=StrictUndefined,
)

# Get CSM versions from canu.yaml
with open(canu_config_file, "r") as file:
    canu_config = yaml.load(file)

csm_options = canu_config["csm_versions"]


@click.command(
    cls=HelpColorsCommand,
    help_headers_color="yellow",
    help_options_color="blue",
)
@click.option(
    "--csm",
    type=click.Choice(csm_options),
    help="CSM network version",
    prompt="CSM network version",
    required=True,
    show_choices=True,
)
@click.option(
    "--architecture",
    "-a",
    type=click.Choice(["Full", "TDS", "V1"], case_sensitive=False),
    help="CSM architecture",
    required=True,
    prompt="Architecture type",
)
@click.option(
    "--shcd",
    help="SHCD file",
    type=click.File("rb"),
    required=True,
)
@click.option(
    "--tabs",
    help="The tabs on the SHCD file to check, e.g. 10G_25G_40G_100G,NMN,HMN.",
)
@click.option(
    "--corners",
    help="The corners on each tab, comma separated e.g. 'J37,U227,J15,T47,J20,U167'.",
)
@click.option(
    "--name",
    "switch_name",
    required=True,
    help="The name of the switch to generate config e.g. 'sw-spine-001'",
    prompt="Switch Name",
)
@click.option(
    "--sls-file",
    help="File containing system SLS JSON data.",
    type=click.File("r"),
)
@click.option(
    "--auth-token",
    envvar="SLS_TOKEN",
    help="Token for SLS authentication",
)
@click.option("--sls-address", default="api-gw-service-nmn.local", show_default=True)
@click.option(
    "--out",
    help="Output results to a file",
    type=click.File("w"),
    default="-",
)
@click.option(
    "--override",
    help="Switch configuration override",
    type=click.Path(),
)
@click.pass_context
def config(
    ctx,
    csm,
    architecture,
    shcd,
    tabs,
    corners,
    switch_name,
    sls_file,
    auth_token,
    sls_address,
    out,
    override,
):
    """Switch config commands.

    In order to generate switch config, a valid SHCD must be passed in and system variables must be read in from either
    CSI output or the SLS API.

    To generate config for a specific switch, a hostname must also be passed in using the `--name HOSTNAME` flag.
    To output the config to a file, append the `--out FILENAME` flag.

    \f
    # noqa: D301

    Args:
        ctx: CANU context settings
        csm: CSM version
        architecture: CSM architecture
        shcd: SHCD file
        tabs: The tabs on the SHCD file to check, e.g. 10G_25G_40G_100G,NMN,HMN.
        corners: The corners on each tab, comma separated e.g. 'J37,U227,J15,T47,J20,U167'.
        switch_name: Switch name
        sls_file: JSON file containing SLS data
        auth_token: Token for SLS authentication
        sls_address: The address of SLS
        out: Name of the output file
        override: Input file to ignore switch configuration
    """
    if architecture.lower() == "full":
        architecture = "network_v2"
        template_folder = "full"
        vendor_folder = "aruba"
    elif architecture.lower() == "tds":
        architecture = "network_v2_tds"
        template_folder = "tds"
        vendor_folder = "aruba"
    elif architecture.lower() == "v1":
        architecture = "network_v1"
        template_folder = "full"
        vendor_folder = "dellmellanox"

    # SHCD Parsing
    sheets = []

    if not tabs:
        wb = load_workbook(shcd, read_only=True)
        click.secho("What tabs would you like to check in the SHCD?")
        tab_options = wb.sheetnames
        for x in tab_options:
            click.secho(f"{x}", fg="green")

        tabs = click.prompt(
            "Please enter the tabs to check separated by a comma, e.g. 10G_25G_40G_100G,NMN,HMN.",
            type=str,
        )

    if corners:
        if len(tabs.split(",")) * 2 != len(corners.split(",")):
            click.secho("Not enough corners.\n", fg="red")
            click.secho(
                f"Make sure each tab: {tabs.split(',')} has 2 corners.\n",
                fg="red",
            )
            click.secho(
                f"There were {len(corners.split(','))} corners entered, but there should be {len(tabs.split(',')) * 2}.",
                fg="red",
            )
            click.secho(
                f"{corners}\n",
                fg="red",
            )
            return

        # Each tab should have 2 corners entered in comma separated
        for i in range(len(tabs.split(","))):
            # 0 -> 0,1
            # 1 -> 2,3
            # 2 -> 4,5

            sheets.append(
                (
                    tabs.split(",")[i],
                    corners.split(",")[i * 2].strip(),
                    corners.split(",")[i * 2 + 1].strip(),
                ),
            )
    else:
        for tab in tabs.split(","):
            click.secho(f"\nFor the Sheet {tab}", fg="green")
            range_start = click.prompt(
                "Enter the cell of the upper left corner (Labeled 'Source')",
                type=str,
            )
            range_end = click.prompt(
                "Enter the cell of the lower right corner",
                type=str,
            )
            sheets.append((tab, range_start, range_end))

    # Create Node factory
    factory = NetworkNodeFactory(
        hardware_schema=hardware_schema_file,
        hardware_data=hardware_spec_file,
        architecture_schema=architecture_schema_file,
        architecture_data=architecture_spec_file,
        architecture_version=architecture,
    )

    # Get nodes from SHCD
    shcd_node_list, shcd_warnings = node_model_from_shcd(
        factory=factory,
        spreadsheet=shcd,
        sheets=sheets,
    )

    # Parse SLS input file.
    if sls_file:
        try:
            input_json = json.load(sls_file)
        except (json.JSONDecodeError, UnicodeDecodeError):
            click.secho(
                f"The file {sls_file.name} is not valid JSON.",
                fg="red",
            )
            return

        # Format the input to be like the SLS JSON
        sls_json = [
            network[x] for network in [input_json.get("Networks", {})] for x in network
        ]

    else:
        # Get SLS config
        token = environ.get("SLS_TOKEN")

        # Token file takes precedence over the environmental variable
        if auth_token != token:
            try:
                with open(auth_token) as f:
                    data = json.load(f)
                    token = data["access_token"]

            except Exception:
                click.secho(
                    "Invalid token file, generate another token or try again.",
                    fg="white",
                    bg="red",
                )
                return

        # SLS
        url = "https://" + sls_address + "/apis/sls/v1/networks"
        try:
            response = requests.get(
                url,
                headers={
                    "Content-Type": "application/json",
                    "Authorization": f"Bearer {token}",
                },
                verify=False,
            )
            response.raise_for_status()

            sls_json = response.json()

        except requests.exceptions.ConnectionError:
            return click.secho(
                f"Error connecting to SLS {sls_address}, check the address or pass in a new address using --sls-address.",
                fg="white",
                bg="red",
            )
        except requests.exceptions.HTTPError:
            bad_token_reason = (
                "environmental variable 'SLS_TOKEN' is correct."
                if auth_token == token
                else "token is valid, or generate a new one."
            )
            return click.secho(
                f"Error connecting SLS {sls_address}, check that the {bad_token_reason}",
                fg="white",
                bg="red",
            )
    sls_variables = parse_sls_for_config(sls_json)

    # For versions of csm < 1.2, the SLS Hostnames need to be renamed
    if csm:
        if float(csm) < 1.2:
            sls_variables = rename_sls_hostnames(sls_variables)

    if override:
        try:
            with open(os.path.join(override), "r") as f:
                override = yaml.load(f)
        except FileNotFoundError:
            click.secho(
                "The override yaml file was not found, check that you entered the right file name and path.",
                fg="red",
            )
            exit(1)

    switch_config, devices, unknown = generate_switch_config(
        architecture,
        shcd_node_list,
        factory,
        switch_name,
        sls_variables,
        template_folder,
        vendor_folder,
        override,
    )

    dash = "-" * 60
    click.echo("\n")
    click.echo(dash)

    if override:
        click.secho(f"{switch_name} Override Switch Config", fg="yellow")
    else:
        click.secho(f"{switch_name} Switch Config", fg="bright_white")
    click.echo(switch_config, file=out)

    if len(unknown) > 0:
        click.secho("\nWarning", fg="red")

        click.secho(
            "\nThe following devices were discovered in the input data, but the CANU model cannot determine "
            + "the type and generate a configuration.\nApplying this configuration without considering these "
            + "devices will likely result in loss of contact with these devices."
            + "\nEnsure valid input, submit a bug to CANU and manually add these devices to the configuration.",
            fg="red",
        )
        click.secho(dash)
        for x in unknown:
            click.secho(x, fg="bright_white")
    return


def get_shasta_name(name, mapper):
    """Parse mapper to get Shasta name."""
    for node in mapper:
        shasta_name = node[1]
        if shasta_name in name:
            return shasta_name


def generate_switch_config(
    architecture,
    shcd_node_list,
    factory,
    switch_name,
    sls_variables,
    template_folder,
    vendor_folder,
    override,
):
    """Generate switch config.

    Args:
        architecture: CSM architecture
        shcd_node_list: List of nodes from the SHCD
        factory: Node factory object
        switch_name: Switch hostname
        sls_variables: Dictionary containing SLS variables
        template_folder: Architecture folder contaning the switch templates
        vendor_folder: Vendor folder contaning the template_folder
        override: Input file that defines what config should be ignored

    Returns:
        switch_config: The generated switch configuration
    """
    node_shasta_name = get_shasta_name(switch_name, factory.lookup_mapper())

    if node_shasta_name is None:
        return Exception(
            click.secho(
                f"For switch {switch_name}, the type cannot be determined. Please check the switch name and try again.",
                fg="red",
            ),
        )
    elif node_shasta_name not in ["sw-cdu", "sw-leaf-bmc", "sw-leaf", "sw-spine"]:
        return Exception(
            click.secho(
                f"{switch_name} is not a switch. Only switch config can be generated.",
                fg="red",
            ),
        )

    is_primary, primary, secondary = switch_is_primary(switch_name)

    templates = {
        "sw-spine": {
            "primary": f"{vendor_folder}/{template_folder}/sw-spine.primary.j2",
            "secondary": f"{vendor_folder}/{template_folder}/sw-spine.secondary.j2",
        },
        "sw-cdu": {
            "primary": f"{vendor_folder}/common/sw-cdu.primary.j2",
            "secondary": f"{vendor_folder}/common/sw-cdu.secondary.j2",
        },
        "sw-leaf": {
            "primary": f"{vendor_folder}/{template_folder}/sw-leaf.primary.j2",
            "secondary": f"{vendor_folder}/{template_folder}/sw-leaf.secondary.j2",
        },
        "sw-leaf-bmc": {
            "primary": f"{vendor_folder}/{template_folder}/sw-leaf-bmc.j2",
            "secondary": f"{vendor_folder}/{template_folder}/sw-leaf-bmc.j2",
        },
    }
    template_name = templates[node_shasta_name][
        "primary" if is_primary else "secondary"
    ]
    template = env.get_template(template_name)
    variables = {
        "HOSTNAME": switch_name,
        "NCN_W001": sls_variables["ncn_w001"],
        "NCN_W002": sls_variables["ncn_w002"],
        "NCN_W003": sls_variables["ncn_w003"],
        "CAN": sls_variables["CAN"],
        "CAN_NETMASK": sls_variables["CAN_NETMASK"],
        "CAN_NETWORK_IP": sls_variables["CAN_NETWORK_IP"],
        "CAN_PREFIX_LEN": sls_variables["CAN_PREFIX_LEN"],
        "CMN": sls_variables["CMN"],
        "CMN_NETMASK": sls_variables["CMN_NETMASK"],
        "CMN_NETWORK_IP": sls_variables["CMN_NETWORK_IP"],
        "CMN_PREFIX_LEN": sls_variables["CMN_PREFIX_LEN"],
        "MTL_NETMASK": sls_variables["MTL_NETMASK"],
        "MTL_PREFIX_LEN": sls_variables["MTL_PREFIX_LEN"],
        "NMN": sls_variables["NMN"],
        "NMN_NETMASK": sls_variables["NMN_NETMASK"],
        "NMN_NETWORK_IP": sls_variables["NMN_NETWORK_IP"],
        "NMN_PREFIX_LEN": sls_variables["NMN_PREFIX_LEN"],
        "HMN": sls_variables["HMN"],
        "HMN_NETMASK": sls_variables["HMN_NETMASK"],
        "HMN_NETWORK_IP": sls_variables["HMN_NETWORK_IP"],
        "HMN_PREFIX_LEN": sls_variables["HMN_PREFIX_LEN"],
        "HMN_MTN": sls_variables["HMN_MTN"],
        "HMN_MTN_NETMASK": sls_variables["HMN_MTN_NETMASK"],
        "HMN_MTN_NETWORK_IP": sls_variables["HMN_MTN_NETWORK_IP"],
        "HMN_MTN_PREFIX_LEN": sls_variables["HMN_MTN_PREFIX_LEN"],
        "NMN_MTN": sls_variables["NMN_MTN"],
        "NMN_MTN_NETMASK": sls_variables["NMN_MTN_NETMASK"],
        "NMN_MTN_NETWORK_IP": sls_variables["NMN_MTN_NETWORK_IP"],
        "NMN_MTN_PREFIX_LEN": sls_variables["NMN_MTN_PREFIX_LEN"],
        "HMNLB": sls_variables["HMNLB"],
        "HMNLB_NETMASK": sls_variables["HMNLB_NETMASK"],
        "HMNLB_NETWORK_IP": sls_variables["HMNLB_NETWORK_IP"],
        "HMNLB_PREFIX_LEN": sls_variables["HMNLB_PREFIX_LEN"],
        "NMNLB": sls_variables["NMNLB"],
        "NMNLB_NETMASK": sls_variables["NMNLB_NETMASK"],
        "NMNLB_NETWORK_IP": sls_variables["NMNLB_NETWORK_IP"],
        "NMNLB_PREFIX_LEN": sls_variables["NMNLB_PREFIX_LEN"],
        "HMN_IP_GATEWAY": sls_variables["HMN_IP_GATEWAY"],
        "MTL_IP_GATEWAY": sls_variables["MTL_IP_GATEWAY"],
        "NMN_IP_GATEWAY": sls_variables["NMN_IP_GATEWAY"],
        "CAN_IP_GATEWAY": sls_variables["CAN_IP_GATEWAY"],
        "CAN_IP_PRIMARY": sls_variables["CAN_IP_PRIMARY"],
        "CAN_IP_SECONDARY": sls_variables["CAN_IP_SECONDARY"],
        "CMN_IP_GATEWAY": sls_variables["CMN_IP_GATEWAY"],
        "CMN_IP_PRIMARY": sls_variables["CMN_IP_PRIMARY"],
        "CMN_IP_SECONDARY": sls_variables["CMN_IP_SECONDARY"],
        "NMN_MTN_CABINETS": sls_variables["NMN_MTN_CABINETS"],
        "HMN_MTN_CABINETS": sls_variables["HMN_MTN_CABINETS"],
    }
    cabling = {}
    cabling["nodes"], unknown = get_switch_nodes(
        switch_name,
        shcd_node_list,
        factory,
        sls_variables,
    )

    if switch_name not in sls_variables["HMN_IPs"].keys():
        click.secho(f"Cannot find {switch_name} in CSI / SLS nodes.", fg="red")
        exit(1)

    variables["HMN_IP"] = sls_variables["HMN_IPs"][switch_name]
    variables["MTL_IP"] = sls_variables["MTL_IPs"][switch_name]
    variables["NMN_IP"] = sls_variables["NMN_IPs"][switch_name]

    last_octet = variables["HMN_IP"].split(".")[3]
    variables["LOOPBACK_IP"] = "10.2.0." + last_octet
    variables["IPV6_IP"] = "2001:db8:beef:99::" + last_octet + "/128"

    if node_shasta_name in ["sw-spine", "sw-leaf", "sw-cdu"]:
        # Get connections to switch pair
        pair_connections = get_pair_connections(cabling["nodes"], switch_name)
        length_connections = len(pair_connections)

        if length_connections == 3:
            variables["VSX_KEEPALIVE"] = pair_connections[0]
            variables["VSX_ISL_PORT1"] = pair_connections[1]
            variables["VSX_ISL_PORT2"] = pair_connections[2]
        elif length_connections == 2:
            variables["VSX_KEEPALIVE"] = "mgmt0"
            variables["VSX_ISL_PORT1"] = pair_connections[0]
            variables["VSX_ISL_PORT2"] = pair_connections[1]

    # get VLANs and IPs for CDU switches
    if "sw-cdu" in node_shasta_name:
        nodes_by_name = {}
        nodes_by_id = {}
        destination_rack_list = []
        variables["NMN_MTN_VLANS"] = []
        variables["HMN_MTN_VLANS"] = []

        for node in shcd_node_list:
            node_tmp = node.serialize()
            name = node_tmp["common_name"]
            nodes_by_name[name] = node_tmp
            nodes_by_id[node_tmp["id"]] = node_tmp
        for port in nodes_by_name[switch_name]["ports"]:
            destination_rack = nodes_by_id[port["destination_node_id"]]["location"][
                "rack"
            ]

            destination_rack_list.append(int(re.search(r"\d+", destination_rack)[0]))
        for cabinets in (
            sls_variables["NMN_MTN_CABINETS"] + sls_variables["HMN_MTN_CABINETS"]
        ):
            ip_address = netaddr.IPNetwork(cabinets["CIDR"])
            is_primary = switch_is_primary(switch_name)
            sls_rack_int = int(re.search(r"\d+", (cabinets["Name"]))[0])
            if sls_rack_int in destination_rack_list:
                if cabinets in sls_variables["NMN_MTN_CABINETS"]:
                    variables["NMN_MTN_VLANS"].append(cabinets)
                    variables["NMN_MTN_VLANS"][-1][
                        "PREFIX_LENGTH"
                    ] = ip_address.prefixlen
                    if is_primary[0]:
                        ip = str(ip_address[2])
                        variables["NMN_MTN_VLANS"][-1]["IP"] = ip
                    else:
                        ip = str(ip_address[3])
                        variables["NMN_MTN_VLANS"][-1]["IP"] = ip

                if cabinets in sls_variables["HMN_MTN_CABINETS"]:
                    variables["HMN_MTN_VLANS"].append(cabinets)
                    variables["HMN_MTN_VLANS"][-1][
                        "PREFIX_LENGTH"
                    ] = ip_address.prefixlen
                    if is_primary[0]:
                        ip = str(ip_address[2])
                        variables["HMN_MTN_VLANS"][-1]["IP"] = ip
                    else:
                        ip = str(ip_address[3])
                        variables["HMN_MTN_VLANS"][-1]["IP"] = ip

    switch_config = template.render(
        variables=variables,
        cabling=cabling,
    )
    devices = set()
    for node in cabling["nodes"]:
        devices.add(node["subtype"])
    if architecture == "network_v1":
        dell_options_file = os.path.join(
            project_root,
            "canu",
            "validate",
            "switch",
            "config",
            "dell_options.yaml",
        )
        mellanox_options_file = os.path.join(
            project_root,
            "canu",
            "validate",
            "switch",
            "config",
            "mellanox_options.yaml",
        )
        dell_options = yaml.load(open(dell_options_file))
        mellanox_options = yaml.load(open(mellanox_options_file))
        if "sw-cdu" or "sw-leaf-bmc" in node_shasta_name:
            v1_config = ""
            dell_switch = Host(node_shasta_name, "dellOS10", dell_options)
            dell_config_hier = HConfig(host=dell_switch)
            dell_config_hier.load_from_string(switch_config).set_order_weight()
            for line in dell_config_hier.all_children_sorted():
                v1_config = v1_config + line.cisco_style_text() + "\n"
        if "sw-spine" in node_shasta_name:
            v1_config = ""
            mellanox_switch = Host(node_shasta_name, "onyx", mellanox_options)
            mellanox_config_hier = HConfig(host=mellanox_switch)
            mellanox_config_hier.load_from_string(switch_config).set_order_weight()
            for line in mellanox_config_hier.all_children_sorted():
                v1_config = v1_config + line.cisco_style_text() + "\n"
        return v1_config, devices, unknown

    if override:
        options_file = os.path.join(
            project_root,
            "canu",
            "validate",
            "switch",
            "config",
            "options.yaml",
        )
        if switch_name in override:
            options = yaml.load(open(options_file))
            host = Host(switch_name, "aoscx", options)
            override_config = (
                "# OVERRIDE CONFIG\n"
                + "# The configuration below has been ignored and is not included in the GENERATED CONFIG\n"
            )
            override_config_hier = HConfig(host=host)
            override_config_hier.load_from_string(switch_config).add_tags(
                override[switch_name],
            )
            for line in override_config_hier.all_children_sorted_by_tags(
                "override",
                None,
            ):
                override_config = override_config + "\n" + "#" + line.cisco_style_text()
            override_config = override_config + "\n# GENERATED CONFIG\n"
            for line in override_config_hier.all_children_sorted_by_tags(
                None,
                "override",
            ):
                # add two spaces to indented config to match aruba formatting.
                if line.cisco_style_text().startswith("  "):
                    override_config = (
                        override_config + "\n" + "  " + line.cisco_style_text()
                    )
                else:
                    override_config = override_config + "\n" + line.cisco_style_text()

            return override_config, devices, unknown
    return switch_config, devices, unknown


def get_pair_connections(nodes, switch_name):
    """Given a hostname and nodes, return connections to the primary or secondary switch.

    Args:
        nodes: List of nodes connected to the switch
        switch_name: Switch hostname

    Returns:
        List of connections to the paired switch
    """
    is_primary, primary, secondary = switch_is_primary(switch_name)

    if is_primary:
        pair_hostname = secondary
    else:
        pair_hostname = primary

    connections = []
    for x in nodes:
        if pair_hostname in x["config"]["DESCRIPTION"]:
            connections.append(x["config"]["PORT"])

    connections = natsort.natsorted(connections)
    return connections


def get_switch_nodes(switch_name, shcd_node_list, factory, sls_variables):
    """Get the nodes connected to the switch ports.

    Args:
        switch_name: Switch hostname
        shcd_node_list: List of nodes from the SHCD
        factory: Node factory object
        sls_variables: Dictionary containing SLS variables.

    Returns:
        List of nodes connected to the switch
        List of unknown nodes
    """
    nodes = []
    nodes_by_name = {}
    nodes_by_id = {}
    unknown = []

    # Make 2 dictionaries for easy node lookup
    for node in shcd_node_list:
        node_tmp = node.serialize()
        name = node_tmp["common_name"]

        nodes_by_name[name] = node_tmp
        nodes_by_id[node_tmp["id"]] = node_tmp

    if switch_name not in nodes_by_name.keys():
        click.secho(
            f"For switch {switch_name}, the type cannot be determined. Please check the switch name and try again.",
            fg="red",
        )
        exit(1)

    for port in nodes_by_name[switch_name]["ports"]:
        destination_node_id = port["destination_node_id"]
        destination_node_name = nodes_by_id[port["destination_node_id"]]["common_name"]
        destination_rack = nodes_by_id[port["destination_node_id"]]["location"]["rack"]
        source_port = port["port"]
        destination_port = port["destination_port"]
        destination_slot = port["destination_slot"]

        shasta_name = get_shasta_name(destination_node_name, factory.lookup_mapper())

        primary_port = get_primary_port(nodes_by_name, switch_name, destination_node_id)
        if shasta_name == "ncn-m":
            new_node = {
                "subtype": "master",
                "slot": destination_slot,
                "destination_port": destination_port,
                "config": {
                    "DESCRIPTION": f"{switch_name}:{source_port}==>{destination_node_name}:{destination_slot}:{destination_port}",
                    "PORT": f"{source_port}",
                    "LAG_NUMBER": primary_port,
                },
            }
            nodes.append(new_node)
        elif shasta_name == "ncn-s":
            # ncn-s also needs destination_port to find the match
            primary_port_ncn_s = get_primary_port(
                nodes_by_name,
                switch_name,
                destination_node_id,
                destination_port,
            )
            new_node = {
                "subtype": "storage",
                "slot": destination_slot,
                "destination_port": destination_port,
                "config": {
                    "DESCRIPTION": f"{switch_name}:{source_port}==>{destination_node_name}:{destination_slot}:{destination_port}",
                    "PORT": f"{source_port}",
                    "LAG_NUMBER": primary_port_ncn_s,
                    "LAG_NUMBER_V1": primary_port,
                },
            }
            nodes.append(new_node)
        elif shasta_name == "ncn-w":
            new_node = {
                "subtype": "worker",
                "slot": destination_slot,
                "destination_port": destination_port,
                "config": {
                    "DESCRIPTION": f"{switch_name}:{source_port}==>{destination_node_name}:{destination_slot}:{destination_port}",
                    "PORT": f"{source_port}",
                    "LAG_NUMBER": primary_port,
                },
            }
            nodes.append(new_node)
        elif shasta_name == "cec":
            destination_rack_int = int(re.search(r"\d+", destination_rack)[0])
            for cabinets in sls_variables["HMN_MTN_CABINETS"]:
                sls_rack_int = int(re.search(r"\d+", (cabinets["Name"]))[0])
                if destination_rack_int == sls_rack_int:
                    hmn_mtn_vlan = cabinets["VlanID"]
            new_node = {
                "subtype": "cec",
                "slot": None,
                "config": {
                    "DESCRIPTION": f"{switch_name}:{source_port}==>{destination_node_name}:{destination_port}",
                    "INTERFACE_NUMBER": f"{source_port}",
                    "NATIVE_VLAN": hmn_mtn_vlan,
                },
            }
            nodes.append(new_node)
        elif shasta_name == "cmm":
            destination_rack_int = int(re.search(r"\d+", destination_rack)[0])
            for cabinets in sls_variables["NMN_MTN_CABINETS"]:
                sls_rack_int = int(re.search(r"\d+", (cabinets["Name"]))[0])
                if destination_rack_int == sls_rack_int:
                    nmn_mtn_vlan = cabinets["VlanID"]
            for cabinets in sls_variables["HMN_MTN_CABINETS"]:
                sls_rack_int = int(re.search(r"\d+", (cabinets["Name"]))[0])
                if destination_rack_int == sls_rack_int:
                    hmn_mtn_vlan = cabinets["VlanID"]
            new_node = {
                "subtype": "cmm",
                "slot": None,
                "config": {
                    "DESCRIPTION": f"{switch_name}:{source_port}==>{destination_node_name}:{destination_port}",
                    "PORT": f"{source_port}",
                    "LAG_NUMBER": primary_port,
                    "NATIVE_VLAN": nmn_mtn_vlan,
                    "TAGGED_VLAN": hmn_mtn_vlan,
                },
            }
            nodes.append(new_node)
        elif shasta_name == "uan":
            new_node = {
                "subtype": "uan",
                "slot": destination_slot,
                "destination_port": destination_port,
                "config": {
                    "DESCRIPTION": f"{switch_name}:{source_port}==>{destination_node_name}:{destination_slot}:{destination_port}",
                    "PORT": f"{source_port}",
                    "LAG_NUMBER": primary_port,
                },
            }
            nodes.append(new_node)
        elif shasta_name == "sw-spine":
            # sw-leaf ==> sw-spine
            if switch_name.startswith("sw-leaf"):
                is_primary, primary, secondary = switch_is_primary(switch_name)
                digits = re.findall(r"(\d+)", primary)[0]
                lag_number = 100 + int(digits)

            # sw-cdu ==> sw-spine
            elif switch_name.startswith("sw-cdu"):
                lag_number = 255

            # sw-leaf-bmc ==> sw-spine
            elif switch_name.startswith("sw-leaf-bmc"):
                lag_number = 255

            # sw-spine ==> sw-spine
            elif switch_name.startswith("sw-spine"):
                lag_number = 256
            new_node = {
                "subtype": "spine",
                "slot": None,
                "config": {
                    "DESCRIPTION": f"{switch_name}:{source_port}==>{destination_node_name}:{destination_port}",
                    "LAG_NUMBER": lag_number,
                    "PORT": f"{source_port}",
                },
            }
            nodes.append(new_node)
        elif shasta_name == "sw-cdu":
            is_primary, primary, secondary = switch_is_primary(destination_node_name)
            # sw-spine ==> sw-cdu
            if switch_name.startswith("sw-spine"):
                digits = re.findall(r"(\d+)", primary)[0]
                lag_number = 200 + int(digits)

            # sw-cdu ==> sw-cdu
            elif switch_name.startswith("sw-cdu"):
                lag_number = 256
            new_node = {
                "subtype": "cdu",
                "slot": None,
                "primary": is_primary,
                "config": {
                    "DESCRIPTION": f"{switch_name}:{source_port}==>{destination_node_name}:{destination_port}",
                    "LAG_NUMBER": lag_number,
                    "PORT": f"{source_port}",
                },
            }
            nodes.append(new_node)
        elif shasta_name == "sw-leaf":
            # sw-spine ==> sw-leaf
            is_primary, primary, secondary = switch_is_primary(destination_node_name)
            if switch_name.startswith("sw-spine"):
                digits = re.findall(r"(\d+)", primary)[0]
                lag_number = 100 + int(digits)

            # sw-leaf-bmc ==> sw-leaf
            elif switch_name.startswith("sw-leaf-bmc"):
                lag_number = 255

            # sw-leaf ==> sw-leaf
            elif switch_name.startswith("sw-leaf"):
                lag_number = 256
            new_node = {
                "subtype": "leaf",
                "slot": None,
                "primary": is_primary,
                "config": {
                    "DESCRIPTION": f"{switch_name}:{source_port}==>{destination_node_name}:{destination_port}",
                    "LAG_NUMBER": lag_number,
                    "PORT": f"{source_port}",
                },
            }
            nodes.append(new_node)
        elif shasta_name == "sw-leaf-bmc":
            # sw-leaf ==> sw-leaf-bmc
            if switch_name.startswith("sw-leaf"):
                digits = re.findall(r"(\d+)", destination_node_name)[0]
                lag_number = 150 + int(digits)

            # sw-spine ==> sw-leaf-bmc
            elif switch_name.startswith("sw-spine"):
                digits = re.findall(r"(\d+)", destination_node_name)[0]
                lag_number = 150 + int(digits)
            new_node = {
                "subtype": "leaf-bmc",
                "slot": None,
                "config": {
                    "DESCRIPTION": f"{switch_name}:{source_port}==>{destination_node_name}:{destination_port}",
                    "LAG_NUMBER": lag_number,
                    "PORT": f"{source_port}",
                },
            }
            nodes.append(new_node)
        elif shasta_name == "sw-edge":
            pass
        else:  # pragma: no cover
            print("*********************************")
            print("Cannot determine destination connection")
            print("Source: ", switch_name)
            print("Port: ", port)
            print("Destination: ", destination_node_name)
            print("shasta_name", shasta_name)
            print("*********************************")
            unknown_description = f"{switch_name}:{source_port}==>{destination_node_name}:{destination_port}"
            new_node = {
                "subtype": "unknown",
                "slot": None,
                "config": {
                    "DESCRIPTION": unknown_description,
                },
            }
            nodes.append(new_node)
            unknown.append(unknown_description)
    return nodes, unknown


def switch_is_primary(switch):
    """Determine if the switch is primary or secondary.

    Args:
        switch: Switch hostname

    Returns:
        is_primary: Bool if the switch is primary
        primary: Primary switch hostname
        secondary: Secondary switch hostname
    """
    # Determine if PRIMARY or SECONDARY
    digits = re.findall(r"(\d+)", switch)[0]
    middle = re.findall(r"(?:sw-)([a-z-]+)", switch)[0]

    if int(digits) % 2 == 0:  # Switch is Secondary
        is_primary = False
        primary = f"sw-{middle.rstrip('-')}-{int(digits)-1 :03d}"
        secondary = switch
    else:  # Switch is Primary
        is_primary = True
        secondary = f"sw-{middle.rstrip('-')}-{int(digits)+1 :03d}"
        primary = switch

    return is_primary, primary, secondary


def parse_sls_for_config(input_json):
    """Parse the `sls_file.json` file or the JSON from SLS `/networks` API for config variables.

    Args:
        input_json: JSON from the SLS `/networks` API

    Returns:
        sls_variables: Dictionary containing SLS variables.
    """
    networks_list = []

    sls_variables = {
        "CAN": None,
        "CAN_NETMASK": None,
        "CAN_PREFIX_LEN": None,
        "CAN_NETWORK_IP": None,
        "CMN": None,
        "CMN_NETMASK": None,
        "CMN_PREFIX_LEN": None,
        "CMN_NETWORK_IP": None,
        "HMN": None,
        "HMN_NETMASK": None,
        "HMN_NETWORK_IP": None,
        "HMN_PREFIX_LEN": None,
        "MTL": None,
        "MTL_NETMASK": None,
        "MTL_NETWORK_IP": None,
        "MTL_PREFIX_LEN": None,
        "NMN": None,
        "NMN_NETMASK": None,
        "NMN_NETWORK_IP": None,
        "NMN_PREFIX_LEN": None,
        "HMN_MTN": None,
        "HMN_MTN_NETMASK": None,
        "HMN_MTN_NETWORK_IP": None,
        "HMN_MTN_PREFIX_LEN": None,
        "NMN_MTN": None,
        "NMN_MTN_NETMASK": None,
        "NMN_MTN_NETWORK_IP": None,
        "NMN_MTN_PREFIX_LEN": None,
        "HMNLB": None,
        "HMNLB_NETMASK": None,
        "HMNLB_NETWORK_IP": None,
        "HMNLB_PREFIX_LEN": None,
        "NMNLB": None,
        "NMNLB_NETMASK": None,
        "NMNLB_NETWORK_IP": None,
        "NMNLB_PREFIX_LEN": None,
        "CAN_IP_GATEWAY": None,
        "CMN_IP_GATEWAY": None,
        "HMN_IP_GATEWAY": None,
        "MTL_IP_GATEWAY": None,
        "NMN_IP_GATEWAY": None,
        "ncn_w001": None,
        "ncn_w002": None,
        "ncn_w003": None,
        "CAN_IP_PRIMARY": None,
        "CAN_IP_SECONDARY": None,
        "CMN_IP_PRIMARY": None,
        "CMN_IP_SECONDARY": None,
        "HMN_IPs": defaultdict(),
        "MTL_IPs": defaultdict(),
        "NMN_IPs": defaultdict(),
        "NMN_MTN_CABINETS": [],
        "HMN_MTN_CABINETS": [],
    }

    for sls_network in input_json:
        name = sls_network.get("Name", "")

        if name == "CAN":
            sls_variables["CAN"] = netaddr.IPNetwork(
                sls_network.get("ExtraProperties", {}).get(
                    "CIDR",
                    "",
                ),
            )
            sls_variables["CAN_NETMASK"] = sls_variables["CAN"].netmask
            sls_variables["CAN_PREFIX_LEN"] = sls_variables["CAN"].prefixlen
            sls_variables["CAN_NETWORK_IP"] = sls_variables["CAN"].ip
            for subnets in sls_network.get("ExtraProperties", {}).get("Subnets", {}):
                if subnets["Name"] == "bootstrap_dhcp":
                    sls_variables["CAN_IP_GATEWAY"] = subnets["Gateway"]
                    for ip in subnets["IPReservations"]:
                        if ip["Name"] == "can-switch-1":
                            sls_variables["CAN_IP_PRIMARY"] = ip["IPAddress"]
                        elif ip["Name"] == "can-switch-2":
                            sls_variables["CAN_IP_SECONDARY"] = ip["IPAddress"]

        elif name == "CMN":
            sls_variables["CMN"] = netaddr.IPNetwork(
                sls_network.get("ExtraProperties", {}).get(
                    "CIDR",
                    "",
                ),
            )
            sls_variables["CMN_NETMASK"] = sls_variables["CMN"].netmask
            sls_variables["CMN_PREFIX_LEN"] = sls_variables["CMN"].prefixlen
            sls_variables["CMN_NETWORK_IP"] = sls_variables["CMN"].ip
            for subnets in sls_network.get("ExtraProperties", {}).get("Subnets", {}):
                if subnets["Name"] == "bootstrap_dhcp":
                    sls_variables["CMN_IP_GATEWAY"] = subnets["Gateway"]
                    for ip in subnets["IPReservations"]:
                        if ip["Name"] == "cmn-switch-1":
                            sls_variables["CMN_IP_PRIMARY"] = ip["IPAddress"]
                        elif ip["Name"] == "cmn-switch-2":
                            sls_variables["CMN_IP_SECONDARY"] = ip["IPAddress"]

        elif name == "HMN":
            sls_variables["HMN"] = netaddr.IPNetwork(
                sls_network.get("ExtraProperties", {}).get(
                    "CIDR",
                    "",
                ),
            )
            sls_variables["HMN_NETMASK"] = sls_variables["HMN"].netmask
            sls_variables["HMN_PREFIX_LEN"] = sls_variables["HMN"].prefixlen
            sls_variables["HMN_NETWORK_IP"] = sls_variables["HMN"].ip
            for subnets in sls_network.get("ExtraProperties", {}).get("Subnets", {}):
                if subnets["Name"] == "network_hardware":
                    sls_variables["HMN_IP_GATEWAY"] = subnets["Gateway"]
                    for ip in subnets["IPReservations"]:
                        sls_variables["HMN_IPs"][ip["Name"]] = ip["IPAddress"]

        elif name == "MTL":
            sls_variables["MTL"] = netaddr.IPNetwork(
                sls_network.get("ExtraProperties", {}).get(
                    "CIDR",
                    "",
                ),
            )
            sls_variables["MTL_NETMASK"] = sls_variables["MTL"].netmask
            sls_variables["MTL_PREFIX_LEN"] = sls_variables["MTL"].prefixlen
            sls_variables["MTL_NETWORK_IP"] = sls_variables["MTL"].ip
            for subnets in sls_network.get("ExtraProperties", {}).get("Subnets", {}):
                if subnets["Name"] == "network_hardware":
                    sls_variables["MTL_IP_GATEWAY"] = subnets["Gateway"]
                    for ip in subnets["IPReservations"]:
                        sls_variables["MTL_IPs"][ip["Name"]] = ip["IPAddress"]

        elif name == "NMN":
            sls_variables["NMN"] = netaddr.IPNetwork(
                sls_network.get("ExtraProperties", {}).get(
                    "CIDR",
                    "",
                ),
            )
            sls_variables["NMN_NETMASK"] = sls_variables["NMN"].netmask
            sls_variables["NMN_PREFIX_LEN"] = sls_variables["NMN"].prefixlen
            sls_variables["NMN_NETWORK_IP"] = sls_variables["NMN"].ip
            for subnets in sls_network.get("ExtraProperties", {}).get("Subnets", {}):
                if subnets["Name"] == "bootstrap_dhcp":
                    for ip in subnets["IPReservations"]:
                        if ip["Name"] == "ncn-w001":
                            sls_variables["ncn_w001"] = ip["IPAddress"]
                        elif ip["Name"] == "ncn-w002":
                            sls_variables["ncn_w002"] = ip["IPAddress"]
                        elif ip["Name"] == "ncn-w003":
                            sls_variables["ncn_w003"] = ip["IPAddress"]
                elif subnets["Name"] == "network_hardware":
                    sls_variables["NMN_IP_GATEWAY"] = subnets["Gateway"]
                    for ip in subnets["IPReservations"]:
                        sls_variables["NMN_IPs"][ip["Name"]] = ip["IPAddress"]

        elif name == "NMN_MTN":
            sls_variables["NMN_MTN"] = netaddr.IPNetwork(
                sls_network.get("ExtraProperties", {}).get(
                    "CIDR",
                    "",
                ),
            )
            sls_variables["NMN_MTN_NETMASK"] = sls_variables["NMN_MTN"].netmask
            sls_variables["NMN_MTN_PREFIX_LEN"] = sls_variables["NMN_MTN"].prefixlen
            sls_variables["NMN_MTN_NETWORK_IP"] = sls_variables["NMN_MTN"].ip
            sls_variables["NMN_MTN_CABINETS"] = list(
                sls_network.get("ExtraProperties", {}).get("Subnets", {}),
            )

        elif name == "HMN_MTN":
            sls_variables["HMN_MTN"] = netaddr.IPNetwork(
                sls_network.get("ExtraProperties", {}).get(
                    "CIDR",
                    "",
                ),
            )
            sls_variables["HMN_MTN_NETMASK"] = sls_variables["HMN_MTN"].netmask
            sls_variables["HMN_MTN_PREFIX_LEN"] = sls_variables["HMN_MTN"].prefixlen
            sls_variables["HMN_MTN_NETWORK_IP"] = sls_variables["HMN_MTN"].ip
            sls_variables["HMN_MTN_CABINETS"] = list(
                sls_network.get("ExtraProperties", {}).get("Subnets", {}),
            )
        elif name == "HMNLB":
            sls_variables["HMNLB"] = netaddr.IPNetwork(
                sls_network.get("ExtraProperties", {}).get(
                    "CIDR",
                    "",
                ),
            )
            sls_variables["HMNLB_NETMASK"] = sls_variables["HMNLB"].netmask
            sls_variables["HMNLB_PREFIX_LEN"] = sls_variables["HMNLB"].prefixlen
            sls_variables["HMNLB_NETWORK_IP"] = sls_variables["HMNLB"].ip
            sls_variables["HMNLB_CABINETS"] = list(
                sls_network.get("ExtraProperties", {}).get("Subnets", {}),
            )
        elif name == "NMNLB":
            sls_variables["NMNLB"] = netaddr.IPNetwork(
                sls_network.get("ExtraProperties", {}).get(
                    "CIDR",
                    "",
                ),
            )
            sls_variables["NMNLB_NETMASK"] = sls_variables["NMNLB"].netmask
            sls_variables["NMNLB_PREFIX_LEN"] = sls_variables["NMNLB"].prefixlen
            sls_variables["NMNLB_NETWORK_IP"] = sls_variables["NMNLB"].ip
            sls_variables["NMNLB_CABINETS"] = list(
                sls_network.get("ExtraProperties", {}).get("Subnets", {}),
            )
        for subnets in sls_network.get("ExtraProperties", {}).get("Subnets", {}):

            vlan = subnets.get("VlanID", "")
            networks_list.append([name, vlan])

    networks_list = {tuple(x) for x in networks_list}
    return sls_variables


def rename_sls_hostnames(sls_variables):
    """Parse and rename SLS switch names.

    The operation needs to be done in two passes to prevent naming conflicts.

    Args:
        sls_variables: Dictionary containing SLS variables.

    Returns:
        sls_variables: Dictionary containing renamed SLS variables.
    """
    # First pass rename leaf ==> leaf-bmc
    for key, value in sls_variables["HMN_IPs"].copy().items():
        new_name = key.replace("-leaf-", "-leaf-bmc-")
        sls_variables["HMN_IPs"].pop(key)
        sls_variables["HMN_IPs"][new_name] = value

    for key, value in sls_variables["MTL_IPs"].copy().items():
        new_name = key.replace("-leaf-", "-leaf-bmc-")
        sls_variables["MTL_IPs"].pop(key)
        sls_variables["MTL_IPs"][new_name] = value

    for key, value in sls_variables["NMN_IPs"].copy().items():
        new_name = key.replace("-leaf-", "-leaf-bmc-")
        sls_variables["NMN_IPs"].pop(key)
        sls_variables["NMN_IPs"][new_name] = value

    # Second pass rename agg ==> leaf
    for key, value in sls_variables["HMN_IPs"].copy().items():
        new_name = key.replace("-agg-", "-leaf-")
        sls_variables["HMN_IPs"].pop(key)
        sls_variables["HMN_IPs"][new_name] = value

    for key, value in sls_variables["MTL_IPs"].copy().items():
        new_name = key.replace("-agg-", "-leaf-")
        sls_variables["MTL_IPs"].pop(key)
        sls_variables["MTL_IPs"][new_name] = value

    for key, value in sls_variables["NMN_IPs"].copy().items():
        new_name = key.replace("-agg-", "-leaf-")
        sls_variables["NMN_IPs"].pop(key)
        sls_variables["NMN_IPs"][new_name] = value

    return sls_variables


def get_primary_port(
    nodes_by_name,
    switch_name,
    destination_node_id,
    destination_port=None,
):
    """Return the primary switch port number for a connection to a node.

    Args:
        nodes_by_name: Dictionary containing the node list where hostname is the key
        switch_name: Switch hostname
        destination_node_id: The node id of the destination device.
        destination_port: (Optional, only used with ncn-s) The destination port

    Returns:
        port: Port number of the primary connection to a device.
    """
    is_primary, primary, secondary = switch_is_primary(switch_name)

    for y in nodes_by_name[primary]["ports"]:
        if y["destination_node_id"] == destination_node_id:
            if not destination_port:
                return y["port"]
            # Since ncn-s can have multiple connections to a device, returns the correct one
            elif destination_port and y["destination_port"] == destination_port:
                return y["port"]
