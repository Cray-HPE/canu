# MIT License
#
# (C) Copyright [2021] Hewlett Packard Enterprise Development LP
#
# Permission is hereby granted, free of charge, to any person obtaining a
# copy of this software and associated documentation files (the "Software"),
# to deal in the Software without restriction, including without limitation
# the rights to use, copy, modify, merge, publish, distribute, sublicense,
# and/or sell copies of the Software, and to permit persons to whom the
# Software is furnished to do so, subject to the following conditions:
#
# The above copyright notice and this permission notice shall be included
# in all copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL
# THE AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR
# OTHER LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE,
# ARISING FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR
# OTHER DEALINGS IN THE SOFTWARE.
"""CANU commands that generate the config of the entire Shasta network."""
import json
import os
from os import environ, makedirs, path
from pathlib import Path
import sys

import click
from click_help_colors import HelpColorsCommand
from jinja2 import Environment, FileSystemLoader, StrictUndefined
from network_modeling.NetworkNodeFactory import NetworkNodeFactory
from openpyxl import load_workbook
import requests
from ruamel.yaml import YAML
import urllib3

from canu.generate.switch.config.config import (
    generate_switch_config,
    get_shasta_name,
    parse_sls_for_config,
    rename_sls_hostnames,
)
from canu.utils.cache import cache_directory
from canu.validate.shcd.shcd import node_model_from_shcd

yaml = YAML()

# To disable warnings about unsecured HTTPS requests
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


# Get project root directory
if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):  # pragma: no cover
    project_root = sys._MEIPASS
else:
    prog = __file__
    project_root = Path(__file__).resolve().parent.parent.parent.parent.parent

# Schema and Data files
hardware_schema_file = path.join(
    project_root,
    "network_modeling",
    "schema",
    "cray-network-hardware-schema.yaml",
)
hardware_spec_file = path.join(
    project_root,
    "network_modeling",
    "models",
    "cray-network-hardware.yaml",
)
architecture_schema_file = path.join(
    project_root,
    "network_modeling",
    "schema",
    "cray-network-architecture-schema.yaml",
)
architecture_spec_file = path.join(
    project_root,
    "network_modeling",
    "models",
    "cray-network-architecture.yaml",
)

canu_cache_file = path.join(cache_directory(), "canu_cache.yaml")
canu_config_file = path.join(project_root, "canu", "canu.yaml")

# Import templates
network_templates_folder = path.join(
    project_root,
    "network_modeling",
    "configs",
    "templates",
)
env = Environment(
    loader=FileSystemLoader(network_templates_folder),
    undefined=StrictUndefined,
)

# Get CSM versions from canu.yaml
with open(canu_config_file, "r") as file:
    canu_config = yaml.load(file)

csm_options = canu_config["csm_versions"]


@click.command(
    cls=HelpColorsCommand,
    help_headers_color="yellow",
    help_options_color="blue",
)
@click.option(
    "--csm",
    type=click.Choice(csm_options),
    help="CSM network version",
    prompt="CSM network version",
    required=True,
    show_choices=True,
)
@click.option(
    "--architecture",
    "-a",
    type=click.Choice(["Full", "TDS", "V1"], case_sensitive=False),
    help="CSM architecture",
    required=True,
    prompt="Architecture type",
)
@click.option(
    "--shcd",
    help="SHCD file",
    type=click.File("rb"),
    required=True,
)
@click.option(
    "--tabs",
    help="The tabs on the SHCD file to check, e.g. 10G_25G_40G_100G,NMN,HMN.",
)
@click.option(
    "--corners",
    help="The corners on each tab, comma separated e.g. 'J37,U227,J15,T47,J20,U167'.",
)
@click.option(
    "--sls-file",
    help="File containing system SLS JSON data.",
    type=click.File("r"),
)
@click.option(
    "--auth-token",
    envvar="SLS_TOKEN",
    help="Token for SLS authentication",
)
@click.option("--sls-address", default="api-gw-service-nmn.local", show_default=True)
@click.option(
    "--folder",
    help="Folder to store config files",
    required=True,
    prompt="Folder for configs",
)
@click.option(
    "--override",
    help="Switch configuration override",
    type=click.Path(),
)
@click.pass_context
def config(
    ctx,
    csm,
    architecture,
    shcd,
    tabs,
    corners,
    sls_file,
    auth_token,
    sls_address,
    folder,
    override,
):
    """Generate the config of all switches (Aruba, Dell, or Mellanox) on the network using the SHCD.

    In order to generate network switch config, a valid SHCD must be passed in and system variables must be read in from either
    an SLS output file or the SLS API.

    Use the `--folder FOLDERNAME` flag to output all the switch configs to a folder.
    \f
    # noqa: D301

    Args:
        ctx: CANU context settings
        csm: CSM version
        architecture: CSM architecture
        shcd: SHCD file
        tabs: The tabs on the SHCD file to check, e.g. 10G_25G_40G_100G,NMN,HMN.
        corners: The corners on each tab, comma separated e.g. 'J37,U227,J15,T47,J20,U167'.
        sls_file: Directory containing the CSI json file
        auth_token: Token for SLS authentication
        sls_address: The address of SLS
        folder: Folder to store config files
        override: Input file that defines what config should be ignored
    """
    if architecture.lower() == "full":
        architecture = "network_v2"
        template_folder = "full"
        vendor_folder = "aruba"
    elif architecture.lower() == "tds":
        architecture = "network_v2_tds"
        template_folder = "tds"
        vendor_folder = "aruba"
    elif architecture.lower() == "v1":
        architecture = "network_v1"
        template_folder = "full"
        vendor_folder = "dellmellanox"

    # SHCD Parsing
    sheets = []

    if not tabs:
        wb = load_workbook(shcd, read_only=True)
        click.secho("What tabs would you like to check in the SHCD?")
        tab_options = wb.sheetnames
        for x in tab_options:
            click.secho(f"{x}", fg="green")

        tabs = click.prompt(
            "Please enter the tabs to check separated by a comma, e.g. 10G_25G_40G_100G,NMN,HMN.",
            type=str,
        )

    if corners:
        if len(tabs.split(",")) * 2 != len(corners.split(",")):
            click.secho("Not enough corners.\n", fg="red")
            click.secho(
                f"Make sure each tab: {tabs.split(',')} has 2 corners.\n",
                fg="red",
            )
            click.secho(
                f"There were {len(corners.split(','))} corners entered, but there should be {len(tabs.split(',')) * 2}.",
                fg="red",
            )
            click.secho(
                f"{corners}\n",
                fg="red",
            )
            return

        # Each tab should have 2 corners entered in comma separated
        for i in range(len(tabs.split(","))):
            # 0 -> 0,1
            # 1 -> 2,3
            # 2 -> 4,5

            sheets.append(
                (
                    tabs.split(",")[i],
                    corners.split(",")[i * 2].strip(),
                    corners.split(",")[i * 2 + 1].strip(),
                ),
            )
    else:
        for tab in tabs.split(","):
            click.secho(f"\nFor the Sheet {tab}", fg="green")
            range_start = click.prompt(
                "Enter the cell of the upper left corner (Labeled 'Source')",
                type=str,
            )
            range_end = click.prompt(
                "Enter the cell of the lower right corner",
                type=str,
            )
            sheets.append((tab, range_start, range_end))

    # Create Node factory
    factory = NetworkNodeFactory(
        hardware_schema=hardware_schema_file,
        hardware_data=hardware_spec_file,
        architecture_schema=architecture_schema_file,
        architecture_data=architecture_spec_file,
        architecture_version=architecture,
    )

    # Get nodes from SHCD
    shcd_node_list, shcd_warnings = node_model_from_shcd(
        factory=factory,
        spreadsheet=shcd,
        sheets=sheets,
    )

    # Parse SLS input file.
    if sls_file:
        try:
            input_json = json.load(sls_file)
        except (json.JSONDecodeError, UnicodeDecodeError):
            click.secho(
                f"The file {sls_file.name} is not valid JSON.",
                fg="red",
            )
            return

        # Format the input to be like the SLS JSON
        sls_json = [
            network[x] for network in [input_json.get("Networks", {})] for x in network
        ]

    else:
        # Get SLS config
        token = environ.get("SLS_TOKEN")

        # Token file takes precedence over the environmental variable
        if auth_token != token:
            try:
                with open(auth_token) as f:
                    data = json.load(f)
                    token = data["access_token"]

            except Exception:
                click.secho(
                    "Invalid token file, generate another token or try again.",
                    fg="white",
                    bg="red",
                )
                return

        # SLS
        url = "https://" + sls_address + "/apis/sls/v1/networks"
        try:
            response = requests.get(
                url,
                headers={
                    "Content-Type": "application/json",
                    "Authorization": f"Bearer {token}",
                },
                verify=False,
            )
            response.raise_for_status()

            sls_json = response.json()

        except requests.exceptions.ConnectionError:
            return click.secho(
                f"Error connecting to SLS {sls_address}, check the address or pass in a new address using --sls-address.",
                fg="white",
                bg="red",
            )
        except requests.exceptions.HTTPError:
            bad_token_reason = (
                "environmental variable 'SLS_TOKEN' is correct."
                if auth_token == token
                else "token is valid, or generate a new one."
            )
            return click.secho(
                f"Error connecting SLS {sls_address}, check that the {bad_token_reason}",
                fg="white",
                bg="red",
            )
    sls_variables = parse_sls_for_config(sls_json)

    # For versions of csm < 1.2, the SLS Hostnames need to be renamed
    if csm:
        if float(csm) < 1.2:
            sls_variables = rename_sls_hostnames(sls_variables)

    if override:
        try:
            with open(os.path.join(override), "r") as f:
                override = yaml.load(f)
        except FileNotFoundError:
            click.secho(
                "The override yaml file was not found, check that you entered the right file name and path.",
                fg="red",
            )
            exit(1)

    # make folder
    if not path.exists(folder):
        makedirs(folder)
    all_devices = {
        "cdu",
        "cec",
        "cmm",
        "leaf",
        "leaf-bmc",
        "master",
        "spine",
        "storage",
        "uan",
        "worker",
    }
    config_devices = set()
    all_unknown = []
    for node in shcd_node_list:
        switch_name = node.common_name()
        node_shasta_name = get_shasta_name(switch_name, factory.lookup_mapper())

        if node_shasta_name in ["sw-cdu", "sw-leaf-bmc", "sw-leaf", "sw-spine"]:

            switch_config, devices, unknown = generate_switch_config(
                csm,
                architecture,
                shcd_node_list,
                factory,
                switch_name,
                sls_variables,
                template_folder,
                vendor_folder,
                override,
            )
            all_unknown.extend(unknown)
            config_devices.update(devices)
            with open(f"{folder}/{switch_name}.cfg", "w+") as f:
                f.write(switch_config)
            if "# OVERRIDE" in switch_config:
                click.secho(f"{switch_name} Override Config Generated", fg="yellow")
            else:
                click.secho(f"{switch_name} Config Generated", fg="bright_white")
    missing_devices = all_devices.difference(config_devices)
    dash = "-" * 60
    if len(missing_devices) > 0:
        click.secho("\nWarning", fg="red")

        click.secho(
            "\nThe following devices typically exist, but were not found in the current configuration.",
            fg="red",
        )
        click.secho(
            "If the network is supposed to have these devices, check the SHCD to ensure all tabs were included.",
            fg="red",
        )
        click.secho(
            "If the network does not contain these devices, disregard this warning.",
        )
        click.secho(dash)
        for x in missing_devices:
            click.secho(x, fg="bright_white")

    if len(all_unknown) > 0:
        click.secho("\nWarning", fg="red")

        click.secho(
            "\nThe following devices were discovered in the input data, but the CANU model cannot determine "
            + "the type and generate a configuration.\nApplying this configuration without considering these "
            + "devices will likely result in loss of contact with these devices."
            + "\nEnsure valid input, submit a bug to CANU and manually add these devices to the configuration.",
            fg="red",
        )
        click.secho(dash)
        for x in all_unknown:
            click.secho(x, fg="bright_white")

    return
