# MIT License
#
# (C) Copyright [2022] Hewlett Packard Enterprise Development LP
#
# Permission is hereby granted, free of charge, to any person obtaining a
# copy of this software and associated documentation files (the "Software"),
# to deal in the Software without restriction, including without limitation
# the rights to use, copy, modify, merge, publish, distribute, sublicense,
# and/or sell copies of the Software, and to permit persons to whom the
# Software is furnished to do so, subject to the following conditions:
#
# The above copyright notice and this permission notice shall be included
# in all copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL
# THE AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR
# OTHER LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE,
# ARISING FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR
# OTHER DEALINGS IN THE SOFTWARE.
"""Test CANU validate shcd commands."""
import json
from os import path
from pathlib import Path

from click import testing
from openpyxl import Workbook

from canu.cli import cli

test_file_directory = Path(__file__).resolve().parent


test_file = "test_file.xlsx"
architecture = "tds"
tabs = "25G_10G"
corners = "I14,S48"
cache_minutes = 0
runner = testing.CliRunner()


def test_validate_shcd():
    """Test that the `canu validate shcd` command runs and returns valid cabling."""
    with runner.isolated_filesystem():
        generate_test_file(test_file)
        result = runner.invoke(
            cli,
            [
                "--cache",
                cache_minutes,
                "validate",
                "shcd",
                "--architecture",
                architecture,
                "--shcd",
                test_file,
                "--tabs",
                tabs,
                "--corners",
                corners,
            ],
        )
        assert result.exit_code == 0
        assert "sw-spine-002 connects to 17 nodes:" in str(result.output)


def test_validate_shcd_full():
    """Test that the `canu validate shcd` command runs and returns valid cabling with '--architecture full' flag."""
    full_architecture = "full"
    with runner.isolated_filesystem():
        generate_test_file(test_file)
        result = runner.invoke(
            cli,
            [
                "--cache",
                cache_minutes,
                "validate",
                "shcd",
                "--architecture",
                full_architecture,
                "--shcd",
                test_file,
                "--tabs",
                tabs,
                "--corners",
                corners,
            ],
        )
        assert result.exit_code == 0
        assert "sw-leaf-002 connects to 17 nodes:" in str(result.output)


def test_validate_shcd_missing_file():
    """Test that the `canu validate shcd` command fails on missing file."""
    with runner.isolated_filesystem():

        result = runner.invoke(
            cli,
            [
                "--cache",
                cache_minutes,
                "validate",
                "shcd",
                "--architecture",
                architecture,
                "--tabs",
                tabs,
                "--corners",
                corners,
            ],
        )
        assert result.exit_code == 2
        assert "Error: Missing option '--shcd'." in str(result.output)


def test_validate_shcd_bad_file():
    """Test that the `canu validate shcd` command fails on bad file."""
    bad_file = "does_not_exist.xlsx"
    with runner.isolated_filesystem():

        result = runner.invoke(
            cli,
            [
                "--cache",
                cache_minutes,
                "validate",
                "shcd",
                "--architecture",
                architecture,
                "--shcd",
                bad_file,
                "--tabs",
                tabs,
                "--corners",
                corners,
            ],
        )
        assert result.exit_code == 2
        assert "Error: Invalid value for '--shcd':" in str(result.output)


def test_validate_shcd_missing_tabs():
    """Test that the `canu validate shcd` command prompts for missing tabs."""
    with runner.isolated_filesystem():
        generate_test_file(test_file)
        result = runner.invoke(
            cli,
            [
                "--cache",
                cache_minutes,
                "validate",
                "shcd",
                "--architecture",
                architecture,
                "--shcd",
                test_file,
                "--corners",
                corners,
            ],
            input="25G_10G\n",
        )
        assert result.exit_code == 0
        assert "sw-spine-001 connects to 18 nodes:" in str(result.output)


def test_validate_shcd_bad_tab():
    """Test that the `canu validate shcd` command fails on bad tab name."""
    bad_tab = "NMN"
    with runner.isolated_filesystem():
        generate_test_file(test_file)
        result = runner.invoke(
            cli,
            [
                "--cache",
                cache_minutes,
                "validate",
                "shcd",
                "--architecture",
                architecture,
                "--shcd",
                test_file,
                "--tabs",
                bad_tab,
                "--corners",
                corners,
            ],
        )
        assert result.exit_code == 1
        assert "Tab NMN not found in test_file.xlsx" in str(result.output)


def test_validate_shcd_corner_prompt():
    """Test that the `canu validate shcd` command prompts for corner input and runs."""
    with runner.isolated_filesystem():
        generate_test_file(test_file)
        result = runner.invoke(
            cli,
            [
                "--cache",
                cache_minutes,
                "validate",
                "shcd",
                "--architecture",
                architecture,
                "--shcd",
                test_file,
                "--tabs",
                tabs,
            ],
            input="I14\nS48",
        )
        assert result.exit_code == 0
        assert "sw-spine-002 connects to 17 nodes:" in str(result.output)


def test_validate_shcd_corners_too_narrow():
    """Test that the `canu validate shcd` command fails on too narrow area."""
    corners_too_narrow = "I16,P48"
    with runner.isolated_filesystem():
        generate_test_file(test_file)
        result = runner.invoke(
            cli,
            [
                "--cache",
                cache_minutes,
                "validate",
                "shcd",
                "--architecture",
                architecture,
                "--shcd",
                test_file,
                "--tabs",
                tabs,
                "--corners",
                corners_too_narrow,
            ],
        )
        assert result.exit_code == 1
        assert "Not enough columns exist." in str(result.output)


def test_validate_shcd_corners_too_high():
    """Test that the `canu validate shcd` command fails on empty headers."""
    corners_too_high = "H16,S48"
    with runner.isolated_filesystem():
        generate_test_file(test_file)
        result = runner.invoke(
            cli,
            [
                "--cache",
                cache_minutes,
                "validate",
                "shcd",
                "--architecture",
                architecture,
                "--shcd",
                test_file,
                "--tabs",
                tabs,
                "--corners",
                corners_too_high,
            ],
        )
        assert result.exit_code == 1
        assert "On tab 25G_10G, header column Source not found." in str(result.output)
        assert "On tab 25G_10G, the header is formatted incorrectly." in str(
            result.output,
        )


def test_validate_shcd_corners_bad_cell():
    """Test that the `canu validate shcd` command fails on bad cell."""
    corners_bad_cell = "16,S48"
    with runner.isolated_filesystem():
        generate_test_file(test_file)
        result = runner.invoke(
            cli,
            [
                "--cache",
                cache_minutes,
                "validate",
                "shcd",
                "--architecture",
                architecture,
                "--shcd",
                test_file,
                "--tabs",
                tabs,
                "--corners",
                corners_bad_cell,
            ],
        )
        assert result.exit_code == 1
        assert "Bad range of cells entered for tab 25G_10G." in str(result.output)


def test_validate_shcd_not_enough_corners():
    """Test that the `canu validate shcd` command fails on not enough corners."""
    not_enough_corners = "H16"
    with runner.isolated_filesystem():
        generate_test_file(test_file)
        result = runner.invoke(
            cli,
            [
                "--cache",
                cache_minutes,
                "validate",
                "shcd",
                "--architecture",
                architecture,
                "--shcd",
                test_file,
                "--tabs",
                tabs,
                "--corners",
                not_enough_corners,
            ],
        )
        assert result.exit_code == 0
        assert "There were 1 corners entered, but there should be 2." in str(
            result.output,
        )


def test_validate_shcd_bad_headers():
    """Test that the `canu validate shcd` command fails on bad headers."""
    bad_header_tab = "Bad_Headers"
    with runner.isolated_filesystem():
        generate_test_file(test_file)
        result = runner.invoke(
            cli,
            [
                "--cache",
                cache_minutes,
                "validate",
                "shcd",
                "--architecture",
                architecture,
                "--shcd",
                test_file,
                "--tabs",
                bad_header_tab,
                "--corners",
                corners,
            ],
        )
        assert result.exit_code == 1
        assert "On tab Bad_Headers, header column Slot not found" in str(result.output)


def test_validate_shcd_bad_architectural_definition():
    """Test that the `canu validate shcd` command fails with bad connections."""
    corners_bad_row = "I14,S50"
    with runner.isolated_filesystem():
        generate_test_file(test_file)
        result = runner.invoke(
            cli,
            [
                "--cache",
                cache_minutes,
                "validate",
                "shcd",
                "--architecture",
                architecture,
                "--shcd",
                test_file,
                "--tabs",
                tabs,
                "--corners",
                corners_bad_row,
            ],
        )
        assert result.exit_code == 1
        assert "No architectural definition found to allow connection between" in str(
            result.output,
        )


def test_validate_shcd_port_reuse():
    """Test that the `canu validate shcd` command fails when a port is used multiple times."""
    multiple_connections_tab = "More_connections"
    multiple_connections_corners = "I14,S20"
    with runner.isolated_filesystem():
        generate_test_file(test_file)
        result = runner.invoke(
            cli,
            [
                "--cache",
                cache_minutes,
                "validate",
                "shcd",
                "--architecture",
                architecture,
                "--shcd",
                test_file,
                "--tabs",
                multiple_connections_tab,
                "--corners",
                multiple_connections_corners,
                "--log",
                "DEBUG",
            ],
        )
        assert result.exit_code == 1
        assert "Failed to connect sw-spine-001 to sw-spine-002 bi-directionally" in str(
            result.output,
        )


def test_validate_shcd_json():
    """Test that the `canu validate shcd` command runs and returns valid JSON cabling."""
    with runner.isolated_filesystem():
        generate_test_file(test_file)
        result = runner.invoke(
            cli,
            [
                "--cache",
                cache_minutes,
                "validate",
                "shcd",
                "--architecture",
                architecture,
                "--shcd",
                test_file,
                "--tabs",
                tabs,
                "--corners",
                corners,
                "--json",
            ],
        )
        assert result.exit_code == 0

        result_json = json.loads(result.output)

        assert result_json["topology"][-1]["common_name"] == "sw-cdu-002"
        assert result_json["topology"][-1]["architecture"] == "mountain_compute_leaf"
        assert result_json["topology"][-1]["location"]["rack"] == "x3000"
        assert result_json["topology"][-1]["location"]["elevation"] == "u13"


def test_validate_shcd_vi():
    """Test that the `canu validate shcd` command runs and returns valid cabling for Dell and Mellanox V1 arch."""
    architecture_v1 = "v1"
    test_file_name_v1 = "Architecture_Golden_Config_Dellanox.xlsx"
    test_file_v1 = path.join(test_file_directory, "data", test_file_name_v1)
    tabs_v1 = "SWITCH_TO_SWITCH,NON_COMPUTE_NODES,HARDWARE_MANAGEMENT,COMPUTE_NODES"
    corners_v1 = "J14,T30,J14,T34,J14,T28,J14,T27"

    with runner.isolated_filesystem():
        generate_test_file(test_file)
        result = runner.invoke(
            cli,
            [
                "--cache",
                cache_minutes,
                "validate",
                "shcd",
                "--architecture",
                architecture_v1,
                "--shcd",
                test_file_v1,
                "--tabs",
                tabs_v1,
                "--corners",
                corners_v1,
            ],
        )
        assert result.exit_code == 0
        assert "sw-spine-002 connects to 18 nodes:" in str(result.output)


def test_validate_shcd_subrack():
    """Test that the `canu validate shcd` command runs and returns valid cabling including SubRack."""
    tabs_subrack = "25G_10G,HMN"
    corners_subrack = "I14,S48,I14,T19"
    full_architecture = "full"
    with runner.isolated_filesystem():
        generate_test_file(test_file)
        result = runner.invoke(
            cli,
            [
                "--cache",
                cache_minutes,
                "validate",
                "shcd",
                "--architecture",
                full_architecture,
                "--shcd",
                test_file,
                "--tabs",
                tabs_subrack,
                "--corners",
                corners_subrack,
                "--log",
                "DEBUG",
            ],
        )
        assert result.exit_code == 0
        assert "15: cn008 connects to 1 nodes: [16]" in str(result.output)
        assert "16: SubRack002-RCM connects to 2 nodes: [15, 17]" in str(result.output)
        assert "17: sw-leaf-bmc-002 connects to 2 nodes: [18, 16]" in str(result.output)
        assert "18: cn004 connects to 2 nodes: [19, 17]" in str(result.output)
        assert "19: SubRack001-RCM connects to 1 nodes: [18]" in str(result.output)


def generate_test_file(file_name):
    """Generate xlsx sheet for testing."""
    wb = Workbook()
    test_file = file_name
    ws1 = wb.active
    ws1.title = "25G_10G"
    ws1["I14"] = "Source"
    ws1["J14"] = "Rack"
    ws1["K14"] = "Location"
    ws1["L14"] = "Slot"
    # None
    ws1["N14"] = "Port"
    ws1["O14"] = "Destination"
    ws1["P14"] = "Rack"
    ws1["Q14"] = "Location"
    # None
    ws1["S14"] = "Port"

    test_data = [
        [
            "sw-25g01",
            "x3000",
            "u12",
            "",
            "-",
            "51",
            "sw-25g02",
            "x3000",
            "u13",
            "-",
            "51",
        ],
        [
            "sw-25g01",
            "x3000",
            "u12",
            "",
            "-",
            "j52",
            "sw-25g02",
            "x3000",
            "u13",
            "-",
            "52",
        ],
        [
            "sw-smn01",
            "x3000",
            "U14",
            "",
            "-",
            "49",
            "sw-25g01",
            "x3000",
            "u12",
            "-",
            "48",
        ],
        [
            "sw-smn01",
            "x3000",
            "U14",
            "",
            "",
            "50",
            "sw-25g02",
            "x3000",
            "u13",
            "-",
            "48",
        ],
        [
            "sw-25g01",
            "x3000",
            "u12",
            "",
            "-",
            "47",
            "sw-25g02",
            "x3000",
            "u13",
            "-",
            "47",
        ],
        [
            "uan01",
            "x3000",
            "u19",
            "ocp",
            "-",
            "1",
            "sw-25g01",
            "x3000",
            "u12",
            "-",
            "16",
        ],
        [
            "uan01",
            "x3000",
            "u19",
            "ocp",
            "-",
            "2",
            "sw-25g01",
            "x3000",
            "u12",
            "-",
            "17",
        ],
        [
            "uan01",
            "x3000",
            "u19",
            "pci",
            "-",
            "1",
            "sw-25g02",
            "x3000",
            "u13",
            "-",
            "16",
        ],
        [
            "uan01",
            "x3000",
            "u19",
            "s1",
            "-",
            "2",
            "sw-25g02",
            "x3000",
            "u13",
            "-",
            "17",
        ],
        [
            "sn03",
            "x3000",
            "u09",
            "ocp",
            "-",
            "1",
            "sw-25g01",
            "x3000",
            "u12",
            "-",
            "15",
        ],
        [
            "sn03",
            "x3000",
            "u09",
            "ocp",
            "-",
            "2	",
            "sw-25g02",
            "x3000",
            "u13",
            "-",
            "15",
        ],
        [
            "sn03",
            "x3000",
            "u09",
            "s1	",
            "-",
            "1",
            "sw-25g01",
            "x3000",
            "u12",
            "-",
            "14",
        ],
        [
            "sn03",
            "x3000",
            "u09",
            "s1	",
            "-",
            "2",
            "sw-25g02",
            "x3000",
            "u13",
            "-",
            "14",
        ],
        [
            "sn02",
            "x3000",
            "u08",
            "ocp",
            "-",
            "1",
            "sw-25g01",
            "x3000",
            "u12",
            "-",
            "13",
        ],
        [
            "sn02",
            "x3000",
            "u08",
            "ocp",
            "-",
            "2	",
            "sw-25g02",
            "x3000",
            "u13",
            "-",
            "13",
        ],
        [
            "sn02",
            "x3000",
            "u08",
            "s1	",
            "-",
            "1",
            "sw-25g01",
            "x3000",
            "u12",
            "-",
            "12",
        ],
        [
            "sn02",
            "x3000",
            "u08",
            "s1	",
            "-",
            "2	",
            "sw-25g02",
            "x3000",
            "u13",
            "-",
            "12",
        ],
        [
            "sn01",
            "x3000",
            "u07",
            "ocp",
            "-",
            "1",
            "sw-25g01",
            "x3000",
            "u12",
            "-",
            "11",
        ],
        [
            "sn01",
            "x3000",
            "u07",
            "ocp",
            "-",
            "2	",
            "sw-25g02",
            "x3000",
            "u13",
            "-",
            "11",
        ],
        [
            "sn01",
            "x3000",
            "u07",
            "s1	",
            "-",
            "1",
            "sw-25g01",
            "x3000",
            "u12",
            "-",
            "10",
        ],
        [
            "sn01",
            "x3000",
            "u07",
            "s1	",
            "-",
            "2	",
            "sw-25g02",
            "x3000",
            "u13",
            "-",
            "10",
        ],
        [
            "wn03",
            "x3000",
            "u06",
            "ocp",
            "-",
            "1",
            "sw-25g01",
            "x3000",
            "u12",
            "-",
            "9",
        ],
        [
            "wn03",
            "x3000",
            "u06",
            "ocp",
            "-",
            "2	",
            "sw-25g02",
            "x3000",
            "u13",
            "-",
            "9",
        ],
        [
            "wn02",
            "x3000",
            "u05",
            "ocp",
            "-",
            "1",
            "sw-25g01",
            "x3000",
            "u12",
            "-",
            "8",
        ],
        [
            "wn02",
            "x3000",
            "u05",
            "ocp",
            "-",
            "2	",
            "sw-25g02",
            "x3000",
            "u13",
            "-",
            "8",
        ],
        [
            "wn01",
            "x3000",
            "u04",
            "ocp",
            "-",
            "1",
            "sw-25g01",
            "x3000",
            "u12",
            "-",
            "7",
        ],
        [
            "wn01",
            "x3000",
            "u04",
            "ocp",
            "-",
            "2	",
            "sw-25g02",
            "x3000",
            "u13",
            "-",
            "7",
        ],
        [
            "mn03",
            "x3000",
            "u03",
            "ocp",
            "-",
            "1",
            "sw-25g01",
            "x3000",
            "u12",
            "-",
            "5",
        ],
        [
            "mn03",
            "x3000",
            "u03",
            "s1	",
            "-",
            "1	",
            "sw-25g02",
            "x3000",
            "u13",
            "-",
            "5",
        ],
        [
            "mn02",
            "x3000",
            "u02",
            "ocp",
            "-",
            "1",
            "sw-25g01",
            "x3000",
            "u12",
            "-",
            "3",
        ],
        [
            "mn02",
            "x3000",
            "u02",
            "s1	",
            "-",
            "1	",
            "sw-25g02",
            "x3000",
            "u13",
            "-",
            "3",
        ],
        [
            "mn01",
            "x3000",
            "u01",
            "ocp",
            "-",
            "1",
            "sw-25g01",
            "x3000",
            "u12",
            "-",
            "1",
        ],
        [
            "sw-cdu-001",
            "x3000",
            "u01",
            "",
            "-",
            "1	",
            "sw-cdu-002",
            "x3000",
            "u13",
            "-",
            "20",
        ],
        [
            "CAN switch",
            "cfcanb4s1",
            "",
            "",
            "-",
            "9",
            "sw-25g01",
            "x3000",
            "u12",
            "-",
            "36",
        ],
        [
            "CAN switch",
            "cfcanb4s1",
            "",
            "",
            "-",
            "11",
            "sw-25g02",
            "x3000",
            "u13",
            "-",
            "36",
        ],
        # BAD ROW, do not include in normal run
        [
            "mn99",
            "x3000",
            "u12",
            "",
            "-",
            "1",
            "mn98",
            "x3000",
            "u13",
            "-",
            "1",
        ],
    ]
    for row in range(0, 36):
        for col in range(0, 11):
            ws1.cell(column=col + 9, row=row + 15, value=f"{test_data[row][col]}")

    # Tab 2 "Bad_Headers"
    ws2 = wb.create_sheet(title="Bad_Headers")
    ws2["I14"] = "Source"
    ws2["J14"] = "Rack"
    ws2["K14"] = "Location"
    # Missing Header
    # None
    ws2["M14"] = "Port"
    ws2["N14"] = "Destination"
    ws2["O14"] = "Rack"
    ws2["P14"] = "Location"
    # None
    ws2["R14"] = "Port"

    # Tab 3 "More_connections" containing bad connections
    ws3 = wb.create_sheet(title="More_connections")
    ws3["I14"] = "Source"
    ws3["J14"] = "Rack"
    ws3["K14"] = "Location"
    ws3["L14"] = "Slot"
    # None
    ws3["N14"] = "Port"
    ws3["O14"] = "Destination"
    ws3["P14"] = "Rack"
    ws3["Q14"] = "Location"
    # None
    ws3["S14"] = "Port"

    test_data3 = [
        [
            "sw-25g01",
            "x3000",
            "u12",
            "",
            "-",
            "51",
            "sw-25g02",
            "x3000",
            "u13",
            "-",
            "51",
        ],
        [
            "sw-25g01",
            "x3000",
            "u12",
            "",
            "-",
            "52",
            "sw-25g02",
            "x3000",
            "u13",
            "-",
            "52",
        ],
        [
            "sw-25g01",
            "x3000",
            "u12",
            "",
            "-",
            "52",
            "sw-25g02",
            "x3000",
            "u13",
            "-",
            "51",
        ],
        [
            "sw-cdu01",
            "x3000",
            "u12",
            "",
            "-",
            "52",
            "sw-smn99",
            "x3000",
            "u13",
            "-",
            "52",
        ],
        [
            "mn-99",
            "x3000",
            "u12",
            "",
            "-",
            "52",
            "sw-25g01",
            "x3000",
            "u13",
            "-",
            "52",
        ],
        [
            "mn-99",
            "x3000",
            "u12",
            "",
            "-",
            "50",
            "sw-25g02",
            "x3000",
            "u13",
            "-",
            "52",
        ],
        [
            "mn-99",
            "x3000",
            "u12",
            "",
            "-",
            "51",
            "sw-smn98",
            "x3000",
            "u13",
            "-",
            "52",
        ],
        [
            "mn-99",
            "x3000",
            "u12",
            "",
            "-",
            "52",
            "sw-smn99",
            "x3000",
            "u13",
            "-",
            "52",
        ],
        [
            "sw-100g01",
            "x3000",
            "u12",
            "",
            "-",
            "52",
            "sw-smn99",
            "x3000",
            "u13",
            "-",
            "52",
        ],
    ]
    for row in range(0, 9):
        for col in range(0, 11):
            ws3.cell(column=col + 9, row=row + 15, value=f"{test_data3[row][col]}")

    # Tab 3 "HMN" containing subrack connections
    ws3 = wb.create_sheet(title="HMN")
    ws3["I14"] = "Source"
    ws3["J14"] = "Rack"
    ws3["K14"] = "Location"
    ws3["L14"] = "Slot"
    ws3["M14"] = "Parent"
    # None
    ws3["O14"] = "Port"
    ws3["P14"] = "Destination"
    ws3["Q14"] = "Rack"
    ws3["R14"] = "Location"
    # None
    ws3["T14"] = "Port"

    test_data_hmn = [
        [
            "cn08",
            "x3000",
            "u15L",
            "",
            "SubRack-002-RCM",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ],
        [
            "SubRack-002-RCM",
            "x3000",
            "u15",
            "",
            "",
            "-",
            "RCM",
            "sw-smn02",
            "x3000",
            "u32",
            "-",
            "36",
        ],
        [
            "nid000004",
            "x3000",
            "u13L",
            "",
            "SubRack-001-RCM",
            "-",
            "j3",
            "sw-smn02",
            "x3000",
            "u32",
            "-",
            "28",
        ],
        [
            "SubRack-001-RCM",
            "x3000",
            "u13",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ],
        [
            "nid000005",
            "x3000",
            "u13L",
            "",
            "",
            "-",
            "j3",
            "JUNK",
            "x3000",
            "u32",
            "-",
            "28",
        ],
    ]
    for row in range(0, 5):
        for col in range(0, 12):
            ws3.cell(column=col + 9, row=row + 15, value=f"{test_data_hmn[row][col]}")

    wb.save(filename=test_file)
