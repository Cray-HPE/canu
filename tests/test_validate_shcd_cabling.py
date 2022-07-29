# MIT License
#
# (C) Copyright [2022] Hewlett Packard Enterprise Development LP
#
# Permission is hereby granted, free of charge, to any person obtaining a
# copy of this software and associated documentation files (the "Software"),
# to deal in the Software without restriction, including without limitation
# the rights to use, copy, modify, merge, publish, distribute, sublicense,
# and/or sell copies of the Software, and to permit persons to whom the
# Software is furnished to do so, subject to the following conditions:
#
# The above copyright notice and this permission notice shall be included
# in all copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL
# THE AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR
# OTHER LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE,
# ARISING FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR
# OTHER DEALINGS IN THE SOFTWARE.
"""Test CANU validate shcd-cabling commands."""
from unittest.mock import patch

from click import testing
from openpyxl import Workbook
import requests
import responses

from canu.cli import cli
from canu.utils.cache import remove_switch_from_cache


architecture = "tds"
username = "admin"
password = "admin"
ip = "192.168.1.1"
ips = "192.168.1.1"
credentials = {"username": username, "password": password}
test_file = "test_file.xlsx"
tabs = "25G_10G"
corners = "I14,S30"
csm = "1.2"
cache_minutes = 0
runner = testing.CliRunner()


@patch("canu.report.switch.cabling.cabling.switch_vendor")
@patch("canu.report.switch.cabling.cabling.netmiko_command")
@responses.activate
def test_validate_shcd_cabling(netmiko_command, switch_vendor):
    """Test that the `canu validate shcd-cabling` command runs and returns valid cabling."""
    with runner.isolated_filesystem():
        switch_vendor.return_value = "aruba"
        netmiko_command.return_value = mac_address_table
        generate_test_file(test_file)
        responses.add(
            responses.POST,
            f"https://{ip}/rest/v10.04/login",
        )
        responses.add(
            responses.GET,
            f"https://{ip}/rest/v10.04/system?attributes=platform_name,hostname,system_mac",
            json=switch_info1,
        )
        responses.add(
            responses.GET,
            f"https://{ip}/rest/v10.04/system/interfaces/*/lldp_neighbors?depth=2",
            json=lldp_neighbors_json1,
        )
        responses.add(
            responses.GET,
            f"https://{ip}/rest/v10.04/system/vrfs/default/neighbors?depth=2",
            json=arp_neighbors_json1,
        )

        responses.add(
            responses.POST,
            f"https://{ip}/rest/v10.04/logout",
        )
        result = runner.invoke(
            cli,
            [
                "--cache",
                cache_minutes,
                "validate",
                "shcd-cabling",
                "--csm",
                csm,
                "--architecture",
                architecture,
                "--ips",
                ips,
                "--username",
                username,
                "--password",
                password,
                "--shcd",
                test_file,
                "--tabs",
                tabs,
                "--corners",
                corners,
                "--log",
                "DEBUG",
            ],
        )
        assert result.exit_code == 0

        assert (
            "sw-spine-001\n"
            + "Rack: x3000    Elevation: u12\n"
            + "--------------------------------------------------------------------------------\n"
            + "Port   SHCD                     Cabling\n"
            + "--------------------------------------------------------------------------------\n"
            + "1      sw-spine-002:1           sw-spine-002:1\n"
            + "2      sw-spine-002:2           sw-spine-002:2\n"
            + "4      sw-leaf-bmc-099:2        ncn-m88:ocp:2\n"
            + "9      ncn-w003:ocp:1           None\n"
            + "15     ncn-s003:ocp:1           None\n"
            + "16     uan001:ocp:1             None\n"
            + "17     uan001:ocp:2             None\n"
            + "47     sw-spine-002:47          None\n"
            + "48     sw-leaf-bmc-001:49       None\n"
            + "5      None                     sw-leaf-bmc-099:5\n"
            + "\n"
            + "sw-spine-002\n"
            + "Rack: x3000    Elevation: u13\n"
            + "--------------------------------------------------------------------------------\n"
            + "Port   SHCD                     Cabling\n"
            + "--------------------------------------------------------------------------------\n"
            + "1      sw-spine-001:1           sw-spine-001:1\n"
            + "2      sw-spine-001:2           sw-spine-001:2\n"
            + "9      ncn-w003:ocp:2           None\n"
            + "16     uan001:pcie-slot1:1      None\n"
            + "17     uan001:pcie-slot1:2      None\n"
            + "47     sw-spine-001:47          None\n"
            + "48     sw-leaf-bmc-001:50       None\n"
        ) in str(result.output)

        assert (
            "Sheet: 25G_10G\n"
            + "Cell: I30      Name: CAN switch\n"
            + "Cell: O17      Name: junk\n"
            + "Cell: O19      Name: junk\n"
        ) in str(result.output)

        assert (
            "Node type could not be determined for the following.\n"
            + "These nodes are not currently included in the model.\n"
            + "(This may be a missing architectural definition/lookup or a spelling error)\n"
            + "--------------------------------------------------------------------------------\n"
            + "sw-spine-001     1/1/3     ===> 00:00:00:00:00:00  XEROX CORPORATION 192.168.1.2:vlan1, 192.168.2.2:vlan3\n"
            + "Nodes that show up as MAC addresses might need to have LLDP enabled.\n"
            + "\n"
            + "The following nodes should be renamed\n"
            + "--------------------------------------------------------------------------------\n"
            + "sw-leaf-bmc99 should be renamed sw-leaf-bmc-099\n"
            + "sw-spine01 should be renamed sw-spine-001\n"
            + "sw-spine02 should be renamed sw-spine-002\n"
        ) in str(result.output)

        remove_switch_from_cache(ip)


@patch("canu.report.switch.cabling.cabling.switch_vendor")
@patch("canu.report.switch.cabling.cabling.netmiko_command")
@responses.activate
def test_validate_shcd_cabling_macs(netmiko_command, switch_vendor):
    """Test that the `canu validate shcd-cabling` command runs with the `--macs` flag and produces valid output."""
    with runner.isolated_filesystem():
        switch_vendor.return_value = "aruba"
        netmiko_command.return_value = mac_address_table
        generate_test_file(test_file)
        responses.add(
            responses.POST,
            f"https://{ip}/rest/v10.04/login",
        )
        responses.add(
            responses.GET,
            f"https://{ip}/rest/v10.04/system?attributes=platform_name,hostname,system_mac",
            json=switch_info1,
        )
        responses.add(
            responses.GET,
            f"https://{ip}/rest/v10.04/system/interfaces/*/lldp_neighbors?depth=2",
            json=lldp_neighbors_json1,
        )
        responses.add(
            responses.GET,
            f"https://{ip}/rest/v10.04/system/vrfs/default/neighbors?depth=2",
            json=arp_neighbors_json1,
        )

        responses.add(
            responses.POST,
            f"https://{ip}/rest/v10.04/logout",
        )
        result = runner.invoke(
            cli,
            [
                "--cache",
                cache_minutes,
                "validate",
                "shcd-cabling",
                "--csm",
                csm,
                "--architecture",
                architecture,
                "--ips",
                ips,
                "--username",
                username,
                "--password",
                password,
                "--shcd",
                test_file,
                "--tabs",
                tabs,
                "--corners",
                corners,
                "--macs",
            ],
        )
        assert result.exit_code == 0

        assert ("  Connecting to 192.168.1.1 - Switch 1 of 1        \n\n") in str(
            result.output,
        )


@patch("canu.report.switch.cabling.cabling.switch_vendor")
@patch("canu.report.switch.cabling.cabling.netmiko_command")
@responses.activate
def test_validate_shcd_cabling_full_architecture(netmiko_command, switch_vendor):
    """Test that the `canu validate shcd-cabling` command runs and returns valid cabling with full architecture."""
    full_architecture = "full"
    with runner.isolated_filesystem():
        switch_vendor.return_value = "aruba"
        netmiko_command.return_value = mac_address_table
        generate_test_file(test_file)
        responses.add(
            responses.POST,
            f"https://{ip}/rest/v10.04/login",
        )
        responses.add(
            responses.GET,
            f"https://{ip}/rest/v10.04/system?attributes=platform_name,hostname,system_mac",
            json=switch_info2,
        )
        responses.add(
            responses.GET,
            f"https://{ip}/rest/v10.04/system/interfaces/*/lldp_neighbors?depth=2",
            json=lldp_neighbors_json2,
        )
        responses.add(
            responses.GET,
            f"https://{ip}/rest/v10.04/system/vrfs/default/neighbors?depth=2",
            json=arp_neighbors_json2,
        )

        responses.add(
            responses.POST,
            f"https://{ip}/rest/v10.04/logout",
        )

        result = runner.invoke(
            cli,
            [
                "--cache",
                cache_minutes,
                "validate",
                "shcd-cabling",
                "--csm",
                csm,
                "--architecture",
                full_architecture,
                "--ips",
                ips,
                "--username",
                username,
                "--password",
                password,
                "--shcd",
                test_file,
                "--tabs",
                tabs,
                "--corners",
                corners,
            ],
        )
        assert result.exit_code == 0

        assert (
            "sw-leaf-001\n"
            + "Rack: x3000    Elevation: u12\n"
            + "--------------------------------------------------------------------------------\n"
            + "Port   SHCD                     Cabling\n"
            + "--------------------------------------------------------------------------------\n"
            + "1      sw-leaf-002:1            sw-leaf-002:1\n"
            + "2      sw-leaf-002:2            None\n"
            + "4      sw-leaf-bmc-099:2        None\n"
            + "9      ncn-w003:ocp:1           None\n"
            + "15     ncn-s003:ocp:1           None\n"
            + "16     uan001:ocp:1             None\n"
            + "17     uan001:ocp:2             None\n"
            + "47     sw-leaf-002:47           None\n"
            + "48     sw-leaf-bmc-001:49       None\n"
            + "\n"
            + "sw-leaf-002\n"
            + "Rack: x3000    Elevation: u13\n"
            + "--------------------------------------------------------------------------------\n"
            + "Port   SHCD                     Cabling\n"
            + "--------------------------------------------------------------------------------\n"
            + "1      sw-leaf-001:1            sw-leaf-001:1\n"
            + "2      sw-leaf-001:2            None\n"
            + "9      ncn-w003:ocp:2           None\n"
            + "16     uan001:pcie-slot1:1      None\n"
            + "17     uan001:pcie-slot1:2      None\n"
            + "47     sw-leaf-001:47           None\n"
            + "48     sw-leaf-bmc-001:50       None\n"
        ) in str(result.output)

        assert (
            "Sheet: 25G_10G\n"
            + "Cell: I30      Name: CAN switch\n"
            + "Cell: O17      Name: junk\n"
            + "Cell: O19      Name: junk\n"
        ) in str(result.output)

        assert (
            "Node type could not be determined for the following.\n"
            + "These nodes are not currently included in the model.\n"
            + "(This may be a missing architectural definition/lookup or a spelling error)\n"
            + "--------------------------------------------------------------------------------\n"
            + "sw-leaf-002      1/1/3     ===> 00:40:a6:00:00:00  Cray, Inc. \n"
            + "Nodes that show up as MAC addresses might need to have LLDP enabled.\n"
            + "\n"
            + "The following nodes should be renamed\n"
            + "--------------------------------------------------------------------------------\n"
            + "sw-leaf01 should be renamed sw-leaf-001\n"
            + "sw-leaf02 should be renamed sw-leaf-002\n"
        ) in str(result.output)
        remove_switch_from_cache(ip)


@patch("canu.report.switch.cabling.cabling.switch_vendor")
@patch("canu.report.switch.cabling.cabling.netmiko_command")
@responses.activate
def test_validate_shcd_cabling_file(netmiko_command, switch_vendor):
    """Test that the `canu validate shcd-cabling` command runs and returns valid cabling with IPs from file."""
    with runner.isolated_filesystem():
        switch_vendor.return_value = "aruba"
        netmiko_command.return_value = mac_address_table
        with open("test.txt", "w") as f:
            f.write("192.168.1.1")

        generate_test_file(test_file)
        responses.add(
            responses.POST,
            f"https://{ip}/rest/v10.04/login",
        )
        responses.add(
            responses.GET,
            f"https://{ip}/rest/v10.04/system?attributes=platform_name,hostname,system_mac",
            json=switch_info1,
        )
        responses.add(
            responses.GET,
            f"https://{ip}/rest/v10.04/system/interfaces/*/lldp_neighbors?depth=2",
            json=lldp_neighbors_json1,
        )
        responses.add(
            responses.GET,
            f"https://{ip}/rest/v10.04/system/vrfs/default/neighbors?depth=2",
            json=arp_neighbors_json1,
        )

        responses.add(
            responses.POST,
            f"https://{ip}/rest/v10.04/logout",
        )

        result = runner.invoke(
            cli,
            [
                "--cache",
                cache_minutes,
                "validate",
                "shcd-cabling",
                "--csm",
                csm,
                "--architecture",
                architecture,
                "--ips-file",
                "test.txt",
                "--username",
                username,
                "--password",
                password,
                "--shcd",
                test_file,
                "--tabs",
                tabs,
                "--corners",
                corners,
            ],
        )

        assert result.exit_code == 0

        assert (
            "sw-spine-001\n"
            + "Rack: x3000    Elevation: u12\n"
            + "--------------------------------------------------------------------------------\n"
            + "Port   SHCD                     Cabling\n"
            + "--------------------------------------------------------------------------------\n"
            + "1      sw-spine-002:1           sw-spine-002:1\n"
            + "2      sw-spine-002:2           sw-spine-002:2\n"
            + "4      sw-leaf-bmc-099:2        ncn-m88:ocp:2\n"
            + "9      ncn-w003:ocp:1           None\n"
            + "15     ncn-s003:ocp:1           None\n"
            + "16     uan001:ocp:1             None\n"
            + "17     uan001:ocp:2             None\n"
            + "47     sw-spine-002:47          None\n"
            + "48     sw-leaf-bmc-001:49       None\n"
            + "5      None                     sw-leaf-bmc-099:5\n"
            + "\n"
            + "sw-spine-002\n"
            + "Rack: x3000    Elevation: u13\n"
            + "--------------------------------------------------------------------------------\n"
            + "Port   SHCD                     Cabling\n"
            + "--------------------------------------------------------------------------------\n"
            + "1      sw-spine-001:1           sw-spine-001:1\n"
            + "2      sw-spine-001:2           sw-spine-001:2\n"
            + "9      ncn-w003:ocp:2           None\n"
            + "16     uan001:pcie-slot1:1      None\n"
            + "17     uan001:pcie-slot1:2      None\n"
            + "47     sw-spine-001:47          None\n"
            + "48     sw-leaf-bmc-001:50       None\n"
        ) in str(result.output)

        assert (
            "Sheet: 25G_10G\n"
            + "Cell: I30      Name: CAN switch\n"
            + "Cell: O17      Name: junk\n"
            + "Cell: O19      Name: junk\n"
        ) in str(result.output)

        assert (
            "Node type could not be determined for the following.\n"
            + "These nodes are not currently included in the model.\n"
            + "(This may be a missing architectural definition/lookup or a spelling error)\n"
            + "--------------------------------------------------------------------------------\n"
            + "sw-spine-001     1/1/3     ===> 00:00:00:00:00:00  XEROX CORPORATION 192.168.1.2:vlan1, 192.168.2.2:vlan3\n"
            + "Nodes that show up as MAC addresses might need to have LLDP enabled.\n"
            + "\n"
            + "The following nodes should be renamed\n"
            + "--------------------------------------------------------------------------------\n"
            + "sw-leaf-bmc99 should be renamed sw-leaf-bmc-099\n"
            + "sw-spine01 should be renamed sw-spine-001\n"
            + "sw-spine02 should be renamed sw-spine-002\n"
        ) in str(result.output)
        remove_switch_from_cache(ip)


def test_validate_shcd_cabling_missing_ips():
    """Test that the `canu validate shcd-cabling` command errors on missing IP address."""
    with runner.isolated_filesystem():
        generate_test_file(test_file)
        result = runner.invoke(
            cli,
            [
                "--cache",
                cache_minutes,
                "validate",
                "shcd-cabling",
                "--csm",
                csm,
                "--architecture",
                architecture,
                "--username",
                username,
                "--password",
                password,
                "--shcd",
                test_file,
                "--tabs",
                tabs,
                "--corners",
                corners,
            ],
        )
        assert result.exit_code == 2
        assert (
            "Error: Missing one of the required mutually exclusive options from 'Network cabling IPv4 input sources' option group"
            in str(result.output)
        )


def test_validate_shcd_cabling_mutually_exclusive_ips_and_file():
    """Test that the `canu validate shcd-cabling` command only accepts IPs from command line OR file input, not both."""
    with runner.isolated_filesystem():
        generate_test_file(test_file)
        result = runner.invoke(
            cli,
            [
                "--cache",
                cache_minutes,
                "validate",
                "shcd-cabling",
                "--csm",
                csm,
                "--architecture",
                architecture,
                "--username",
                username,
                "--password",
                password,
                "--shcd",
                test_file,
                "--tabs",
                tabs,
                "--corners",
                corners,
                "--ips",
                ips,
                "--ips-file",
                "file.txt",
            ],
        )
        assert result.exit_code == 2
        assert (
            "Error: Mutually exclusive options from 'Network cabling IPv4 input sources'"
            in str(result.output)
        )


def test_validate_shcd_cabling_invalid_ip():
    """Test that the `canu validate shcd-cabling` command errors on invalid IP address."""
    invalid_ip = "999.999.999.999"

    with runner.isolated_filesystem():
        generate_test_file(test_file)
        result = runner.invoke(
            cli,
            [
                "--cache",
                cache_minutes,
                "validate",
                "shcd-cabling",
                "--csm",
                csm,
                "--architecture",
                architecture,
                "--ips",
                invalid_ip,
                "--username",
                username,
                "--password",
                password,
                "--shcd",
                test_file,
                "--tabs",
                tabs,
                "--corners",
                corners,
            ],
        )
        assert result.exit_code == 2
        assert (
            "Error: Invalid value for '--ips': These items are not ipv4 addresses: ['999.999.999.999']"
            in str(result.output)
        )


def test_validate_shcd_cabling_invalid_ip_file():
    """Test that the `canu validate shcd-cabling` command errors on invalid IPs from a file."""
    invalid_ip = "999.999.999.999"

    with runner.isolated_filesystem():
        with open("test.txt", "w") as f:
            f.write(invalid_ip)

        generate_test_file(test_file)
        result = runner.invoke(
            cli,
            [
                "--cache",
                cache_minutes,
                "validate",
                "shcd-cabling",
                "--csm",
                csm,
                "--architecture",
                architecture,
                "--ips-file",
                "test.txt",
                "--username",
                username,
                "--password",
                password,
                "--shcd",
                test_file,
                "--tabs",
                tabs,
                "--corners",
                corners,
            ],
        )
        assert result.exit_code == 2
        assert "Error: Invalid value:" in str(result.output)


@patch("canu.report.switch.cabling.cabling.switch_vendor")
@patch("canu.report.switch.cabling.cabling.netmiko_command")
@responses.activate
def test_validate_shcd_cabling_bad_ip(netmiko_command, switch_vendor):
    """Test that the `canu validate shcd-cabling` command errors on bad IP address."""
    bad_ip = "192.168.1.99"

    with runner.isolated_filesystem():
        switch_vendor.return_value = "aruba"
        netmiko_command.return_value = mac_address_table
        responses.add(
            responses.POST,
            f"https://{bad_ip}/rest/v10.04/login",
            body=requests.exceptions.ConnectionError(
                "Failed to establish a new connection: [Errno 60] Operation timed out'))",
            ),
        )
        generate_test_file(test_file)
        result = runner.invoke(
            cli,
            [
                "--cache",
                cache_minutes,
                "validate",
                "shcd-cabling",
                "--csm",
                csm,
                "--architecture",
                architecture,
                "--ips",
                bad_ip,
                "--username",
                username,
                "--password",
                password,
                "--shcd",
                test_file,
                "--tabs",
                tabs,
                "--corners",
                corners,
            ],
        )
        assert result.exit_code == 0
        assert "check the IP address and try again" in str(result.output)


@patch("canu.report.switch.cabling.cabling.switch_vendor")
@patch("canu.report.switch.cabling.cabling.netmiko_command")
@responses.activate
def test_validate_shcd_cabling_bad_ip_file(netmiko_command, switch_vendor):
    """Test that the `canu validate shcd-cabling` command errors on a bad IP from a file."""
    bad_ip = "192.168.1.99"

    with runner.isolated_filesystem():
        switch_vendor.return_value = "aruba"
        netmiko_command.return_value = mac_address_table
        with open("test.txt", "w") as f:
            f.write(bad_ip)

        responses.add(
            responses.POST,
            f"https://{bad_ip}/rest/v10.04/login",
            body=requests.exceptions.ConnectionError(
                "Failed to establish a new connection: [Errno 60] Operation timed out'))",
            ),
        )
        generate_test_file(test_file)
        result = runner.invoke(
            cli,
            [
                "--cache",
                cache_minutes,
                "validate",
                "shcd-cabling",
                "--csm",
                csm,
                "--architecture",
                architecture,
                "--ips-file",
                "test.txt",
                "--username",
                username,
                "--password",
                password,
                "--shcd",
                test_file,
                "--tabs",
                tabs,
                "--corners",
                corners,
            ],
        )
        assert result.exit_code == 0
        assert "check the IP address and try again" in str(result.output)


@patch("canu.report.switch.cabling.cabling.switch_vendor")
@patch("canu.report.switch.cabling.cabling.netmiko_command")
@responses.activate
def test_validate_shcd_cabling_bad_password(netmiko_command, switch_vendor):
    """Test that the `canu validate shcd-cabling` command errors on bad credentials."""
    bad_password = "foo"

    with runner.isolated_filesystem():
        switch_vendor.return_value = "aruba"
        netmiko_command.return_value = mac_address_table
        responses.add(
            responses.POST,
            f"https://{ip}/rest/v10.04/login",
            body=requests.exceptions.HTTPError("Client Error: Unauthorized for url"),
        )
        generate_test_file(test_file)
        result = runner.invoke(
            cli,
            [
                "--cache",
                cache_minutes,
                "validate",
                "shcd-cabling",
                "--csm",
                csm,
                "--architecture",
                architecture,
                "--ips",
                ip,
                "--username",
                username,
                "--password",
                bad_password,
                "--shcd",
                test_file,
                "--tabs",
                tabs,
                "--corners",
                corners,
            ],
        )
        assert result.exit_code == 0
        assert "IP, username, or password" in str(result.output)


def test_validate_shcd_cabling_missing_file():
    """Test that the `canu validate shcd-cabling` command fails on missing file."""
    with runner.isolated_filesystem():
        generate_test_file(test_file)
        result = runner.invoke(
            cli,
            [
                "--cache",
                cache_minutes,
                "validate",
                "shcd-cabling",
                "--csm",
                csm,
                "--architecture",
                architecture,
                "--ips",
                ips,
                "--username",
                username,
                "--password",
                password,
                "--tabs",
                tabs,
                "--corners",
                corners,
            ],
        )
        assert result.exit_code == 2
        assert "Error: Missing option '--shcd'." in str(result.output)


def test_validate_shcd_cabling_bad_file():
    """Test that the `canu validate shcd-cabling` command fails on bad file."""
    bad_file = "does_not_exist.xlsx"
    with runner.isolated_filesystem():
        generate_test_file(test_file)
        result = runner.invoke(
            cli,
            [
                "--cache",
                cache_minutes,
                "validate",
                "shcd-cabling",
                "--csm",
                csm,
                "--architecture",
                architecture,
                "--ips",
                ips,
                "--username",
                username,
                "--password",
                password,
                "--shcd",
                bad_file,
                "--tabs",
                tabs,
                "--corners",
                corners,
            ],
        )
        assert result.exit_code == 2
        assert "Error: Invalid value for '--shcd':" in str(result.output)


@patch("canu.report.switch.cabling.cabling.switch_vendor")
@patch("canu.report.switch.cabling.cabling.netmiko_command")
@responses.activate
def test_validate_shcd_cabling_missing_tabs(netmiko_command, switch_vendor):
    """Test that the `canu validate shcd-cabling` command prompts for missing tabs."""
    with runner.isolated_filesystem():
        switch_vendor.return_value = "aruba"
        netmiko_command.return_value = mac_address_table
        generate_test_file(test_file)
        responses.add(
            responses.POST,
            f"https://{ip}/rest/v10.04/login",
        )
        responses.add(
            responses.GET,
            f"https://{ip}/rest/v10.04/system?attributes=platform_name,hostname,system_mac",
            json=switch_info1,
        )
        responses.add(
            responses.GET,
            f"https://{ip}/rest/v10.04/system/interfaces/*/lldp_neighbors?depth=2",
            json=lldp_neighbors_json1,
        )
        responses.add(
            responses.GET,
            f"https://{ip}/rest/v10.04/system/vrfs/default/neighbors?depth=2",
            json=arp_neighbors_json1,
        )

        responses.add(
            responses.POST,
            f"https://{ip}/rest/v10.04/logout",
        )

        result = runner.invoke(
            cli,
            [
                "--cache",
                cache_minutes,
                "validate",
                "shcd-cabling",
                "--csm",
                csm,
                "--architecture",
                architecture,
                "--ips",
                ips,
                "--username",
                username,
                "--password",
                password,
                "--shcd",
                test_file,
                "--corners",
                corners,
            ],
            input="25G_10G\n",
        )
        assert result.exit_code == 0

        assert (
            "sw-spine-001\n"
            + "Rack: x3000    Elevation: u12\n"
            + "--------------------------------------------------------------------------------\n"
            + "Port   SHCD                     Cabling\n"
            + "--------------------------------------------------------------------------------\n"
            + "1      sw-spine-002:1           sw-spine-002:1\n"
            + "2      sw-spine-002:2           sw-spine-002:2\n"
            + "4      sw-leaf-bmc-099:2        ncn-m88:ocp:2\n"
            + "9      ncn-w003:ocp:1           None\n"
            + "15     ncn-s003:ocp:1           None\n"
            + "16     uan001:ocp:1             None\n"
            + "17     uan001:ocp:2             None\n"
            + "47     sw-spine-002:47          None\n"
            + "48     sw-leaf-bmc-001:49       None\n"
            + "5      None                     sw-leaf-bmc-099:5\n"
            + "\n"
            + "sw-spine-002\n"
            + "Rack: x3000    Elevation: u13\n"
            + "--------------------------------------------------------------------------------\n"
            + "Port   SHCD                     Cabling\n"
            + "--------------------------------------------------------------------------------\n"
            + "1      sw-spine-001:1           sw-spine-001:1\n"
            + "2      sw-spine-001:2           sw-spine-001:2\n"
            + "9      ncn-w003:ocp:2           None\n"
            + "16     uan001:pcie-slot1:1      None\n"
            + "17     uan001:pcie-slot1:2      None\n"
            + "47     sw-spine-001:47          None\n"
            + "48     sw-leaf-bmc-001:50       None\n"
        ) in str(result.output)

        assert (
            "Sheet: 25G_10G\n"
            + "Cell: I30      Name: CAN switch\n"
            + "Cell: O17      Name: junk\n"
            + "Cell: O19      Name: junk\n"
        ) in str(result.output)

        assert (
            "Node type could not be determined for the following.\n"
            + "These nodes are not currently included in the model.\n"
            + "(This may be a missing architectural definition/lookup or a spelling error)\n"
            + "--------------------------------------------------------------------------------\n"
            + "sw-spine-001     1/1/3     ===> 00:00:00:00:00:00  XEROX CORPORATION 192.168.1.2:vlan1, 192.168.2.2:vlan3\n"
            + "Nodes that show up as MAC addresses might need to have LLDP enabled.\n"
            + "\n"
            + "The following nodes should be renamed\n"
            + "--------------------------------------------------------------------------------\n"
            + "sw-leaf-bmc99 should be renamed sw-leaf-bmc-099\n"
            + "sw-spine01 should be renamed sw-spine-001\n"
            + "sw-spine02 should be renamed sw-spine-002\n"
        ) in str(result.output)

        remove_switch_from_cache(ip)


@patch("canu.report.switch.cabling.cabling.switch_vendor")
@patch("canu.report.switch.cabling.cabling.netmiko_command")
@responses.activate
def test_validate_shcd_cabling_bad_tab(netmiko_command, switch_vendor):
    """Test that the `canu validate shcd-cabling` command fails on bad tab name."""
    bad_tab = "BAD_TAB"
    with runner.isolated_filesystem():
        switch_vendor.return_value = "aruba"
        netmiko_command.return_value = mac_address_table
        generate_test_file(test_file)
        responses.add(
            responses.POST,
            f"https://{ip}/rest/v10.04/login",
        )
        responses.add(
            responses.GET,
            f"https://{ip}/rest/v10.04/system?attributes=platform_name,hostname,system_mac",
            json=switch_info1,
        )
        responses.add(
            responses.GET,
            f"https://{ip}/rest/v10.04/system/interfaces/*/lldp_neighbors?depth=2",
            json=lldp_neighbors_json1,
        )
        responses.add(
            responses.GET,
            f"https://{ip}/rest/v10.04/system/vrfs/default/neighbors?depth=2",
            json=arp_neighbors_json1,
        )

        responses.add(
            responses.POST,
            f"https://{ip}/rest/v10.04/logout",
        )
        result = runner.invoke(
            cli,
            [
                "--cache",
                cache_minutes,
                "validate",
                "shcd-cabling",
                "--csm",
                csm,
                "--architecture",
                architecture,
                "--ips",
                ips,
                "--username",
                username,
                "--password",
                password,
                "--shcd",
                test_file,
                "--tabs",
                bad_tab,
                "--corners",
                corners,
            ],
        )
        assert result.exit_code == 1
        assert "Tab BAD_TAB not found in test_file.xlsx" in str(result.output)


@patch("canu.report.switch.cabling.cabling.switch_vendor")
@patch("canu.report.switch.cabling.cabling.netmiko_command")
@responses.activate
def test_validate_shcd_cabling_corner_prompt(netmiko_command, switch_vendor):
    """Test that the `canu validate shcd-cabling` command prompts for corner input and runs."""
    with runner.isolated_filesystem():
        switch_vendor.return_value = "aruba"
        netmiko_command.return_value = mac_address_table
        responses.add(
            responses.POST,
            f"https://{ip}/rest/v10.04/login",
        )
        responses.add(
            responses.GET,
            f"https://{ip}/rest/v10.04/system?attributes=platform_name,hostname,system_mac",
            json=switch_info1,
        )
        responses.add(
            responses.GET,
            f"https://{ip}/rest/v10.04/system/interfaces/*/lldp_neighbors?depth=2",
            json=lldp_neighbors_json1,
        )
        responses.add(
            responses.GET,
            f"https://{ip}/rest/v10.04/system/vrfs/default/neighbors?depth=2",
            json=arp_neighbors_json1,
        )

        responses.add(
            responses.POST,
            f"https://{ip}/rest/v10.04/logout",
        )
        generate_test_file(test_file)
        result = runner.invoke(
            cli,
            [
                "--cache",
                cache_minutes,
                "validate",
                "shcd-cabling",
                "--csm",
                csm,
                "--architecture",
                architecture,
                "--ips",
                ips,
                "--username",
                username,
                "--password",
                password,
                "--shcd",
                test_file,
                "--tabs",
                tabs,
            ],
            input="I14\nS30",
        )
        assert result.exit_code == 0

        assert (
            "sw-spine-001\n"
            + "Rack: x3000    Elevation: u12\n"
            + "--------------------------------------------------------------------------------\n"
            + "Port   SHCD                     Cabling\n"
            + "--------------------------------------------------------------------------------\n"
            + "1      sw-spine-002:1           sw-spine-002:1\n"
            + "2      sw-spine-002:2           sw-spine-002:2\n"
            + "4      sw-leaf-bmc-099:2        ncn-m88:ocp:2\n"
            + "9      ncn-w003:ocp:1           None\n"
            + "15     ncn-s003:ocp:1           None\n"
            + "16     uan001:ocp:1             None\n"
            + "17     uan001:ocp:2             None\n"
            + "47     sw-spine-002:47          None\n"
            + "48     sw-leaf-bmc-001:49       None\n"
            + "5      None                     sw-leaf-bmc-099:5\n"
            + "\n"
            + "sw-spine-002\n"
            + "Rack: x3000    Elevation: u13\n"
            + "--------------------------------------------------------------------------------\n"
            + "Port   SHCD                     Cabling\n"
            + "--------------------------------------------------------------------------------\n"
            + "1      sw-spine-001:1           sw-spine-001:1\n"
            + "2      sw-spine-001:2           sw-spine-001:2\n"
            + "9      ncn-w003:ocp:2           None\n"
            + "16     uan001:pcie-slot1:1      None\n"
            + "17     uan001:pcie-slot1:2      None\n"
            + "47     sw-spine-001:47          None\n"
            + "48     sw-leaf-bmc-001:50       None\n"
        ) in str(result.output)

        assert (
            "Sheet: 25G_10G\n"
            + "Cell: I30      Name: CAN switch\n"
            + "Cell: O17      Name: junk\n"
            + "Cell: O19      Name: junk\n"
        ) in str(result.output)

        assert (
            "Node type could not be determined for the following.\n"
            + "These nodes are not currently included in the model.\n"
            + "(This may be a missing architectural definition/lookup or a spelling error)\n"
            + "--------------------------------------------------------------------------------\n"
            + "sw-spine-001     1/1/3     ===> 00:00:00:00:00:00  XEROX CORPORATION 192.168.1.2:vlan1, 192.168.2.2:vlan3\n"
            + "Nodes that show up as MAC addresses might need to have LLDP enabled.\n"
            + "\n"
            + "The following nodes should be renamed\n"
            + "--------------------------------------------------------------------------------\n"
            + "sw-leaf-bmc99 should be renamed sw-leaf-bmc-099\n"
            + "sw-spine01 should be renamed sw-spine-001\n"
            + "sw-spine02 should be renamed sw-spine-002\n"
        ) in str(result.output)
        remove_switch_from_cache(ip)


@patch("canu.report.switch.cabling.cabling.switch_vendor")
@patch("canu.report.switch.cabling.cabling.netmiko_command")
@responses.activate
def test_validate_shcd_cabling_corners_too_narrow(netmiko_command, switch_vendor):
    """Test that the `canu validate shcd cabling` command fails on too narrow area."""
    corners_too_narrow = "I16,P48"
    with runner.isolated_filesystem():
        switch_vendor.return_value = "aruba"
        netmiko_command.return_value = mac_address_table
        responses.add(
            responses.POST,
            f"https://{ip}/rest/v10.04/login",
        )
        responses.add(
            responses.GET,
            f"https://{ip}/rest/v10.04/system?attributes=platform_name,hostname,system_mac",
            json=switch_info1,
        )
        responses.add(
            responses.GET,
            f"https://{ip}/rest/v10.04/system/interfaces/*/lldp_neighbors?depth=2",
            json=lldp_neighbors_json1,
        )
        responses.add(
            responses.GET,
            f"https://{ip}/rest/v10.04/system/vrfs/default/neighbors?depth=2",
            json=arp_neighbors_json1,
        )

        responses.add(
            responses.POST,
            f"https://{ip}/rest/v10.04/logout",
        )
        generate_test_file(test_file)
        result = runner.invoke(
            cli,
            [
                "--cache",
                cache_minutes,
                "validate",
                "shcd-cabling",
                "--csm",
                csm,
                "--architecture",
                architecture,
                "--ips",
                ips,
                "--username",
                username,
                "--password",
                password,
                "--shcd",
                test_file,
                "--tabs",
                tabs,
                "--corners",
                corners_too_narrow,
            ],
        )
        assert result.exit_code == 1
        assert "Not enough columns exist." in str(result.output)
        remove_switch_from_cache(ip)


@patch("canu.report.switch.cabling.cabling.switch_vendor")
@patch("canu.report.switch.cabling.cabling.netmiko_command")
@responses.activate
def test_validate_shcd_cabling_corners_too_high(netmiko_command, switch_vendor):
    """Test that the `canu validate shcd-cabling` command fails on empty cells."""
    corners_too_high = "H16,S48"
    with runner.isolated_filesystem():
        switch_vendor.return_value = "aruba"
        netmiko_command.return_value = mac_address_table
        responses.add(
            responses.POST,
            f"https://{ip}/rest/v10.04/login",
        )
        responses.add(
            responses.GET,
            f"https://{ip}/rest/v10.04/system?attributes=platform_name,hostname,system_mac",
            json=switch_info1,
        )
        responses.add(
            responses.GET,
            f"https://{ip}/rest/v10.04/system/interfaces/*/lldp_neighbors?depth=2",
            json=lldp_neighbors_json1,
        )
        responses.add(
            responses.GET,
            f"https://{ip}/rest/v10.04/system/vrfs/default/neighbors?depth=2",
            json=arp_neighbors_json1,
        )

        responses.add(
            responses.POST,
            f"https://{ip}/rest/v10.04/logout",
        )
        generate_test_file(test_file)
        result = runner.invoke(
            cli,
            [
                "--cache",
                cache_minutes,
                "validate",
                "shcd-cabling",
                "--csm",
                csm,
                "--architecture",
                architecture,
                "--ips",
                ips,
                "--username",
                username,
                "--password",
                password,
                "--shcd",
                test_file,
                "--tabs",
                tabs,
                "--corners",
                corners_too_high,
            ],
        )
        assert result.exit_code == 1
        assert "On tab 25G_10G, header column Source not found." in str(result.output)
        assert "On tab 25G_10G, the header is formatted incorrectly." in str(
            result.output,
        )
        remove_switch_from_cache(ip)


@patch("canu.report.switch.cabling.cabling.switch_vendor")
@patch("canu.report.switch.cabling.cabling.netmiko_command")
@responses.activate
def test_validate_shcd_cabling_corners_bad_cell(netmiko_command, switch_vendor):
    """Test that the `canu validate shcd-cabling` command fails on bad cell."""
    corners_bad_cell = "16,S48"
    with runner.isolated_filesystem():
        switch_vendor.return_value = "aruba"
        netmiko_command.return_value = mac_address_table
        responses.add(
            responses.POST,
            f"https://{ip}/rest/v10.04/login",
        )
        responses.add(
            responses.GET,
            f"https://{ip}/rest/v10.04/system?attributes=platform_name,hostname,system_mac",
            json=switch_info1,
        )
        responses.add(
            responses.GET,
            f"https://{ip}/rest/v10.04/system/interfaces/*/lldp_neighbors?depth=2",
            json=lldp_neighbors_json1,
        )
        responses.add(
            responses.GET,
            f"https://{ip}/rest/v10.04/system/vrfs/default/neighbors?depth=2",
            json=arp_neighbors_json1,
        )

        responses.add(
            responses.POST,
            f"https://{ip}/rest/v10.04/logout",
        )
        generate_test_file(test_file)
        result = runner.invoke(
            cli,
            [
                "--cache",
                cache_minutes,
                "validate",
                "shcd-cabling",
                "--csm",
                csm,
                "--architecture",
                architecture,
                "--ips",
                ips,
                "--username",
                username,
                "--password",
                password,
                "--shcd",
                test_file,
                "--tabs",
                tabs,
                "--corners",
                corners_bad_cell,
            ],
        )
        assert result.exit_code == 1
        assert "Bad range of cells entered for tab 25G_10G." in str(result.output)
        remove_switch_from_cache(ip)


def test_validate_shcd_cabling_not_enough_corners():
    """Test that the `canu validate shcd-cabling` command fails on not enough corners."""
    not_enough_corners = "H16"
    with runner.isolated_filesystem():
        generate_test_file(test_file)
        result = runner.invoke(
            cli,
            [
                "--cache",
                cache_minutes,
                "validate",
                "shcd-cabling",
                "--csm",
                csm,
                "--architecture",
                architecture,
                "--ips",
                ips,
                "--username",
                username,
                "--password",
                password,
                "--shcd",
                test_file,
                "--tabs",
                tabs,
                "--corners",
                not_enough_corners,
            ],
        )
        assert result.exit_code == 0
        assert "There were 1 corners entered, but there should be 2." in str(
            result.output,
        )
        remove_switch_from_cache(ip)


@patch("canu.report.switch.cabling.cabling.switch_vendor")
@patch("canu.report.switch.cabling.cabling.netmiko_command")
@responses.activate
def test_validate_shcd_cabling_bad_headers(netmiko_command, switch_vendor):
    """Test that the `canu validate shcd-cabling` command fails on bad headers."""
    bad_header_tab = "Bad_Headers"
    with runner.isolated_filesystem():
        switch_vendor.return_value = "aruba"
        netmiko_command.return_value = mac_address_table
        responses.add(
            responses.POST,
            f"https://{ip}/rest/v10.04/login",
        )
        responses.add(
            responses.GET,
            f"https://{ip}/rest/v10.04/system?attributes=platform_name,hostname,system_mac",
            json=switch_info1,
        )
        responses.add(
            responses.GET,
            f"https://{ip}/rest/v10.04/system/interfaces/*/lldp_neighbors?depth=2",
            json=lldp_neighbors_json1,
        )
        responses.add(
            responses.GET,
            f"https://{ip}/rest/v10.04/system/vrfs/default/neighbors?depth=2",
            json=arp_neighbors_json1,
        )

        responses.add(
            responses.POST,
            f"https://{ip}/rest/v10.04/logout",
        )
        generate_test_file(test_file)
        result = runner.invoke(
            cli,
            [
                "--cache",
                cache_minutes,
                "validate",
                "shcd-cabling",
                "--csm",
                csm,
                "--architecture",
                architecture,
                "--ips",
                ips,
                "--username",
                username,
                "--password",
                password,
                "--shcd",
                test_file,
                "--tabs",
                bad_header_tab,
                "--corners",
                corners,
            ],
        )
        assert result.exit_code == 1
        assert "On tab Bad_Headers, header column Slot not found" in str(result.output)
        remove_switch_from_cache(ip)


@patch("canu.report.switch.cabling.cabling.switch_vendor")
@patch("canu.report.switch.cabling.cabling.netmiko_command")
@responses.activate
def test_validate_shcd_cabling_bad_architectural_definition(
    netmiko_command,
    switch_vendor,
):
    """Test that the `canu validate shcd-cabling` command fails with bad connections."""
    corners_bad_row = "I14,S31"
    with runner.isolated_filesystem():
        switch_vendor.return_value = "aruba"
        netmiko_command.return_value = mac_address_table
        responses.add(
            responses.POST,
            f"https://{ip}/rest/v10.04/login",
        )
        responses.add(
            responses.GET,
            f"https://{ip}/rest/v10.04/system?attributes=platform_name,hostname,system_mac",
            json=switch_info1,
        )
        responses.add(
            responses.GET,
            f"https://{ip}/rest/v10.04/system/interfaces/*/lldp_neighbors?depth=2",
            json=lldp_neighbors_json1,
        )
        responses.add(
            responses.GET,
            f"https://{ip}/rest/v10.04/system/vrfs/default/neighbors?depth=2",
            json=arp_neighbors_json1,
        )

        responses.add(
            responses.POST,
            f"https://{ip}/rest/v10.04/logout",
        )
        generate_test_file(test_file)
        result = runner.invoke(
            cli,
            [
                "--cache",
                cache_minutes,
                "validate",
                "shcd-cabling",
                "--csm",
                csm,
                "--architecture",
                architecture,
                "--ips",
                ips,
                "--username",
                username,
                "--password",
                password,
                "--shcd",
                test_file,
                "--tabs",
                tabs,
                "--corners",
                corners_bad_row,
            ],
        )
        assert result.exit_code == 1
        assert "The plan-of-record architectural definition does not allow connections" in str(
            result.output,
        )
        remove_switch_from_cache(ip)


# Switch 1
switch_info1 = {
    "hostname": "sw-spine01",
    "platform_name": "X86-64",
    "system_mac": "aa:aa:aa:aa:aa:aa",
}

lldp_neighbors_json1 = {
    "1%2F1%2F1": {
        "bb:bb:bb:bb:bb:bb,1/1/1": {
            "chassis_id": "bb:bb:bb:bb:bb:bb",
            "mac_addr": "bb:bb:bb:bb:bb:cc",
            "neighbor_info": {
                "chassis_description": "Test switch description",
                "chassis_name": "sw-spine02",
                "port_description": "",
                "port_id_subtype": "if_name",
            },
            "port_id": "1/1/1",
        },
    },
    "1%2F1%2F2": {
        "aa:bb:cc:88:00:00,1/1/2": {
            "chassis_id": "aa:bb:cc:88:00:00",
            "mac_addr": "aa:bb:cc:88:00:03",
            "neighbor_info": {
                "chassis_description": "Test switch2 description",
                "chassis_name": "sw-spine02",
                "port_description": "1/1/2",
                "port_id_subtype": "if_name",
            },
            "port_id": "1/1/2",
        },
    },
    "1%2F1%2F3": {
        "00:00:00:00:00:00,00:00:00:00:00:00": {
            "chassis_id": "00:00:00:00:00:00",
            "mac_addr": "00:00:00:00:00:00",
            "neighbor_info": {
                "chassis_description": "",
                "chassis_name": "",
                "port_description": "",
                "port_id_subtype": "link_local_addr",
            },
            "port_id": "00:00:00:00:00:00",
        },
        "11:11:11:11:11:11,11:11:11:11:11:11": {
            "chassis_id": "11:11:11:11:11:11",
            "mac_addr": "11:11:11:11:11:11",
            "neighbor_info": {
                "chassis_description": "",
                "chassis_name": "",
                "port_description": "",
                "port_id_subtype": "link_local_addr",
            },
            "port_id": "11:11:11:11:11:11",
        },
    },
    "1%2F1%2F4": {
        "aa:aa:aa:aa:aa:aa,aa:aa:aa:aa:aa:aa": {
            "chassis_id": "aa:aa:aa:aa:aa:aa",
            "mac_addr": "aa:aa:aa:aa:aa:aa",
            "neighbor_info": {
                "chassis_description": "NCN description",
                "chassis_name": "ncn-m88",
                "port_description": "mgmt1",
                "port_id_subtype": "link_local_addr",
            },
            "port_id": "aa:aa:aa:aa:aa:aa",
        },
    },
    "1%2F1%2F5": {
        "99:99:99:99:99:99,1/1/5": {
            "chassis_id": "99:99:99:99:99:99",
            "mac_addr": "99:99:99:99:99:99",
            "neighbor_info": {
                "chassis_description": "sw-leaf-bmc-99",
                "chassis_name": "sw-leaf-bmc99",
                "port_description": "1/1/5",
                "port_id_subtype": "if_name",
            },
            "port_id": "1/1/5",
        },
    },
}

arp_neighbors_json1 = {
    "192.168.1.2,vlan1": {
        "mac": "00:00:00:00:00:00",
        "ip_address": "192.168.1.2",
        "port": {"vlan1": "/rest/v10.04/system/interfaces/vlan1"},
    },
    "192.168.1.3,vlan2": {
        "mac": "11:11:11:11:11:11",
        "ip_address": "192.168.1.3",
        "port": {"vlan2": "/rest/v10.04/system/interfaces/vlan2"},
    },
    "192.168.2.2,vlan3": {
        "mac": "00:00:00:00:00:00",
        "ip_address": "192.168.2.2",
        "port": {"vlan3": "/rest/v10.04/system/interfaces/vlan3"},
    },
}

# Switch 2
switch_info2 = {
    "hostname": "sw-leaf02",
    "platform_name": "X86-64",
    "system_mac": "bb:bb:bb:bb:bb:bb",
}

lldp_neighbors_json2 = {
    "1%2F1%2F1": {
        "aa:aa:aa:aa:aa:aa,1/1/1": {
            "chassis_id": "aa:aa:aa:aa:aa:aa",
            "mac_addr": "aa:aa:aa:aa:aa:bb",
            "neighbor_info": {
                "chassis_description": "Test switch description",
                "chassis_name": "sw-leaf01",
                "port_description": "",
                "port_id_subtype": "if_name",
            },
            "port_id": "1/1/1",
        },
    },
}

arp_neighbors_json2 = {
    "192.168.1.2,vlan1": {
        "mac": "00:00:00:00:00:00",
        "ip_address": "192.168.1.2",
        "port": {"vlan1": "/rest/v10.04/system/interfaces/vlan1"},
    },
    "192.168.1.3,vlan2": {
        "mac": "11:11:11:11:11:11",
        "ip_address": "192.168.1.3",
        "port": {"vlan2": "/rest/v10.04/system/interfaces/vlan2"},
    },
    "192.168.2.2,vlan3": {
        "mac": "00:00:00:00:00:00",
        "ip_address": "192.168.2.2",
        "port": {"vlan3": "/rest/v10.04/system/interfaces/vlan3"},
    },
}


def generate_test_file(file_name):
    """Generate xlsx sheet for testing."""
    wb = Workbook()
    test_file = file_name
    ws1 = wb.active
    ws1.title = "25G_10G"
    ws1["I14"] = "Source"
    ws1["J14"] = "Rack"
    ws1["K14"] = "Location"
    ws1["L14"] = "Slot"
    # None
    ws1["N14"] = "Port"
    ws1["O14"] = "Destination"
    ws1["P14"] = "Rack"
    ws1["Q14"] = "Location"
    # None
    ws1["S14"] = "Port"

    test_data = [
        [
            "sw-25g01",
            "x3000",
            "u12",
            "",
            "-",
            "1",
            "sw-25g02",
            "x3000",
            "u13",
            "-",
            "1",
        ],
        [
            "sw-25g01",
            "x3000",
            "u12",
            "",
            "-",
            "2",
            "sw-25g02",
            "x3000",
            "u13",
            "-",
            "2",
        ],
        [
            "sw-25g01",
            "x3000",
            "u12",
            "",
            "-",
            "3",
            "junk",
            "x3000",
            "u13",
            "-",
            "2",
        ],
        [
            "sw-25g01",
            "x3000",
            "u12",
            "",
            "-",
            "4",
            "sw-smn99",
            "x3000",
            "u13",
            "-",
            "2",
        ],
        [
            "sw-25g01",
            "x3000",
            "u12",
            "",
            "-",
            "5",
            "junk",
            "x3000",
            "u13",
            "-",
            "2",
        ],
        [
            "sw-smn01",
            "x3000",
            "U14",
            "",
            "-",
            "49",
            "sw-25g01",
            "x3000",
            "u12",
            "-",
            "48",
        ],
        [
            "sw-smn01",
            "x3000",
            "U14",
            "",
            "",
            "50",
            "sw-25g02",
            "x3000",
            "u13",
            "-",
            "48",
        ],
        [
            "sw-25g01",
            "x3000",
            "u12",
            "",
            "-",
            "47",
            "sw-25g02",
            "x3000",
            "u13",
            "-",
            "47",
        ],
        [
            "uan01",
            "x3000",
            "u19",
            "ocp",
            "-",
            "1",
            "sw-25g01",
            "x3000",
            "u12",
            "-",
            "16",
        ],
        [
            "uan01",
            "x3000",
            "u19",
            "ocp",
            "-",
            "2",
            "sw-25g01",
            "x3000",
            "u12",
            "-",
            "17",
        ],
        [
            "uan01",
            "x3000",
            "u19",
            "pcie-slot1",
            "-",
            "1",
            "sw-25g02",
            "x3000",
            "u13",
            "-",
            "16",
        ],
        [
            "uan01",
            "x3000",
            "u19",
            "pcie-slot1",
            "-",
            "2",
            "sw-25g02",
            "x3000",
            "u13",
            "-",
            "17",
        ],
        [
            "sn03",
            "x3000",
            "u09",
            "ocp",
            "-",
            "1",
            "sw-25g01",
            "x3000",
            "u12",
            "-",
            "15",
        ],
        [
            "wn03",
            "x3000",
            "u06",
            "ocp",
            "-",
            "1",
            "sw-25g01",
            "x3000",
            "u12",
            "-",
            "9",
        ],
        [
            "wn03",
            "x3000",
            "u06",
            "ocp",
            "-",
            "2",
            "sw-25g02",
            "x3000",
            "u13",
            "-",
            "9",
        ],
        [
            "CAN switch",
            "cfcanb4s1",
            "",
            "",
            "-",
            "9",
            "sw-25g01",
            "x3000",
            "u12",
            "-",
            "36",
        ],
        # BAD ROW, do not include in normal run
        [
            "mn99",
            "x3000",
            "u12",
            "",
            "-",
            "1",
            "mn98",
            "x3000",
            "u13",
            "-",
            "1",
        ],
    ]
    for row in range(0, 17):
        for col in range(0, 11):
            ws1.cell(column=col + 9, row=row + 15, value=f"{test_data[row][col]}")

    # Tab 2 "Bad_Headers"
    ws2 = wb.create_sheet(title="Bad_Headers")
    ws2["I14"] = "Source"
    ws2["J14"] = "Rack"
    ws2["K14"] = "Location"
    # Missing Header
    # None
    ws2["M14"] = "Port"
    ws2["N14"] = "Destination"
    ws2["O14"] = "Rack"
    ws2["P14"] = "Location"
    # None
    ws2["R14"] = "Port"

    # Tab 3 "More_connections" containing bad connections
    ws3 = wb.create_sheet(title="More_connections")
    ws3["I14"] = "Source"
    ws3["J14"] = "Rack"
    ws3["K14"] = "Location"
    ws3["L14"] = "Slot"
    # None
    ws3["N14"] = "Port"
    ws3["O14"] = "Destination"
    ws3["P14"] = "Rack"
    ws3["Q14"] = "Location"
    # None
    ws3["S14"] = "Port"

    test_data3 = [
        [
            "sw-25g01",
            "x3000",
            "u12",
            "",
            "-",
            "51",
            "sw-25g02",
            "x3000",
            "u13",
            "-",
            "51",
        ],
        [
            "sw-25g01",
            "x3000",
            "u12",
            "",
            "-",
            "52",
            "sw-25g02",
            "x3000",
            "u13",
            "-",
            "52",
        ],
        [
            "sw-25g01",
            "x3000",
            "u12",
            "",
            "-",
            "52",
            "sw-25g02",
            "x3000",
            "u13",
            "-",
            "51",
        ],
        [
            "sw-cdu01",
            "x3000",
            "u12",
            "",
            "-",
            "52",
            "sw-smn99",
            "x3000",
            "u13",
            "-",
            "52",
        ],
        [
            "mn-99",
            "x3000",
            "u12",
            "",
            "-",
            "52",
            "sw-25g01",
            "x3000",
            "u13",
            "-",
            "52",
        ],
        [
            "mn-99",
            "x3000",
            "u12",
            "",
            "-",
            "50",
            "sw-25g02",
            "x3000",
            "u13",
            "-",
            "52",
        ],
        [
            "mn-99",
            "x3000",
            "u12",
            "",
            "-",
            "51",
            "sw-smn98",
            "x3000",
            "u13",
            "-",
            "52",
        ],
        [
            "mn-99",
            "x3000",
            "u12",
            "",
            "-",
            "52",
            "sw-smn99",
            "x3000",
            "u13",
            "-",
            "52",
        ],
        [
            "sw-100g01",
            "x3000",
            "u12",
            "",
            "-",
            "52",
            "sw-smn99",
            "x3000",
            "u13",
            "-",
            "52",
        ],
    ]
    for row in range(0, 9):
        for col in range(0, 11):
            ws3.cell(column=col + 9, row=row + 15, value=f"{test_data3[row][col]}")

    wb.save(filename=test_file)


mac_address_table = (
    "MAC age-time            : 300 seconds\n"
    + "Number of MAC addresses : 90\n"
    + "\n"
    + "MAC Address          VLAN     Type                      Port\n"
    + "--------------------------------------------------------------\n"
    + "00:40:a6:00:00:00    2        dynamic                   1/1/3\n"
)
