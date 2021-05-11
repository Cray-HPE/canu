"""Test CANU validate shcd commands."""

import click.testing
from openpyxl import Workbook

from canu.cli import cli


test_file = "test_file.xlsx"
architecture = "tds"
tabs = "25G_10G"
corners = "I14,S48"
shasta = "1.4"
cache_minutes = 0
runner = click.testing.CliRunner()


def test_validate_shcd():
    """Test that the `canu validate shcd` command runs and returns valid cabling."""
    with runner.isolated_filesystem():
        generate_test_file(test_file)
        result = runner.invoke(
            cli,
            [
                "--shasta",
                shasta,
                "--cache",
                cache_minutes,
                "validate",
                "shcd",
                "--architecture",
                architecture,
                "--shcd",
                test_file,
                "--tabs",
                tabs,
                "--corners",
                corners,
            ],
        )
        assert result.exit_code == 0
        assert "sw-spine-002 connects to 12 nodes:" in str(result.output)


def test_validate_shcd_full():
    """Test that the `canu validate shcd` command runs and returns valid cabling with '--architecture full' flag."""
    full_architecture = "full"
    with runner.isolated_filesystem():
        generate_test_file(test_file)
        result = runner.invoke(
            cli,
            [
                "--shasta",
                shasta,
                "--cache",
                cache_minutes,
                "validate",
                "shcd",
                "--architecture",
                full_architecture,
                "--shcd",
                test_file,
                "--tabs",
                tabs,
                "--corners",
                corners,
            ],
        )
        assert result.exit_code == 0
        assert "sw-leaf-002 connects to 12 nodes:" in str(result.output)


def test_validate_shcd_missing_file():
    """Test that the `canu validate shcd` command fails on missing file."""
    with runner.isolated_filesystem():

        result = runner.invoke(
            cli,
            [
                "--shasta",
                shasta,
                "--cache",
                cache_minutes,
                "validate",
                "shcd",
                "--architecture",
                architecture,
                "--tabs",
                tabs,
                "--corners",
                corners,
            ],
        )
        assert result.exit_code == 2
        assert "Error: Missing option '--shcd'." in str(result.output)


def test_validate_shcd_bad_file():
    """Test that the `canu validate shcd` command fails on bad file."""
    bad_file = "does_not_exist.xlsx"
    with runner.isolated_filesystem():

        result = runner.invoke(
            cli,
            [
                "--shasta",
                shasta,
                "--cache",
                cache_minutes,
                "validate",
                "shcd",
                "--architecture",
                architecture,
                "--shcd",
                bad_file,
                "--tabs",
                tabs,
                "--corners",
                corners,
            ],
        )
        assert result.exit_code == 2
        assert (
            "Error: Invalid value for '--shcd': Could not open file: does_not_exist.xlsx: No such file or directory"
            in str(result.output)
        )


def test_validate_shcd_missing_tabs():
    """Test that the `canu validate shcd` command fails on missing tabs."""
    with runner.isolated_filesystem():
        generate_test_file(test_file)
        result = runner.invoke(
            cli,
            [
                "--shasta",
                shasta,
                "--cache",
                cache_minutes,
                "validate",
                "shcd",
                "--architecture",
                architecture,
                "--shcd",
                test_file,
            ],
        )
        assert result.exit_code == 2
        assert "Error: Missing option '--tabs'." in str(result.output)


def test_validate_shcd_bad_tab():
    """Test that the `canu validate shcd` command fails on bad tab name."""
    bad_tab = "NMN"
    with runner.isolated_filesystem():
        generate_test_file(test_file)
        result = runner.invoke(
            cli,
            [
                "--shasta",
                shasta,
                "--cache",
                cache_minutes,
                "validate",
                "shcd",
                "--architecture",
                architecture,
                "--shcd",
                test_file,
                "--tabs",
                bad_tab,
                "--corners",
                corners,
            ],
        )
        assert result.exit_code == 1
        assert "Tab NMN not found in test_file.xlsx" in str(result.output)


def test_validate_shcd_corner_prompt():
    """Test that the `canu validate shcd` command prompts for corner input and runs."""
    with runner.isolated_filesystem():
        generate_test_file(test_file)
        result = runner.invoke(
            cli,
            [
                "--shasta",
                shasta,
                "--cache",
                cache_minutes,
                "validate",
                "shcd",
                "--architecture",
                architecture,
                "--shcd",
                test_file,
                "--tabs",
                tabs,
            ],
            input="I16\nS49",
        )
        assert result.exit_code == 0
        assert "sw-spine-002 connects to 12 nodes:" in str(result.output)


def test_validate_shcd_corners_too_narrow():
    """Test that the `canu validate shcd` command fails on too narrow area."""
    corners_too_narrow = "I16,R48"
    with runner.isolated_filesystem():
        generate_test_file(test_file)
        result = runner.invoke(
            cli,
            [
                "--shasta",
                shasta,
                "--cache",
                cache_minutes,
                "validate",
                "shcd",
                "--architecture",
                architecture,
                "--shcd",
                test_file,
                "--tabs",
                tabs,
                "--corners",
                corners_too_narrow,
            ],
        )
        assert result.exit_code == 1
        assert (
            "Ensure that the upper left corner (Labeled 'Source'), and the lower right corner of the table is entered."
            in str(result.output)
        )


def test_validate_shcd_corners_too_high():
    """Test that the `canu validate shcd` command fails on empty cells."""
    corners_too_high = "H16,S48"
    with runner.isolated_filesystem():
        generate_test_file(test_file)
        result = runner.invoke(
            cli,
            [
                "--shasta",
                shasta,
                "--cache",
                cache_minutes,
                "validate",
                "shcd",
                "--architecture",
                architecture,
                "--shcd",
                test_file,
                "--tabs",
                tabs,
                "--corners",
                corners_too_high,
            ],
        )
        assert result.exit_code == 1
        assert "Ensure the range entered does not contain a row of empty cells." in str(
            result.output
        )


def test_validate_shcd_corners_bad_cell():
    """Test that the `canu validate shcd` command fails on bad cell."""
    corners_bad_cell = "16,S48"
    with runner.isolated_filesystem():
        generate_test_file(test_file)
        result = runner.invoke(
            cli,
            [
                "--shasta",
                shasta,
                "--cache",
                cache_minutes,
                "validate",
                "shcd",
                "--architecture",
                architecture,
                "--shcd",
                test_file,
                "--tabs",
                tabs,
                "--corners",
                corners_bad_cell,
            ],
        )
        assert result.exit_code == 1
        assert "Bad range of cells entered for tab 25G_10G." in str(result.output)


def test_validate_shcd_not_enough_corners():
    """Test that the `canu validate shcd` command fails on not enough corners."""
    not_enough_corners = "H16"
    with runner.isolated_filesystem():
        generate_test_file(test_file)
        result = runner.invoke(
            cli,
            [
                "--shasta",
                shasta,
                "--cache",
                cache_minutes,
                "validate",
                "shcd",
                "--architecture",
                architecture,
                "--shcd",
                test_file,
                "--tabs",
                tabs,
                "--corners",
                not_enough_corners,
            ],
        )
        assert result.exit_code == 0
        assert "There were 1 corners entered, but there should be 2." in str(
            result.output
        )


def test_validate_shcd_bad_headers():
    """Test that the `canu validate shcd` command fails on bad headers."""
    bad_header_tab = "Bad_Headers"
    with runner.isolated_filesystem():
        generate_test_file(test_file)
        result = runner.invoke(
            cli,
            [
                "--shasta",
                shasta,
                "--cache",
                cache_minutes,
                "validate",
                "shcd",
                "--architecture",
                architecture,
                "--shcd",
                test_file,
                "--tabs",
                bad_header_tab,
                "--corners",
                corners,
            ],
        )
        assert result.exit_code == 1
        assert "On tab Bad_Headers, header column Slot not found" in str(result.output)


def test_validate_shcd_bad_architectural_definition():
    """Test that the `canu validate shcd` command fails with bad connections."""
    corners_bad_row = "I14,S50"
    with runner.isolated_filesystem():
        generate_test_file(test_file)
        result = runner.invoke(
            cli,
            [
                "--shasta",
                shasta,
                "--cache",
                cache_minutes,
                "validate",
                "shcd",
                "--architecture",
                architecture,
                "--shcd",
                test_file,
                "--tabs",
                tabs,
                "--corners",
                corners_bad_row,
            ],
        )
        assert result.exit_code == 1
        assert "No architectural definition found to allow connection between" in str(
            result.output
        )


def test_validate_shcd_multiple_connections():
    """Test that the `canu validate shcd` command runs and returns valid cabling."""
    multiple_connections_tab = "More_connections"
    multiple_connections_corners = "I14,S20"
    with runner.isolated_filesystem():
        generate_test_file(test_file)
        result = runner.invoke(
            cli,
            [
                "--shasta",
                shasta,
                "--cache",
                cache_minutes,
                "validate",
                "shcd",
                "--architecture",
                architecture,
                "--shcd",
                test_file,
                "--tabs",
                multiple_connections_tab,
                "--corners",
                multiple_connections_corners,
            ],
        )
        assert result.exit_code == 1
        assert "No architectural definition found to allow connection between" in str(
            result.output
        )


def generate_test_file(file_name):
    """Generate xlsx sheet for testing."""
    wb = Workbook()
    test_file = file_name
    ws1 = wb.active
    ws1.title = "25G_10G"
    ws1["I14"] = "Source"
    ws1["J14"] = "Rack"
    ws1["K14"] = "Location"
    ws1["L14"] = "Slot"
    # None
    ws1["N14"] = "Port"
    ws1["O14"] = "Destination"
    ws1["P14"] = "Rack"
    ws1["Q14"] = "Location"
    # None
    ws1["S14"] = "Port"

    test_data = [
        [
            "sw-25g01",
            "x3000",
            "u12",
            "",
            "-",
            "j51",
            "sw-25g02",
            "x3000",
            "u13",
            "-",
            "j51",
        ],
        [
            "sw-25g01",
            "x3000",
            "u12",
            "",
            "-",
            "j52",
            "sw-25g02",
            "x3000",
            "u13",
            "-",
            "j52",
        ],
        [
            "sw-smn01",
            "x3000",
            "U14",
            "",
            "-",
            "j49",
            "sw-25g01",
            "x3000",
            "u12",
            "-",
            "j48",
        ],
        [
            "sw-smn01",
            "x3000",
            "U14",
            "",
            "",
            "j50",
            "sw-25g02",
            "x3000",
            "u13",
            "-",
            "j48",
        ],
        [
            "sw-25g01",
            "x3000",
            "u12",
            "",
            "-",
            "j47",
            "sw-25g02",
            "x3000",
            "u13",
            "-",
            "j47",
        ],
        [
            "uan01",
            "x3000",
            "u19",
            "ocp",
            "-",
            "j1",
            "sw-25g01",
            "x3000",
            "u12",
            "-",
            "j16",
        ],
        [
            "uan01",
            "x3000",
            "u19",
            "ocp",
            "-",
            "j2",
            "sw-25g02",
            "x3000",
            "u12",
            "-",
            "j17",
        ],
        [
            "uan01",
            "x3000",
            "u19",
            "s1",
            "-",
            "j1",
            "sw-25g01",
            "x3000",
            "u13",
            "-",
            "j16",
        ],
        [
            "uan01",
            "x3000",
            "u19",
            "s1",
            "-",
            "j2",
            "sw-25g02",
            "x3000",
            "u13",
            "-",
            "j17",
        ],
        [
            "sn03",
            "x3000",
            "u09",
            "ocp",
            "-",
            "j1",
            "sw-25g01",
            "x3000",
            "u12",
            "-",
            "j15",
        ],
        [
            "sn03",
            "x3000",
            "u09",
            "ocp",
            "-",
            "j2	",
            "sw-25g02",
            "x3000",
            "u13",
            "-",
            "j15",
        ],
        [
            "sn03",
            "x3000",
            "u09",
            "s1	",
            "-",
            "j1",
            "sw-25g01",
            "x3000",
            "u12",
            "-",
            "j14",
        ],
        [
            "sn03",
            "x3000",
            "u09",
            "s1	",
            "-",
            "j2",
            "sw-25g02",
            "x3000",
            "u13",
            "-",
            "j14",
        ],
        [
            "sn02",
            "x3000",
            "u08",
            "ocp",
            "-",
            "j1",
            "sw-25g01",
            "x3000",
            "u12",
            "-",
            "j13",
        ],
        [
            "sn02",
            "x3000",
            "u08",
            "ocp",
            "-",
            "j2	",
            "sw-25g02",
            "x3000",
            "u13",
            "-",
            "j13",
        ],
        [
            "sn02",
            "x3000",
            "u08",
            "s1	",
            "-",
            "j1",
            "sw-25g01",
            "x3000",
            "u12",
            "-",
            "j12",
        ],
        [
            "sn02",
            "x3000",
            "u08",
            "s1	",
            "-",
            "j2	",
            "sw-25g02",
            "x3000",
            "u13",
            "-",
            "j12",
        ],
        [
            "sn01",
            "x3000",
            "u07",
            "ocp",
            "-",
            "j1",
            "sw-25g01",
            "x3000",
            "u12",
            "-",
            "j11",
        ],
        [
            "sn01",
            "x3000",
            "u07",
            "ocp",
            "-",
            "j2	",
            "sw-25g02",
            "x3000",
            "u13",
            "-",
            "j11",
        ],
        [
            "sn01",
            "x3000",
            "u07",
            "s1	",
            "-",
            "j1",
            "sw-25g01",
            "x3000",
            "u12",
            "-",
            "j10",
        ],
        [
            "sn01",
            "x3000",
            "u07",
            "s1	",
            "-",
            "j2	",
            "sw-25g02",
            "x3000",
            "u13",
            "-",
            "j10",
        ],
        [
            "wn03",
            "x3000",
            "u06",
            "ocp",
            "-",
            "j1",
            "sw-25g01",
            "x3000",
            "u12",
            "-",
            "j9",
        ],
        [
            "wn03",
            "x3000",
            "u06",
            "ocp",
            "-",
            "j2	",
            "sw-25g02",
            "x3000",
            "u13",
            "-",
            "j9",
        ],
        [
            "wn02",
            "x3000",
            "u05",
            "ocp",
            "-",
            "j1",
            "sw-25g01",
            "x3000",
            "u12",
            "-",
            "j8",
        ],
        [
            "wn02",
            "x3000",
            "u05",
            "ocp",
            "-",
            "j2	",
            "sw-25g02",
            "x3000",
            "u13",
            "-",
            "j8",
        ],
        [
            "wn01",
            "x3000",
            "u04",
            "ocp",
            "-",
            "j1",
            "sw-25g01",
            "x3000",
            "u12",
            "-",
            "j7",
        ],
        [
            "wn01",
            "x3000",
            "u04",
            "ocp",
            "-",
            "j2	",
            "sw-25g02",
            "x3000",
            "u13",
            "-",
            "j7",
        ],
        [
            "mn03",
            "x3000",
            "u03",
            "ocp",
            "-",
            "j1",
            "sw-25g01",
            "x3000",
            "u12",
            "-",
            "j5",
        ],
        [
            "mn03",
            "x3000",
            "u03",
            "s1	",
            "-",
            "j1	",
            "sw-25g02",
            "x3000",
            "u13",
            "-",
            "j5",
        ],
        [
            "mn02",
            "x3000",
            "u02",
            "ocp",
            "-",
            "j1",
            "sw-25g01",
            "x3000",
            "u12",
            "-",
            "j3",
        ],
        [
            "mn02",
            "x3000",
            "u02",
            "s1	",
            "-",
            "j1	",
            "sw-25g02",
            "x3000",
            "u13",
            "-",
            "j3",
        ],
        [
            "mn01",
            "x3000",
            "u01",
            "ocp",
            "-",
            "j1",
            "sw-25g01",
            "x3000",
            "u12",
            "-",
            "j1",
        ],
        [
            "mn01",
            "x3000",
            "u01",
            "s1	",
            "-",
            "j1	",
            "sw-25g02",
            "x3000",
            "u13",
            "-",
            "j1",
        ],
        [
            "CAN switch",
            "cfcanb4s1",
            "",
            "",
            "-",
            "port 9",
            "sw-25g01",
            "x3000",
            "u12",
            "-",
            "j36",
        ],
        [
            "CAN switch",
            "cfcanb4s1",
            "",
            "",
            "-",
            "port 11",
            "sw-25g02",
            "x3000",
            "u13",
            "-",
            "j36",
        ],
        # BAD ROW, do not include in normal run
        [
            "mn99",
            "x3000",
            "u12",
            "",
            "-",
            "j1",
            "mn98",
            "x3000",
            "u13",
            "-",
            "j1",
        ],
    ]
    for row in range(0, 36):
        for col in range(0, 11):
            ws1.cell(column=col + 9, row=row + 15, value=f"{test_data[row][col]}")

    # Tab 2 "Bad_Headers"
    ws2 = wb.create_sheet(title="Bad_Headers")
    ws2["I14"] = "Source"
    ws2["J14"] = "Rack"
    ws2["K14"] = "Location"
    # ws2["L14"] = "Slot" # Missing Header
    # None
    ws2["M14"] = "Port"
    ws2["N14"] = "Destination"
    ws2["O14"] = "Rack"
    ws2["P14"] = "Location"
    # None
    ws2["R14"] = "Port"

    # Tab 3 "More_connections" containing bad connections
    ws3 = wb.create_sheet(title="More_connections")
    ws3["I14"] = "Source"
    ws3["J14"] = "Rack"
    ws3["K14"] = "Location"
    ws3["L14"] = "Slot"
    # None
    ws3["N14"] = "Port"
    ws3["O14"] = "Destination"
    ws3["P14"] = "Rack"
    ws3["Q14"] = "Location"
    # None
    ws3["S14"] = "Port"

    test_data3 = [
        [
            "sw-25g01",
            "x3000",
            "u12",
            "",
            "-",
            "j51",
            "sw-25g02",
            "x3000",
            "u13",
            "-",
            "j51",
        ],
        [
            "sw-25g01",
            "x3000",
            "u12",
            "",
            "-",
            "j52",
            "sw-25g02",
            "x3000",
            "u13",
            "-",
            "j52",
        ],
        [
            "sw-25g01",
            "x3000",
            "u12",
            "",
            "-",
            "j52",
            "sw-25g02",
            "x3000",
            "u13",
            "-",
            "j51",
        ],
        [
            "sw-cdu01",
            "x3000",
            "u12",
            "",
            "-",
            "j52",
            "sw-smn99",
            "x3000",
            "u13",
            "-",
            "j52",
        ],
        [
            "mn-99",
            "x3000",
            "u12",
            "",
            "-",
            "j52",
            "sw-25g01",
            "x3000",
            "u13",
            "-",
            "j52",
        ],
        [
            "mn-99",
            "x3000",
            "u12",
            "",
            "-",
            "j50",
            "sw-25g02",
            "x3000",
            "u13",
            "-",
            "j52",
        ],
        [
            "mn-99",
            "x3000",
            "u12",
            "",
            "-",
            "j51",
            "sw-smn98",
            "x3000",
            "u13",
            "-",
            "j52",
        ],
        [
            "mn-99",
            "x3000",
            "u12",
            "",
            "-",
            "j52",
            "sw-smn99",
            "x3000",
            "u13",
            "-",
            "j52",
        ],
        [
            "sw-100g01",
            "x3000",
            "u12",
            "",
            "-",
            "j52",
            "sw-smn99",
            "x3000",
            "u13",
            "-",
            "j52",
        ],
    ]
    for row in range(0, 9):
        for col in range(0, 11):
            ws3.cell(column=col + 9, row=row + 15, value=f"{test_data3[row][col]}")

    wb.save(filename=test_file)
